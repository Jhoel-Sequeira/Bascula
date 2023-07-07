

from datetime import datetime, date, timedelta
import json
import ast
import random
from flask import Flask, jsonify, redirect, render_template, request, send_file, session, url_for
from flask_mysqldb import MySQL
import MySQLdb.cursors
from werkzeug.security import check_password_hash, generate_password_hash
import config
import conexion
import jinja2
from jinja2 import Environment
import MySQLdb.cursors
import time
import uuid
import logging

#IMPORTS PARA GENERAR EL EXCEL
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment,Font,Border,Side


app = Flask(__name__)

# Configuración del registro
handler = logging.FileHandler('app.log')
handler.setLevel(logging.DEBUG)
formatter = logging.Formatter('%(asctime)s %(levelname)s: %(message)s')
handler.setFormatter(formatter)
app.logger.addHandler(handler)

mysql = MySQL()
mysql.init_app(app)

# def generate_session_id():
#     return str(uuid.uuid4())

# app.config['SESSION_TYPE'] = 'filesystem'
# app.config['SESSION_PERMANENT'] = False
# app.config['SESSION_USE_SIGNER'] = True
# app.config['SESSION_COOKIE_NAME'] = 'flask_session_' + generate_session_id()


# CONFIGURACIONES DE LA BASE DE DATOS
app.secret_key = 'your secret key'
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = ''
app.config['MYSQL_DB'] = 'verificacion'
# CONFIGURAR TIEMPO DE LA SESION
#app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(seconds=10)

# CIERRE CONFIGURACIONES
# @app.errorhandler(Exception)
# def handle_error(error):
#     app.logger.error('Error en la aplicación:', exc_info=error)
#     return 'Se ha producido un error en la aplicación', 500

def capturarHora():
    hi = datetime.now()
    return hi


# cursor = mysql.connection.cursor(MySQLdb.cursors.DictCursor) para iniciar
# mysql.connection.commit() Para finalizar la conexion a la base de datos


# cursor.execute("") para ejecutar las instrucciones sql
#PONEMOS LA VALIDACION DE LA CADUCACION DE SESION
def ValidarSesion():
    if not session['user']:
    
        return redirect(url_for('Index'))

@app.route('/')
def Index():
    session.clear()
    
    

    return render_template('login.html')


# Aqui comienza el login

@app.route('/login', methods=["GET", "POST"])
def login():

    if request.method == "POST":
        session_id = request.cookies.get(app.config['SESSION_COOKIE_NAME'])
        usuario = request.form['loginUser']
        Contraseña = request.form['loginPassword']
        if usuario == "" or Contraseña == "":
            return render_template('login.html', errorlogin=1)
        else:
            # MANDAMOS A VERIFICAR LOS DATOS EN ODDO
            # conexion.username = usuario
            # conexion.password = Contraseña
            try:
                print(usuario)
                print('CONTRASEÑA: '+Contraseña)
                uidUser = conexion.obtenerUid(usuario, Contraseña)
                print("uidUser:", uidUser)
                uidUser = conexion.obtenerUid(usuario, Contraseña)
                print("uidUser:", uidUser)
                uid = conexion.Autenticar(usuario, Contraseña, uidUser)
                session['uid'] = uid[0]['id']
                print("verrr")
                print(uid)
                cargo = uid[0]['x_studio_field_xql4c']
                print(cargo[1])
                # session["puntoid"] = uid[0]['almacen'][0]
                # session["punto"] = uid[0]['almacen'][1]

                if cargo[1] == "JEFE DE TECNOLOGÍA" or cargo[1] == "SOPORTE DE INFORMATICA" or cargo[1] == 'SOPORTE DE INFORMA' or cargo[1] == "GERENTE ADMINISTRACIÓN":

                    session["user"] = usuario
                    session["pass"] = Contraseña
                    session["userrole"] = 1
                    session["puntoid"] = 5
                    session["punto"] = "Sin Punto"

                    # session["puntoid"] = uid[0]['almacen'][0]
                    # session["punto"] = uid[0]['almacen'][1]

                    # ESTA CONSULTA ERA PARA SABER EL CARGO DEL USUARIO LOGUEADO
                    # cur = mysql.connection.cursor()
                    # cur.execute("select IdCargo from tb_usuarios Where Id_Usuario = %s",[session["userId"]])
                    # cargo = cur.fetchone()
                    # print(cargo[0]) cargo de oddo
                    session['cargo'] = 3
                    # GUARDAR USUARIOS EN LA BASE DE DATOS

                    cur = mysql.connection.cursor()
                    cur.execute(
                        "SELECT * FROM tb_credenciales Where Usuarios = %s", [session["user"]])
                    usuarioExistente = cur.fetchone()
                    mysql.connection.commit()

                    user_info = conexion.TraerUsuario(
                        str(uid[0]['id']), session['uid'], session['pass'])
                    print("user info:", user_info)
                    print("usuario Existe: ", usuarioExistente)
                    if not usuarioExistente:
                        cur = mysql.connection.cursor()
                        cur.execute("INSERT INTO tb_credenciales (Usuarios,Contraseñas,IdRol) VALUES (%s,%s,%s)", (
                            session["user"], generate_password_hash(Contraseña), session["userrole"]))

                        mysql.connection.commit()
                        idcredenciales = cur.lastrowid
                        # YA TENEMOS LAS CREDENCIALES DE LOS USUARIOS DE ODDO
                        print(idcredenciales)
                        # INSERTAMOS EN LA TABLA EL USUARIO COMO TAL
                        cur = mysql.connection.cursor()
                        cur.execute("INSERT INTO tb_usuarios (NombreUsuario,IdCargo,IdOddo,IdCredenciales,IdEstado,IdPuesto) VALUES (%s,%s,%s,%s,%s,'5')", (
                            user_info[0]['name'], session['cargo'], user_info[0]['id'], idcredenciales, 1))
                        mysql.connection.commit()
                        ultimoUsuario = cur.lastrowid

                        id = usuarioExistente[0]
                        cur = mysql.connection.cursor()
                        cur.execute(
                            "SELECT u.Id_Usuario FROM tb_usuarios as u inner join tb_credenciales as cred on u.IdCredenciales = cred.Id_credenciales Where cred.Id_credenciales = %s", [id])
                        ultimoUSuarios = cur.fetchone()
                        mysql.connection.commit()
                        print(ultimoUSuarios[0])
                        session["userId"] = ultimoUSuarios[0]

                    elif usuarioExistente:
                        print(usuarioExistente[0])
                        id = usuarioExistente[0]
                        cur = mysql.connection.cursor()
                        cur.execute(
                            "SELECT u.Id_Usuario FROM tb_usuarios as u inner join tb_credenciales as cred on u.IdCredenciales = cred.Id_credenciales Where cred.Id_credenciales = %s", [id])
                        ultimoUSuarios = cur.fetchone()
                        mysql.connection.commit()
                        print(ultimoUSuarios[0])
                        session["userId"] = ultimoUSuarios[0]

                    print(user_info[0]['name'])
                    ########################################
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "select * from tb_proveedor Where IdEstado = 1")
                    Proveedores = cur.fetchall()
                    # CONSULTA PARA LOS PUNTOS DE COMPRA
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "select * from tb_puntocompra Where IdEstado = 1")
                    punto = cur.fetchall()
                    # CONSULTA PARA LOS MATERIALES
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "select * from tb_material Where Id_Estado = 1")
                    material = cur.fetchall()
                    # CONSULTA PARA LOS VERIFICADORES
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 2")
                    verificador = cur.fetchall()
                    # CONSULTA PARA LOS DIGITADOR
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 1")
                    digitador = cur.fetchall()
                    return render_template('ajustes.html', Proveedores=Proveedores, Punto=punto, Material=material, Verificador=verificador, Digitador=digitador)
                elif cargo[1] == "DIGITADOR":
                    # ESTE ES DIGITADOR
                    session["puntoid"] = uid[0]['almacen'][0]
                    session["punto"] = uid[0]['almacen'][1]

                    # SELECCIONAMOS LOS PUNTOS QUE TENEMOS EN LA BASE DE DATOS PARA
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "SELECT * FROM tb_puntocompra Where NombrePuntoCompra = %s", [session["punto"]])
                    puntoexiste = cur.fetchall()
                    mysql.connection.commit()

                    if not puntoexiste:
                        cur = mysql.connection.cursor()
                        cur.execute("INSERT INTO tb_puntocompra (Id_PuntoCompra,NombrePuntoCompra,IdEstado) VALUES (%s,%s,%s)", (
                            session["puntoid"], session["punto"], '1'))
                        mysql.connection.commit()

                    print('punto: ', session['punto'])
                    session["pass"] = Contraseña
                    session["user"] = usuario
                    session["userrole"] = 2
                    # ESTA CONSULTA ERA PARA SABER EL CARGO DEL USUARIO LOGUEADO
                    # cur = mysql.connection.cursor()
                    # cur.execute("select IdCargo from tb_usuarios Where Id_Usuario = %s",[session["userId"]])
                    # cargo = cur.fetchone()
                    # print(cargo[0])
                    # ESCRIBIMOS EL CARGO DE DIGITADOR
                    session['cargo'] = 1

                    # vaidamos si el usuario ya esta logueado antes
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "SELECT * FROM tb_credenciales Where Usuarios = %s", [session["user"]])
                    usuarioExistente = cur.fetchall()
                    mysql.connection.commit()

                    user_info = conexion.TraerUsuario(
                        str(uid[0]['id']), session['uid'], session['pass'])
                    print("user info:", user_info)
                    if not usuarioExistente:
                        cur = mysql.connection.cursor()
                        cur.execute("INSERT INTO tb_credenciales (Usuarios,Contraseñas,IdRol) VALUES (%s,%s,%s)", (
                            session["user"], generate_password_hash(Contraseña), session["userrole"]))

                        mysql.connection.commit()
                        idcredenciales = cur.lastrowid
                        # YA TENEMOS LAS CREDENCIALES DE LOS USUARIOS DE ODDO
                        print(idcredenciales)
                        # INSERTAMOS EN LA TABLA EL USUARIO COMO TAL
                        cur = mysql.connection.cursor()
                        cur.execute("INSERT INTO tb_usuarios (NombreUsuario,IdCargo,IdOddo,IdCredenciales,IdEstado,IdPuesto) VALUES (%s,%s,%s,%s,%s,%s)", (
                            user_info[0]['name'], session['cargo'], user_info[0]['id'], idcredenciales, 1,session["puntoid"]))
                        mysql.connection.commit()
                        ultimoUsuario = cur.lastrowid

                        id = idcredenciales
                        cur = mysql.connection.cursor()
                        cur.execute(
                            "SELECT u.Id_Usuario FROM tb_usuarios as u inner join tb_credenciales as cred on u.IdCredenciales = cred.Id_credenciales Where cred.Id_credenciales = %s", [id])
                        ultimoUSuarios = cur.fetchone()
                        mysql.connection.commit()
                        print(ultimoUSuarios[0])
                        session["userId"] = ultimoUSuarios[0]
                    elif usuarioExistente:
                        print("aqui")
                        print(usuarioExistente[0][0])
                        id = usuarioExistente[0][0]
                        print("ssssssssssss")
                        cur = mysql.connection.cursor()
                        cur.execute(
                            "SELECT u.Id_Usuario FROM tb_usuarios as u inner join tb_credenciales as cred on u.IdCredenciales = cred.Id_credenciales Where cred.Id_credenciales = %s", [id])
                        ultimoUSuarios = cur.fetchone()
                        mysql.connection.commit()
                        session["userId"] = ultimoUSuarios[0]

                    cur = mysql.connection.cursor()
                    cur.execute(
                        "select * from tb_proveedor Where IdEstado = 1")
                    Proveedores = cur.fetchall()
                    # CONSULTA PARA LOS PUNTOS DE COMPRA
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "select * from tb_puntocompra Where IdEstado = 1")
                    punto = cur.fetchall()
                    # CONSULTA PARA LOS MATERIALES
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "select * from tb_material Where Id_Estado = 1")
                    material = cur.fetchall()
                    # CONSULTA PARA LOS VERIFICADORES
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 2")
                    verificador = cur.fetchall()
                    # CONSULTA PARA LOS DIGITADOR
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 1")
                    digitador = cur.fetchall()
                    return render_template('ajustes.html', Proveedores=Proveedores, Punto=punto, Material=material, Verificador=verificador, Digitador=digitador)

                elif cargo[1] == "VALIDADORA DE DATOS" or cargo[1] == "ANALISTA DE DATOS Y SISTEMAS":
                    # ESTE ES VALIDADOR
                    session["pass"] = Contraseña
                    session["user"] = usuario
                    session["userrole"] = 2
                    session["puntoid"] = 1
                    session["punto"] = "GRANEL/PLANTA: Receipts"
                    # ESTA CONSULTA ERA PARA SABER EL CARGO DEL USUARIO LOGUEADO
                    # cur = mysql.connection.cursor()
                    # cur.execute("select IdCargo from tb_usuarios Where Id_Usuario = %s",[session["userId"]])
                    # cargo = cur.fetchone()
                    # print(cargo[0])
                    # ESCRIBIMOS EL CARGO DE DIGITADOR
                    session['cargo'] = 5

                    # vaidamos si el usuario ya esta logueado antes
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "SELECT * FROM tb_credenciales Where Usuarios = %s", [session["user"]])
                    usuarioExistente = cur.fetchone()
                    mysql.connection.commit()

                    user_info = conexion.TraerUsuario(
                        str(uid[0]['id']), session['uid'], session['pass'])
                    print("user info:", user_info)
                    if not usuarioExistente:
                        cur = mysql.connection.cursor()
                        cur.execute("INSERT INTO tb_credenciales (Usuarios,Contraseñas,IdRol) VALUES (%s,%s,%s)", (
                            session["user"], generate_password_hash(Contraseña), session["userrole"]))

                        mysql.connection.commit()
                        idcredenciales = cur.lastrowid
                        # YA TENEMOS LAS CREDENCIALES DE LOS USUARIOS DE ODDO
                        print(idcredenciales)
                        # INSERTAMOS EN LA TABLA EL USUARIO COMO TAL
                        cur = mysql.connection.cursor()
                        cur.execute("INSERT INTO tb_usuarios (NombreUsuario,IdCargo,IdOddo,IdCredenciales,IdEstado,IdPuesto) VALUES (%s,%s,%s,%s,%s,'1')", (
                            user_info[0]['name'], session['cargo'], user_info[0]['id'], idcredenciales, 1))
                        mysql.connection.commit()
                        ultimoUsuario = cur.lastrowid

                        id = idcredenciales
                        cur = mysql.connection.cursor()
                        cur.execute(
                            "SELECT u.Id_Usuario FROM tb_usuarios as u inner join tb_credenciales as cred on u.IdCredenciales = cred.Id_credenciales Where cred.Id_credenciales = %s", [id])
                        ultimoUSuarios = cur.fetchone()
                        mysql.connection.commit()
                        print(ultimoUSuarios[0])
                        session["userId"] = ultimoUSuarios[0]

                    elif usuarioExistente:
                        print(usuarioExistente[0])
                        id = usuarioExistente[0]
                        print("ssssssssssss")
                        cur = mysql.connection.cursor()
                        cur.execute(
                            "SELECT u.Id_Usuario FROM tb_usuarios as u inner join tb_credenciales as cred on u.IdCredenciales = cred.Id_credenciales Where cred.Id_credenciales = %s", [id])
                        ultimoUSuarios = cur.fetchone()
                        mysql.connection.commit()
                        session["userId"] = ultimoUSuarios[0]

                    # session["userId"] = ultimoUsuario

                    cur = mysql.connection.cursor()
                    cur.execute(
                        "select * from tb_proveedor Where IdEstado = 1")
                    Proveedores = cur.fetchall()
                    # CONSULTA PARA LOS PUNTOS DE COMPRA
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "select * from tb_puntocompra Where IdEstado = 1")
                    punto = cur.fetchall()
                    # CONSULTA PARA LOS MATERIALES
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "select * from tb_material Where Id_Estado = 1")
                    material = cur.fetchall()
                    # CONSULTA PARA LOS VERIFICADORES
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 2")
                    verificador = cur.fetchall()
                    # CONSULTA PARA LOS DIGITADOR
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 1")
                    digitador = cur.fetchall()
                    return render_template('ajustes.html', Proveedores=Proveedores, Punto=punto, Material=material, Verificador=verificador, Digitador=digitador)

                    # else:
                    #     #VIENE VACIO
                    #     return render_template('login.html', errorlogin=1)
                elif cargo[1] == "SUPERVISOR PET" or cargo[1] == "RESPONSABLE DE VERIFICACION" or cargo[1] == "JEFE DE RECEPCION":
                    # CARGO DE LOS NUEVOS USUARIOS
                     # ESTE ES FISCALES Y JEFES DE PLANTA
                    session["pass"] = Contraseña
                    session["user"] = usuario
                    session["userrole"] = 2
                    session["puntoid"] = 1
                    session["punto"] = "GRANEL/PLANTA: Receipts"
                    # ESTA CONSULTA ERA PARA SABER EL CARGO DEL USUARIO LOGUEADO
                    # cur = mysql.connection.cursor()
                    # cur.execute("select IdCargo from tb_usuarios Where Id_Usuario = %s",[session["userId"]])
                    # cargo = cur.fetchone()
                    # print(cargo[0])
                    # ESCRIBIMOS EL CARGO DE JEFE
                    session['cargo'] = 6

                    # vaidamos si el usuario ya esta logueado antes
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "SELECT * FROM tb_credenciales Where Usuarios = %s", [session["user"]])
                    usuarioExistente = cur.fetchone()
                    mysql.connection.commit()

                    user_info = conexion.TraerUsuario(
                        str(uid[0]['id']), session['uid'], session['pass'])
                    print("user info:", user_info)
                    if not usuarioExistente:
                        cur = mysql.connection.cursor()
                        cur.execute("INSERT INTO tb_credenciales (Usuarios,Contraseñas,IdRol) VALUES (%s,%s,%s)", (
                            session["user"], generate_password_hash(Contraseña), session["userrole"]))

                        mysql.connection.commit()
                        idcredenciales = cur.lastrowid
                        # YA TENEMOS LAS CREDENCIALES DE LOS USUARIOS DE ODDO
                        print(idcredenciales)
                        # INSERTAMOS EN LA TABLA EL USUARIO COMO TAL
                        cur = mysql.connection.cursor()
                        cur.execute("INSERT INTO tb_usuarios (NombreUsuario,IdCargo,IdOddo,IdCredenciales,IdEstado,IdPuesto) VALUES (%s,%s,%s,%s,%s,'1')", (
                            user_info[0]['name'], session['cargo'], user_info[0]['id'], idcredenciales, 1))
                        mysql.connection.commit()
                        ultimoUsuario = cur.lastrowid

                        id = idcredenciales
                        cur = mysql.connection.cursor()
                        cur.execute(
                            "SELECT u.Id_Usuario FROM tb_usuarios as u inner join tb_credenciales as cred on u.IdCredenciales = cred.Id_credenciales Where cred.Id_credenciales = %s", [id])
                        ultimoUSuarios = cur.fetchone()
                        mysql.connection.commit()
                        print(ultimoUSuarios[0])
                        session["userId"] = ultimoUSuarios[0]

                    elif usuarioExistente:
                        print(usuarioExistente[0])
                        id = usuarioExistente[0]
                        print("ssssssssssss")
                        cur = mysql.connection.cursor()
                        cur.execute(
                            "SELECT u.Id_Usuario FROM tb_usuarios as u inner join tb_credenciales as cred on u.IdCredenciales = cred.Id_credenciales Where cred.Id_credenciales = %s", [id])
                        ultimoUSuarios = cur.fetchone()
                        mysql.connection.commit()
                        session["userId"] = ultimoUSuarios[0]

                    # session["userId"] = ultimoUsuario

                    cur = mysql.connection.cursor()
                    cur.execute(
                        "select * from tb_proveedor Where IdEstado = 1")
                    Proveedores = cur.fetchall()
                    # CONSULTA PARA LOS PUNTOS DE COMPRA
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "select * from tb_puntocompra Where IdEstado = 1")
                    punto = cur.fetchall()
                    # CONSULTA PARA LOS MATERIALES
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "select * from tb_material Where Id_Estado = 1")
                    material = cur.fetchall()
                    # CONSULTA PARA LOS VERIFICADORES
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 2")
                    verificador = cur.fetchall()
                    # CONSULTA PARA LOS DIGITADOR
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 1")
                    digitador = cur.fetchall()
                    return render_template('ajustes.html', Proveedores=Proveedores, Punto=punto, Material=material, Verificador=verificador, Digitador=digitador)

            except:
                try:

                    # ESTE ES VERIFICADOR
                    print("else")
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "select * from tb_credenciales Where Usuarios = %s", [usuario])
                    results = cur.fetchone()
                    print("aaaa")
                    print(results)
                    
                    if results:
                        # si trae algo
                        print("SI TRAE")
                        # Recordar el usuario y rol que se logeo
                        session["userId"] = results[0]
                        session["user"] = results[1]
                        session["userrole"] = results[3]
                        cur = mysql.connection.cursor()
                        cur.execute("select IdCargo from tb_usuarios Where Id_Usuario = %s", [
                                    session["userId"]])
                        cargo = cur.fetchone()
                        print(cargo[0])
                        session['cargo'] = cargo[0]
                        # ESTAS CONSULTAS SON PARA TRAER LOS PROVEEDORES Y LOS DATOS DE LOS SELECT
                        # CONSULTA PARA LOS PROVEEDORES
                        cur = mysql.connection.cursor()
                        cur.execute(
                            "select * from tb_proveedor Where IdEstado = 1")
                        Proveedores = cur.fetchall()
                        # CONSULTA PARA LOS PUNTOS DE COMPRA
                        cur = mysql.connection.cursor()
                        cur.execute(
                            "select * from tb_puntocompra Where IdEstado = 1")
                        punto = cur.fetchall()
                        # CONSULTA PARA LOS MATERIALES
                        cur = mysql.connection.cursor()
                        cur.execute(
                            "select * from tb_material Where Id_Estado = 1")
                        material = cur.fetchall()
                        # CONSULTA PARA LOS VERIFICADORES
                        cur = mysql.connection.cursor()
                        cur.execute(
                            "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 2")
                        verificador = cur.fetchall()
                        # CONSULTA PARA LOS DIGITADOR
                        cur = mysql.connection.cursor()
                        cur.execute(
                            "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 1")
                        digitador = cur.fetchall()

                        cur = mysql.connection.cursor()
                        cur.execute("SELECT pu.Id_PuntoCompra,pu.NombrePuntoCompra FROM `tb_usuarios` as u inner join tb_puntocompra as pu on u.IdPuesto = pu.Id_PuntoCompra where u.Id_Usuario = %s", [
                                    session["userId"]])
                        control = cur.fetchone()
                        mysql.connection.commit()
                        session["puntoid"] = control[0]
                        session["punto"] = control[1]

                        # CHEQUEAMOS LAS CONTRASEÑAS PARA VER SI SON IGUALES
                        if not check_password_hash(results[2], Contraseña):
                            # contraseñas incorrectas
                            print("contras")
                            return render_template('login.html', errorlogin=2)
                        return render_template('ajustes.html', Proveedores=Proveedores, Punto=punto, Material=material, Verificador=verificador, Digitador=digitador)

                    # else:
                    #     #VIENE VACIO
                    #     return render_template('login.html', errorlogin=1)
                except:

                    return render_template('login.html', errorlogin=1)

        return render_template('login.html', errorlogin=1)

@app.route('/desbloqueo')
def desbloqueo():
    # try:
        if session['userId']:
            cur = mysql.connection.cursor()
            cur.execute(
                "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 2")
            verificador = cur.fetchall()
            # CONSULTA PARA LOS DIGITADOR
            cur = mysql.connection.cursor()
            cur.execute(
                "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 1")
            digitador = cur.fetchall()

            return render_template('desbloqueo.html',verificadores = verificador,digitadores = digitador)
        else:
            return render_template('otros/error.html')

    # except:
    #     return render_template('otros/error.html')






@app.route('/home')
def home():
    try:
        if session['userId']:
            # ESTAS CONSULTAS SON PARA TRAER LOS PROVEEDORES Y LOS DATOS DE LOS SELECT
            # CONSULTA PARA LOS PROVEEDORES
            cur = mysql.connection.cursor()
            cur.execute("select * from tb_proveedor Where IdEstado = 1")
            Proveedores = cur.fetchall()
            # CONSULTA PARA LOS PUNTOS DE COMPRA
            cur = mysql.connection.cursor()

            cur.execute("select * from tb_puntocompra Where IdEstado = 1")
            punto = cur.fetchall()
            # CONSULTA PARA LOS MATERIALES
            cur = mysql.connection.cursor()
            cur.execute("select * from tb_material Where Id_Estado = 1")
            material = cur.fetchall()
            # CONSULTA PARA LOS VERIFICADORES
            cur = mysql.connection.cursor()
            cur.execute(
                "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 2")
            verificador = cur.fetchall()
            # CONSULTA PARA LOS DIGITADOR
            cur = mysql.connection.cursor()
            cur.execute(
                "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 1")
            digitador = cur.fetchall()
            print(digitador)

            return render_template('home.html', punto=session["punto"], cargo=session['cargo'], Proveedores=Proveedores, Punto=punto, Material=material, Verificador=verificador, Digitador=digitador)
        else:
            return render_template('otros/error.html')

    except:
        return render_template('otros/error.html')


@app.route('/buscarProveedor', methods=["POST", "GET"])
def buscarProveedor():
    if request.method == "POST":
        proveedor = request.form['proveedor']
        cur = mysql.connection.cursor()
        cur.execute(
            "select * from tb_proveedor Where IdEstado = 1 and NombreProveedor like %s", [proveedor+'%'])
        proveedores = cur.fetchall()
        print(proveedores)
        if proveedores:

            return render_template('otros/proveedor-busqueda.html', proveedores=proveedores)
        else:
            return "no"
    else:
        return "No"


@app.route('/buscarProveedorApi', methods=["POST", "GET"])
def buscarProveedorApi():
    if request.method == "POST":
        proveedor = request.form['proveedor']
        proveedores = conexion.buscarProveedor(proveedor, session['cargo'])
        # print(proveedores)
        if proveedores:

            return render_template('otros/proveedor-busqueda.html', proved=proveedores)
        else:
            return "no"
    else:
        return "No"
    
@app.route('/buscarCuadrillaApi', methods=["POST", "GET"])
def buscarCuadrillaApi():
    if request.method == "POST":
        cuadrilla = request.form['cuadrilla']
        cuadrilla = conexion.buscarCuadrilla(cuadrilla, session['cargo'])
        # print(proveedores)
        if cuadrilla:

            return render_template('otros/cuadrilla-busqueda.html', proved=cuadrilla)
        else:
            return "no"
    else:
        return "No"


@app.route('/buscarProveedorAdmin', methods=["POST", "GET"])
def buscarProveedorAdmin():
    if request.method == "POST":
        proveedor = request.form['proveedor']
        print(proveedor)
        cur = mysql.connection.cursor()
        cur.execute("SELECT p.Id_Proveedor,p.NombreProveedor,p.Cedula,e.NombreEstado FROM tb_proveedor as p inner join tb_estado as e ON p.IdEstado = e.Id_Estado where p.NombreProveedor like %s", [
                    proveedor+'%'])
        proveedores = cur.fetchall()
        if proveedores:

            return render_template('tablas/tabla-proveedornuevo.html', prov=proveedores)
        else:
            return "no"
    else:
        return "No"

# BUSCAR LOS USUARIOS PARA VER SI EXISTEN O NO EXISTEN


@app.route('/buscarUsuariosAdmin', methods=["POST", "GET"])
def buscarUsuariosAdmin():
    if request.method == "POST":
        cedula = request.form['usuario']
        cur = mysql.connection.cursor()
        cur.execute(
            "SELECT u.Id_Usuario,u.NombreUsuario as Nombre,c.Usuarios as Usuario,r.NombreRol,e.NombreEstado,u.Cedula,car.NombreCargo FROM tb_usuarios as u inner join tb_credenciales as c ON u.IdCredenciales = c.Id_Credenciales inner join tb_roles as r on c.IdRol = r.Id_Rol  inner join tb_estado as e on u.IdEstado = e.Id_Estado inner join tb_cargo as car on u.IdCargo = car.Id_Cargo where u.NombreUsuario like %s", [cedula+'%'])
        proveedores = cur.fetchall()
        if proveedores:

            return render_template('tablas/tabla-usuario.html', prov=proveedores)
        else:
            return "no"
    else:
        return "No"

# ACTUALIZAMOS EL PROVEEDOR


@app.route('/llamarProveedorEspecifico', methods=["POST", "GET"])
def llamarProveedorEspecifico():
    if request.method == "POST":
        id = request.form['id']
        print(id)
        cur = mysql.connection.cursor()
        cur.execute(
            "SELECT p.Id_Proveedor,p.NombreProveedor,p.Cedula,e.NombreEstado FROM tb_proveedor as p inner join tb_estado as e ON p.IdEstado = e.Id_Estado Where p.Id_Proveedor = %s", [id])
        proveedor = cur.fetchall()
        mysql.connection.commit()
        print(proveedor)

        return render_template('modal/proveedornuevo-modal.html', proveedor=proveedor)
    else:
        return "No"

# ACTUALIZAMOS EL PROVEEDOR


@app.route('/actProve', methods=["POST", "GET"])
def actProve():
    if request.method == "POST":
        id = request.form['id']
        nombre = request.form['nombre']
        cedula = request.form['cedula']
        cur = mysql.connection.cursor()
        cur.execute(
            "UPDATE tb_proveedor set NombreProveedor = %s, Cedula = %s Where Id_Proveedor = %s", (nombre, cedula, id))
        proveedor = cur.fetchall()
        mysql.connection.commit()

        return "listo"
    else:
        return "No"
# ACTUALIZAR USUARIOS


@app.route('/actUsu', methods=["POST", "GET"])
def actUsu():
    if request.method == "POST":
        id = request.form['id']
        nombre = request.form['nombre']
        loginUser = request.form['loginUser']
        cedula = request.form['cedula']
        rol = request.form['rol']
        punto = request.form['punto']
        cargo = request.form['cargo']
        contra = request.form['contra']
        flag = request.form['flag']
        if flag == "nocontra":
            cur = mysql.connection.cursor()
            cur.execute("UPDATE tb_usuarios set NombreUsuario = %s, Cedula = %s, IdCargo = %s,IdPuesto = %s Where Id_Usuario = %s",
                        (nombre, cedula, cargo, punto, id))
            proveedor = cur.fetchall()
            mysql.connection.commit()
            # MODIFICAMOS EL ROL
            cur = mysql.connection.cursor()
            cur.execute(
                "UPDATE tb_credenciales set Usuarios = %s, IdRol = %s Where Id_Credenciales = %s", (loginUser, rol, id))
            proveedor = cur.fetchall()
            mysql.connection.commit()
        elif flag == "contra":
            contra = request.form['contra']
            cur = mysql.connection.cursor()
            cur.execute("UPDATE tb_usuarios set NombreUsuario = %s, Cedula = %s, IdCargo = %s,IdPuesto = %s Where Id_Usuario = %s",
                        (nombre, cedula, cargo, punto, id))
            proveedor = cur.fetchall()
            mysql.connection.commit()
            # MODIFICAMOS EL ROL
            haseho = generate_password_hash(contra)
            cur = mysql.connection.cursor()
            cur.execute("UPDATE tb_credenciales set Usuarios = %s,Contraseñas = %s, IdRol = %s Where Id_Credenciales = %s",
                        (loginUser, haseho, rol, id))
            proveedor = cur.fetchall()
            mysql.connection.commit()

        return "listo"
    else:
        return "No"

# ADD USUARIOS


@app.route('/addUsu', methods=["POST", "GET"])
def addUsu():
    if request.method == "POST":
        nombre = request.form['nombre']
        loginUser = request.form['loginUser']
        cedula = request.form['cedula']
        rol = request.form['rol']
        punto = request.form['punto']
        cargo = request.form['cargo']
        contra = request.form['contra']

        contra = request.form['contra']
        # CREAMOS PRIMERO LAS CREDENCIALES
        haseho = generate_password_hash(contra)
        cur = mysql.connection.cursor()
        cur.execute(
            "INSERT INTO tb_credenciales (Usuarios,Contraseñas,IdRol) VALUES (%s,%s,%s)", (loginUser, haseho, rol))
        cred = cur.fetchall()
        mysql.connection.commit()

        cur = mysql.connection.cursor()
        cur.execute(
            "SELECT * FROM tb_credenciales ORDER by Id_Credenciales DESC LIMIT 1")
        cred = cur.fetchone()
        mysql.connection.commit()
        print(cred)
        # AHORA CREAMOS EL USUARIO COMO TAL
        cur = mysql.connection.cursor()
        cur.execute("INSERT INTO tb_usuarios (NombreUsuario,Cedula,IdCargo,IdCredenciales,IdEstado,IdPuesto) VALUES (%s,%s,%s,%s,1,%s)",
                    (nombre, cedula, cargo, cred[0], punto))
        proveedor = cur.fetchall()
        mysql.connection.commit()
        print("listo")
        return "listo"
    else:
        return "No"

# ELIMINAMOS EL PROVEEDOR


@app.route('/eliminarProveedor', methods=["POST", "GET"])
def eliminarProveedor():
    if request.method == "POST":
        id = request.form['id']
        cur = mysql.connection.cursor()
        cur.execute(
            "Update tb_proveedor set IdEstado = 2 Where Id_Proveedor = %s", [id])

        mysql.connection.commit()

        return "done"
    else:
        return "No"

# ACTUALIZAMOS EL PROVEEDOR


@app.route('/eliminarUsuario', methods=["POST", "GET"])
def eliminarUsuario():
    if request.method == "POST":
        id = request.form['id']
        if int(id) == int(session['userId']):
            return "no"
        else:
            cur = mysql.connection.cursor()
            cur.execute(
                "Update tb_usuarios set IdEstado = 2 Where Id_Usuario = %s", [id])

            mysql.connection.commit()

            return "done"
    else:
        return "No"

# HABILITAMOS EL PROVEEDOR QUE SE ENCUENTRA INACTIVO


@app.route('/habilitarProveedor', methods=["POST", "GET"])
def habilitarProveedor():
    if request.method == "POST":
        id = request.form['id']
        cur = mysql.connection.cursor()
        cur.execute(
            "Update tb_proveedor set IdEstado = 1 Where Id_Proveedor = %s", [id])

        mysql.connection.commit()

        return "done"
    else:
        return "No"

# HABILITAMOS EL USUARIO QUE SE ENCUENTRA INACTIVO


@app.route('/habilitarUsuario', methods=["POST", "GET"])
def habilitarUsuario():
    if request.method == "POST":
        id = request.form['id']
        cur = mysql.connection.cursor()
        cur.execute(
            "Update tb_USuarios set IdEstado = 1 Where Id_Usuario = %s", [id])

        mysql.connection.commit()

        return "done"
    else:
        return "No"


@app.route('/buscarUsuario', methods=["POST", "GET"])
def buscarUsuario():
    if request.method == "POST":
        usuario = request.form['proveedor']
        cur = mysql.connection.cursor()
        cur.execute(
            "select * from tb_usuarios Where IdEstado = 1 AND NombreUsuario like %s", [usuario+'%'])
        proveedores = cur.fetchall()
        print(proveedores)
        return render_template('otros/proveedor-busqueda.html', proveedores=proveedores)


@app.route('/buscarUsuarioNuevo', methods=["POST", "GET"])
def buscarUsuarioNuevo():
    if request.method == "POST":
        usuario = request.form['usuario']
        cur = mysql.connection.cursor()
        cur.execute(
            "select * from tb_usuarios Where IdEstado = 1 AND NombreUsuario like %s", [usuario+'%'])
        usuario = cur.fetchall()
        if usuario:
            return "existe"
        else:
            return "no existe"
        return render_template('otros/proveedor-busqueda.html', proveedores=proveedores)

# DETALLES DE USUARIO Y ESTE MODAL ES PARA AÑADIR USUARIOS NUEVOS DEPENDIENDO DE LA FLAG


@app.route('/detalleUsuarios', methods=["POST", "GET"])
def detalleUsuarios():
    if request.method == "POST":

        flagUSuario = request.form['flagUsuario']
        cedula = request.form['nombre']
        if flagUSuario == "editar":
            id = request.form['id']
            cur = mysql.connection.cursor()
            cur.execute(
                "SELECT u.Id_Usuario,u.NombreUsuario as Nombre,c.Usuarios as Usuario,r.NombreRol,e.NombreEstado,u.Cedula,car.NombreCargo,u.IdPuesto FROM tb_usuarios as u inner join tb_credenciales as c ON u.IdCredenciales = c.Id_Credenciales inner join tb_roles as r on c.IdRol = r.Id_Rol inner join tb_estado as e on u.IdEstado = e.Id_Estado inner join tb_cargo as car on u.IdCargo = car.Id_Cargo where u.Id_Usuario = %s", [id])
            usuario = cur.fetchall()

        # TRAEMOS LOS CARGOS
        cur = mysql.connection.cursor()
        cur.execute("SELECT * From tb_cargo")
        cargos = cur.fetchall()
        # TRAEMOS LOS ROLES
        cur = mysql.connection.cursor()
        cur.execute("SELECT * From tb_roles")
        roles = cur.fetchall()
        # TRAEMOS LOS PUNTOS DE VENTA
        cur = mysql.connection.cursor()
        cur.execute("SELECT * From tb_puntocompra")
        punto = cur.fetchall()
        return render_template('modal/usuarionuevo-modal.html', punto=punto, flagUSuario=flagUSuario, info=usuario, cargos=cargos, roles=roles)

# LISTA DE PROVEEDORES TABLA
@app.route('/listaDesbloqueo', methods=["POST", "GET"])
def listaDesbloqueo():
    if request.method == "POST":
        #MANDAR A LLAMAR LA TABLA DE DESBLOQUEOS
        id = request.form['proveedor']
        if id:
            cur = mysql.connection.cursor()
            cur.execute("select IdVerificacion from tb_desbloqueos where IdVerificacion like %s",[id+'%'])
            numerosBoletas = cur.fetchall()
            verificaciones = []
            print('Numero Boletas: ',numerosBoletas)
            for boleta in numerosBoletas:
                # LLAMAMOS LAS VERIFICACIONES DE CADA NUMERO DE BOLETA
                cur = mysql.connection.cursor()
                cur.execute(
                        "select v.Id_Verificacion,v.NoBoleta,v.Fecha,pc.NombrePuntoCompra, p.NombreProveedor,digi.NombreUsuario as digitador,veri.NombreUsuario as verificador from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_usuarios as digi on v.IdDigitador = digi.Id_Usuario inner join tb_usuarios as veri on v.IdVerificador = veri.Id_Usuario Where v.Id_Verificacion = %s", [boleta[0]])
                verifi = cur.fetchall()
                mysql.connection.commit()
                verificaciones.append(verifi)
            print("trae algo")
            print(verificaciones)
            return render_template('tablas/tabla-desbloqueo.html', verificaciones=verificaciones)
        else:
         
            cur = mysql.connection.cursor()
            cur.execute("select IdVerificacion from tb_desbloqueos")
            numerosBoletas = cur.fetchall()
            print('Numero Boletas: ',numerosBoletas)
            verificaciones = []
            for boleta in numerosBoletas:
                # LLAMAMOS LAS VERIFICACIONES DE CADA NUMERO DE BOLETA
                print('boleta: ',boleta[0])
                cur = mysql.connection.cursor()
                cur.execute(
                        "select v.Id_Verificacion,v.NoBoleta,v.Fecha,pc.NombrePuntoCompra, p.NombreProveedor,digi.NombreUsuario as digitador,veri.NombreUsuario as verificador from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_usuarios as digi on v.IdDigitador = digi.Id_Usuario inner join tb_usuarios as veri on v.IdVerificador = veri.Id_Usuario Where v.Id_Verificacion = %s", [boleta[0]])
                verifi1 = cur.fetchall()
                mysql.connection.commit()
                verificaciones.append(verifi1)
            print("Vacio")
            print(verificaciones)
        return render_template('tablas/tabla-desbloqueo.html', verificaciones=verificaciones)






@app.route('/listaProveedores', methods=["POST", "GET"])
def listaProveedores():
    if request.method == "POST":
        proveedor = request.form['proveedor']
        if session['cargo'] != 5 and session['cargo'] != 6:
            if proveedor == "":
                print("DIGITADOR")
                print(session['cargo'])
                # LLamar la verificacion de ese proveedor
                cur = mysql.connection.cursor()
                cur.execute("select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado,p.NombreProveedor,v.Bahia from tb_verificacion as v inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra left join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor Where v.IdEstado = 3 AND v.IdUsuarioCreacion = %s", [
                            session["userId"]])
                verificaciones = cur.fetchall()
                return render_template('tablas/tabla-proveedores.html', verificaciones=verificaciones)
            else:
                # SELECCIONAR EL ID DEL PROVEEDOR
                cur = mysql.connection.cursor()
                cur.execute(
                    "select * from tb_proveedor Where NombreProveedor = %s", [proveedor])
                proveedornuevo = cur.fetchone()
                # INSERTAMOS LA VERIFICACION
                hi = capturarHora()
                fecha = hi.replace(microsecond=0)
                print(proveedor)
                # if proveedor == "001":
                #     print("entro en la condicion del codigo")
                #     proveedor = 'CUADRILLA CHATARRA'
                # elif proveedor == "002":
                #     proveedor = 'CASETA'
                # elif proveedor == '003':
                #     proveedor = 'MOVIL'
                # elif proveedor == '004':
                #     proveedor = 'CRN'
                # print(proveedor)

                #GENERAMOS EL NUMERO RANDOM PARA EL ENLACE 
                if session['cargo'] == 1: # SOLAMENTE EL DIGITADOR PODRA ASIGNAR EL N DE ENLACE
                    # Obtenemos la hora actual
                    now = datetime.now()

                    # Utilizamos la hora actual para generar una semilla para la función random
                    random.seed(now.hour * 3600 + now.minute * 60 + now.second)

                    # Generamos un número aleatorio entre 0 y 100
                    random_number = random.randint(0, 100)
                    cur = mysql.connection.cursor()
                    cur.execute("INSERT INTO tb_verificacion (Fecha,NoBoleta,IdPuntoCompra,IdEstado,IdUsuarioCreacion) VALUES (%s,%s,%s,3,%s)",
                                (fecha, proveedor, 5, session["userId"]))
                    proveedornuevo = cur.fetchone()
                else:
                    cur = mysql.connection.cursor()
                    cur.execute("INSERT INTO tb_verificacion (Fecha,NoBoleta,IdPuntoCompra,IdEstado,IdUsuarioCreacion) VALUES (%s,%s,%s,3,%s)",
                                (fecha, proveedor, 5, session["userId"]))
                    proveedornuevo = cur.fetchone()
                # LLAMAMOS LAS VERIFICACIONES DEL USUARIO
                cur = mysql.connection.cursor()
                cur.execute("select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado,p.NombreProveedor,v.Bahia from tb_verificacion as v inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra left join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor Where v.IdEstado = 3 AND v.IdUsuarioCreacion = %s", [
                            session["userId"]])
                verificaciones = cur.fetchall()
                mysql.connection.commit()
                print("VERIFICADICA")
                print(verificaciones)
                return render_template('tablas/tabla-proveedores.html', verificaciones=verificaciones)
        elif session['cargo'] == 5:
            print("entro como validador")
            if proveedor == "":
                print("cargooo")
                print(session['cargo'])
                consulta_extra = "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado,p.NombreProveedor,v.Bahia from tb_verificacion as v inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra left join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor Where v.IdEstado = 5 AND "
                # LLamar la verificacion de ese proveedor
                cur = mysql.connection.cursor()
                cur.execute("select v.NoBoleta,v.PO,count(v.NoBoleta) as boleta from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner JOIN tb_usuarios as u on v.IdUsuarioCreacion = u.Id_Usuario inner join tb_cargo as car ON u.IdCargo = car.Id_Cargo Where v.IdEstado = 5 group BY v.PO")
                verificacionesRep = cur.fetchall()
                verificaciones = ""
                banderaConsultas = 0

                for ver in verificacionesRep:
                    if ver[2] == 2:
                        consulta_extra += " v.NoBoleta = '" + \
                            str(ver[0])+"' AND v.PO = '" + str(ver[1])+"' OR "
                        print(ver)
                        print("cambiar bandera")
                        banderaConsultas = 1
                # print(consulta_extra)
                print("Bandera consulta: ", banderaConsultas)
                if banderaConsultas == 1:
                    total = consulta_extra[:-4]
                    print(total)
                    # LLamar las verificaciones validas
                    cur = mysql.connection.cursor()
                    cur.execute("" + total+'group BY v.NoBoleta,v.PO')
                    verificaciones = cur.fetchall()
                    print("verifi")
                    return render_template('tablas/tabla-comparacion.html', verificaciones=verificaciones)
                else:
                    verificaciones = ""
                # TENEMOS QUE AGARRAR LA PO Y COMPARARLA PARA VER SI
                    print("fuera")
                    return render_template('tablas/tabla-comparacion.html', verificaciones=verificaciones)
            else:
                print(proveedor)
                cur = mysql.connection.cursor()
                cur.execute("select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado,p.NombreProveedor,v.Bahia from tb_verificacion as v inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra left join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor Where v.IdEstado = 5 AND v.NoBoleta like %s group by v.PO",[proveedor+'%'])
                verificaciones = cur.fetchall()
                print(verificaciones)
                return render_template('tablas/tabla-comparacion.html', verificaciones=verificaciones)
        


        elif session['cargo'] == 6:
            print("entro como jefe")
            if proveedor == "":
                print("cargooo")
                print(session['cargo'])
                # PARTE DEL FILTRO DE LAS VERIFICACIONES
                #group BY v.NoBoleta,v.PO,v.IdVerificador,v.IdDigitador,v.IdEstado,v.IdProveedor
                consulta_extra = "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado,p.NombreProveedor,v.Bahia from tb_verificacion as v inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra left join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor Where v.IdEstado = 7 group BY v.NoBoleta,v.PO,v.IdVerificador,v.IdDigitador,v.IdEstado,v.IdProveedor " 
                # LLamar la verificacion de ese proveedor
                cur = mysql.connection.cursor()
                cur.execute("select v.NoBoleta,v.PO,count(v.NoBoleta) as boleta from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner JOIN tb_usuarios as u on v.IdUsuarioCreacion = u.Id_Usuario inner join tb_cargo as car ON u.IdCargo = car.Id_Cargo Where v.IdEstado = 7 group BY v.NoBoleta,v.PO,v.IdVerificador,v.IdDigitador,v.IdEstado,v.IdProveedor ") #group BY v.NoBoleta,v.PO,v.IdVerificador,v.IdDigitador,v.IdEstado,v.IdProveedor
                verificacionesRep = cur.fetchall()
                verificaciones = ""
                banderaConsultas = 1
                print('en esta funcion')
                for ver in verificacionesRep:
                    if ver[2] == 2:
                        consulta_extra += " v.NoBoleta = '" + \
                            str(ver[0])+"' AND v.PO = '" + str(ver[1])+"' OR "
                        #print("cambiar bandera")
                        banderaConsultas = 1
                # print(consulta_extra)
                
                if banderaConsultas == 1:
                    total = consulta_extra[:-4]
                    print(total)
                    # LLamar las verificaciones validas
                    cur = mysql.connection.cursor()
                    cur.execute("select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado,p.NombreProveedor,v.Bahia from tb_verificacion as v inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra left join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor Where v.IdEstado = 7 group BY v.NoBoleta,v.PO,v.IdVerificador,v.IdDigitador,v.IdEstado,v.IdProveedor") #group BY v.NoBoleta,v.PO,v.IdVerificador,v.IdDigitador,v.IdEstado,v.IdProveedor
                    verificaciones = cur.fetchall()
                    print("verifi")
                    return render_template('tablas/tabla-comparacion.html', verificaciones=verificaciones)
                else:
                    verificaciones = ""
                # TENEMOS QUE AGARRAR LA PO Y COMPARARLA PARA VER SI
                    print("fuera")
                    return render_template('tablas/tabla-comparacion.html', verificaciones=verificaciones)
            else:
                # SELECCIONAR EL ID DEL PROVEEDOR
                cur = mysql.connection.cursor()
                cur.execute(
                    "select * from tb_proveedor Where NombreProveedor = %s", [proveedor])
                proveedornuevo = cur.fetchone()
                # INSERTAMOS LA VERIFICACION
                # hora = capturarHora()
                # fecha = hora.replace(microsecond=0)
                # print(fecha)
                # cur = mysql.connection.cursor()
                # cur.execute("INSERT INTO tb_verificacion (Fecha,NoBoleta,IdPuntoCompra,IdEstado,IdUsuarioCreacion) VALUES (%s,%s,%s,%s,%s)",
                #             (fecha, proveedor, 5, 3, session["userId"]))
                # proveedornuevo = cur.fetchone()
                # LLAMAMOS LAS VERIFICACIONES DEL USUARIO
                cur = mysql.connection.cursor()
                cur.execute("select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado, p.NombreProveedor,v.Bahia from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra Where v.IdEstado = 7 AND v.NoBoleta LIKE %s", [
                            proveedor + '%'])
                verificaciones = cur.fetchall()
                mysql.connection.commit()
                print("VERIFICADICA")
                print(verificaciones)
                return render_template('tablas/tabla-comparacion.html', verificaciones=verificaciones)
    else:
        return "No"
# DETALLE VERIFICACION


@app.route('/detalleVerificacion', methods=["POST", "GET"])
def detalleVerificacion():
    if request.method == "POST":

        id = request.form['id']
        if session['cargo'] != 5 and session['cargo'] != 6:
            print(id)
            print("aqui no")
            cur = mysql.connection.cursor()
            cur.execute(
                "select IdProveedor from tb_verificacion Where Id_Verificacion = %s", [id])
            nueva = cur.fetchone()
            mysql.connection.commit()
            print("nueva")
            print(nueva[0])
            if nueva[0] != None:
                print("tiene algo")
                cur = mysql.connection.cursor()
                cur.execute(
                    "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado, p.NombreProveedor,digi.NombreUsuario as digitador,veri.NombreUsuario as verificador,v.Bahia,cuad.NombreJefe from tb_verificacion as v inner join tb_cuadrilla as cuad on v.IdJefeCuadrilla = cuad.Id_CuadrillaJefe inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_usuarios as digi on v.IdDigitador = digi.Id_Usuario inner join tb_usuarios as veri on v.IdVerificador = veri.Id_Usuario Where v.IdEstado = 3 AND v.Id_Verificacion = %s", [id])
                verificacion = cur.fetchall()
                mysql.connection.commit()
                # cur = mysql.connection.cursor()
                # cur.execute(
                #     "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado, p.NombreProveedor,digi.NombreUsuario as digitador,veri.NombreUsuario as verificador,v.Bahia from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_usuarios as digi on v.IdDigitador = digi.Id_Usuario inner join tb_usuarios as veri on v.IdVerificador = veri.Id_Usuario Where v.IdEstado = 3 AND v.Id_Verificacion = %s", [id])
                # verificacion = cur.fetchall()
                # mysql.connection.commit()
                # if verificacion:
                #     print("a")
                # else:
                #     cur = mysql.connection.cursor()
                #     cur.execute(
                #         "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado, p.NombreProveedor,digi.NombreUsuario as digitador,veri.NombreUsuario as verificador,v.Bahia,cuad.NombreJefe from tb_verificacion as v inner join tb_cuadrilla as cuad on v.IdJefeCuadrilla = cuad.Id_CuadrillaJefe inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_usuarios as digi on v.IdDigitador = digi.Id_Usuario inner join tb_usuarios as veri on v.IdVerificador = veri.Id_Usuario Where v.IdEstado = 3 AND v.Id_Verificacion = %s", [id])
                #     verificacion = cur.fetchall()
                #     mysql.connection.commit()
                print(verificacion)
            else:
                # MANDAMOS A LLAMAR LOS DATOS QUE TENEMOS EN UNA VERIFICACION INCOMPLETA
                print("no tiene algo")
                cur = mysql.connection.cursor()
                cur.execute(
                    "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado from tb_verificacion as v inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra Where v.IdEstado = 3 AND v.Id_Verificacion = %s", [id])
                verificacion = cur.fetchall()
                mysql.connection.commit()
                print(verificacion)

            # ESTAS CONSULTAS SON PARA TRAER LOS PROVEEDORES Y LOS DATOS DE LOS SELECT
            # CONSULTA PARA LOS PUNTOS DE COMPRA
            cur = mysql.connection.cursor()
            cur.execute("select * from tb_puntocompra Where IdEstado = 1")
            punto = cur.fetchall()
            # CONSULTA PARA LOS MATERIALES
            cur = mysql.connection.cursor()
            cur.execute("select * from tb_material Where Id_Estado = 1")
            material = cur.fetchall()
            # CONSULTA PARA LOS VERIFICADORES
            cur = mysql.connection.cursor()
            cur.execute(
                "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 2")
            verificador = cur.fetchall()
            # CONSULTA PARA LOS DIGITADOR
            cur = mysql.connection.cursor()
            cur.execute(
                "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 1")
            digitador = cur.fetchall()

            # necesito mandar a llamar los dastos del usuario que esta logueado para ponerlo como verificador
            cur = mysql.connection.cursor()
            cur.execute("select u.Id_Usuario,car.NombreCargo,u.NombreUsuario from tb_usuarios as u inner join tb_cargo as car on u.IdCargo = car.Id_Cargo inner join tb_credenciales as cred ON u.IdCredenciales = cred.Id_Credenciales Where cred.Id_Credenciales = %s", [
                        session['userId']])
            usuariolog = cur.fetchone()
            print("USUARIO LOG")
            print(usuariolog)
            # necesito mandar a llamar los dastos del usuario para saber en que punto de venta esta
            cur = mysql.connection.cursor()
            cur.execute("select pc.Id_PuntoCompra,pc.NombrePuntoCompra from tb_usuarios as u inner join tb_puntocompra as pc on u.IdPuesto = pc.Id_PuntoCompra inner join tb_credenciales as cred on u.IdCredenciales = cred.Id_Credenciales where cred.Id_Credenciales = %s", [
                        session['userId']])
            usuariopunto = cur.fetchone()
            print(usuariopunto)
            return render_template('modal/verificaciones-modal.html', idpunto=session['puntoid'], punto=session['punto'], usuariopunto=usuariopunto, usuariolog=usuariolog, verificacion=verificacion, Punto=punto, Material=material, Verificador=verificador, Digitador=digitador)
        elif session['cargo'] == 5:
            # print("VER:",ver)
            print(id)
            print("pesoss")
            cur = mysql.connection.cursor()
            cur.execute(
                "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado, p.NombreProveedor,digi.NombreUsuario as digitador,veri.NombreUsuario as verificador,v.Bahia from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_usuarios as digi on v.IdDigitador = digi.Id_Usuario inner join tb_usuarios as veri on v.IdVerificador = veri.Id_Usuario Where v.IdEstado = 5 AND v.PO= %s", [id])
            nueva = cur.fetchall()
            mysql.connection.commit()
            print("nueva")
            print(nueva)
            if nueva:
                cur = mysql.connection.cursor()
                cur.execute(
                    "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado, p.NombreProveedor,digi.NombreUsuario as digitador,veri.NombreUsuario as verificador,v.Bahia from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_usuarios as digi on v.IdDigitador = digi.Id_Usuario inner join tb_usuarios as veri on v.IdVerificador = veri.Id_Usuario Where v.IdEstado = 5 AND v.PO = %s", [id])
                verificacion = cur.fetchall()
                mysql.connection.commit()
            else:
                cur = mysql.connection.cursor()
                cur.execute(
                    "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra Where v.IdEstado = 5 AND v.Id_Verificacion = %s", [id])
                verificacion = cur.fetchall()
                mysql.connection.commit()

            # ESTAS CONSULTAS SON PARA TRAER LOS PROVEEDORES Y LOS DATOS DE LOS SELECT
            # CONSULTA PARA LOS PUNTOS DE COMPRA
            cur = mysql.connection.cursor()
            cur.execute("select * from tb_puntocompra Where IdEstado = 1")
            punto = cur.fetchall()
            # CONSULTA PARA LOS MATERIALES
            cur = mysql.connection.cursor()
            cur.execute("select * from tb_material Where Id_Estado = 1")
            material = cur.fetchall()
            # CONSULTA PARA LOS VERIFICADORES
            cur = mysql.connection.cursor()
            cur.execute(
                "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 2")
            verificador = cur.fetchall()
            # CONSULTA PARA LOS DIGITADOR
            cur = mysql.connection.cursor()
            cur.execute(
                "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 1")
            digitador = cur.fetchall()

            # necesito mandar a llamar los dastos del usuario que esta logueado para ponerlo como verificador
            cur = mysql.connection.cursor()
            cur.execute("select cred.Id_Credenciales,u.NombreUsuario from tb_usuarios as u inner join tb_credenciales as cred on u.IdCredenciales = cred.Id_Credenciales Where cred.Id_Credenciales = %s", [
                        session['userId']])
            usuariolog = cur.fetchone()
            # necesito mandar a llamar los dastos del usuario para saber en que punto de venta esta
            cur = mysql.connection.cursor()
            cur.execute("select pc.Id_PuntoCompra,pc.NombrePuntoCompra from tb_usuarios as u inner join tb_puntocompra as pc on u.IdPuesto = pc.Id_PuntoCompra inner join tb_credenciales as cred on u.IdCredenciales = cred.Id_Credenciales where cred.Id_Credenciales = %s", [
                        session['userId']])
            usuariopunto = cur.fetchone()
            

            return render_template('modal/comparacion-modal.html',po='',ids= '',validacion = "",val = "", usuariopunto=usuariopunto, usuariolog=usuariolog, verificacion=verificacion, Punto=punto, Material=material, Verificador=verificador, Digitador=digitador)
        
        elif session['cargo'] == 6:
            print("JEFE")
            print(id)
            print("pesoss")
            # TENGO EL ID DE UNA VERIFICACION ( SE NECESITA MANDAR A EXTRAER EL NUNMERO DE BOLETA CON AMBOS PESOS)
            # * MANDAMOS A TRAER LOS PESOS DEL DIGITADOR Y VERIFICADOR
            cur = mysql.connection.cursor()
            cur.execute(
                "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado, p.NombreProveedor,digi.NombreUsuario as digitador,veri.NombreUsuario as verificador,v.Bahia from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_usuarios as digi on v.IdDigitador = digi.Id_Usuario inner join tb_usuarios as veri on v.IdVerificador = veri.Id_Usuario Where v.IdEstado = 7 AND v.Id_Verificacion= %s", [id])
            verificacion = cur.fetchall()
            mysql.connection.commit()
            print("nueva")
            print("datos verificacion: ",verificacion)


            return render_template('modal/modal-jefe.html', verificacion = verificacion)



@app.route('/detalleVerificacionVal', methods=["POST", "GET"])
def detalleVerificacionVal():
    if request.method == "POST":

        id = request.form['id']
        val = request.form['val']
        if session['cargo'] == 5:
            # print("VER:",ver)
            print(id)
            print("pesoss")
            cur = mysql.connection.cursor()
            cur.execute(
                "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado, p.NombreProveedor,digi.NombreUsuario as digitador,veri.NombreUsuario as verificador,v.Bahia from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_usuarios as digi on v.IdDigitador = digi.Id_Usuario inner join tb_usuarios as veri on v.IdVerificador = veri.Id_Usuario Where v.PO= %s", [id])
            nueva = cur.fetchall()
            mysql.connection.commit()
            print("nueva")
            print(nueva)
            if nueva:
                cur = mysql.connection.cursor()
                cur.execute(
                    "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado, p.NombreProveedor,digi.NombreUsuario as digitador,veri.NombreUsuario as verificador,v.Bahia from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_usuarios as digi on v.IdDigitador = digi.Id_Usuario inner join tb_usuarios as veri on v.IdVerificador = veri.Id_Usuario Where v.PO = %s", [id])
                verificacion = cur.fetchall()
                mysql.connection.commit()
            else:
                cur = mysql.connection.cursor()
                cur.execute(
                    "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra Where v.PO = %s", [id])
                verificacion = cur.fetchall()
                mysql.connection.commit()
                print("VERIFICACION: ",verificacion)

            # ESTAS CONSULTAS SON PARA TRAER LOS PROVEEDORES Y LOS DATOS DE LOS SELECT
            # CONSULTA PARA LOS PUNTOS DE COMPRA
            cur = mysql.connection.cursor()
            cur.execute("select * from tb_puntocompra Where IdEstado = 1")
            punto = cur.fetchall()
            # CONSULTA PARA LOS MATERIALES
            cur = mysql.connection.cursor()
            cur.execute("select * from tb_material Where Id_Estado = 1")
            material = cur.fetchall()
            # CONSULTA PARA LOS VERIFICADORES
            cur = mysql.connection.cursor()
            cur.execute(
                "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 2")
            verificador = cur.fetchall()
            # CONSULTA PARA LOS DIGITADOR
            cur = mysql.connection.cursor()
            cur.execute(
                "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 1")
            digitador = cur.fetchall()

            # necesito mandar a llamar los dastos del usuario que esta logueado para ponerlo como verificador
            cur = mysql.connection.cursor()
            cur.execute("select cred.Id_Credenciales,u.NombreUsuario from tb_usuarios as u inner join tb_credenciales as cred on u.IdCredenciales = cred.Id_Credenciales Where cred.Id_Credenciales = %s", [
                        session['userId']])
            usuariolog = cur.fetchone()
            # necesito mandar a llamar los dastos del usuario para saber en que punto de venta esta
            cur = mysql.connection.cursor()
            cur.execute("select pc.Id_PuntoCompra,pc.NombrePuntoCompra from tb_usuarios as u inner join tb_puntocompra as pc on u.IdPuesto = pc.Id_PuntoCompra inner join tb_credenciales as cred on u.IdCredenciales = cred.Id_Credenciales where cred.Id_Credenciales = %s", [
                        session['userId']])
            usuariopunto = cur.fetchone()

            #TENEMOS QUE LLAMAR LOS PESOS DE LAS BASCULAS
            
            cur = mysql.connection.cursor()
            cur.execute(
                "select * from tb_validacion Where IdVerificacion = %s",[verificacion[0][0]])
            validaciondatos = cur.fetchall()
            print(validaciondatos)
            print("VAL",val)
            return render_template('modal/comparacion-modal.html',po='',ids = "",validacion = validaciondatos,val = val, usuariopunto=usuariopunto, usuariolog=usuariolog, verificacion=verificacion, Punto=punto, Material=material, Verificador=verificador, Digitador=digitador)
        
        

# DETALLE VERIFICACION
@app.route('/detalleVerificacionAdmin', methods=["POST", "GET"])
def detalleVerificacionAdmin():
    if request.method == "POST":
        id = request.form['id']
        cur = mysql.connection.cursor()

        cur.execute(
            "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado, p.NombreProveedor,digi.NombreUsuario as digitador,veri.NombreUsuario as verificador,v.Bahia from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_usuarios as digi on v.IdDigitador = digi.Id_Usuario inner join tb_usuarios as veri on v.IdVerificador = veri.Id_Usuario Where v.Id_Verificacion = %s", [id])
        nueva = cur.fetchall()
        mysql.connection.commit()
        if nueva:
            cur = mysql.connection.cursor()
            cur.execute(
                "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado, p.NombreProveedor,digi.NombreUsuario as digitador,veri.NombreUsuario as verificador,v.Bahia,cuad.NombreJefe from tb_verificacion as v inner join tb_cuadrilla as cuad on v.IdJefeCuadrilla = cuad.Id_CuadrillaJefe inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_usuarios as digi on v.IdDigitador = digi.Id_Usuario inner join tb_usuarios as veri on v.IdVerificador = veri.Id_Usuario Where v.Id_Verificacion = %s", [id])
            verificacion = cur.fetchall()
            mysql.connection.commit()
            cur = mysql.connection.cursor()
            
                
        else:
            cur = mysql.connection.cursor()
            cur.execute(
                "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra Where v.Id_Verificacion = %s", [id])
            verificacion = cur.fetchall()
            mysql.connection.commit()
        # ESTAS CONSULTAS SON PARA TRAER LOS PROVEEDORES Y LOS DATOS DE LOS SELECT
        # CONSULTA PARA LOS PUNTOS DE COMPRA
        cur = mysql.connection.cursor()
        cur.execute("select * from tb_puntocompra Where IdEstado = 1")
        punto = cur.fetchall()
        # CONSULTA PARA LOS MATERIALES
        cur = mysql.connection.cursor()
        cur.execute("select * from tb_material Where Id_Estado = 1")
        material = cur.fetchall()
        # CONSULTA PARA LOS VERIFICADORES
        cur = mysql.connection.cursor()
        cur.execute(
            "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 2")
        verificador = cur.fetchall()
        # CONSULTA PARA LOS DIGITADOR
        cur = mysql.connection.cursor()
        cur.execute(
            "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 1")
        digitador = cur.fetchall()
        print(verificacion)
        return render_template('modal/admin-modal.html', verificacion=verificacion, Punto=punto, Material=material, Verificador=verificador, Digitador=digitador)

# DETALLE VERIFICACION


@app.route('/detalleVerificacionUser', methods=["POST", "GET"])
def detalleVerificacionUser():
    if request.method == "POST":
        print("ENTRROOO")
        id = request.form['id']
        cur = mysql.connection.cursor()
        cur.execute(
            "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado, p.NombreProveedor,digi.NombreUsuario as digitador,veri.NombreUsuario as verificador,v.Bahia from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_usuarios as digi on v.IdDigitador = digi.Id_Usuario inner join tb_usuarios as veri on v.IdVerificador = veri.Id_Usuario Where v.Id_Verificacion = %s ", [id])
        nueva = cur.fetchall()
        mysql.connection.commit()
        if nueva:
            cur = mysql.connection.cursor()
            cur.execute(
                "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado, p.NombreProveedor,digi.NombreUsuario as digitador,veri.NombreUsuario as verificador,v.Bahia from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_usuarios as digi on v.IdDigitador = digi.Id_Usuario inner join tb_usuarios as veri on v.IdVerificador = veri.Id_Usuario Where v.Id_Verificacion = %s", [id])
            verificacion = cur.fetchall()
            mysql.connection.commit()
            print(verificacion)
        else:
            cur = mysql.connection.cursor()
            cur.execute(
                "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra Where v.Id_Verificacion = %s and v.IdUsuarioCreacion = %s", (id, session['userId']))
            verificacion = cur.fetchall()
            mysql.connection.commit()
        # ESTAS CONSULTAS SON PARA TRAER LOS PROVEEDORES Y LOS DATOS DE LOS SELECT
        # CONSULTA PARA LOS PUNTOS DE COMPRA
        cur = mysql.connection.cursor()
        cur.execute("select * from tb_puntocompra Where IdEstado = 1")
        punto = cur.fetchall()
        # CONSULTA PARA LOS MATERIALES
        cur = mysql.connection.cursor()
        cur.execute("select * from tb_material Where Id_Estado = 1")
        material = cur.fetchall()
        # CONSULTA PARA LOS VERIFICADORES
        cur = mysql.connection.cursor()
        cur.execute(
            "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 2")
        verificador = cur.fetchall()
        # CONSULTA PARA LOS DIGITADOR
        cur = mysql.connection.cursor()
        cur.execute(
            "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 1")
        digitador = cur.fetchall()
        print(verificacion)
        return render_template('modal/user-modal.html', verificacion=verificacion, Punto=punto, Material=material, Verificador=verificador, Digitador=digitador)


# GUARDAR DATOS GENERALES DE LAS VERIFICACIONES CREADAS
@app.route('/datosGeneralesVerificacion', methods=["POST", "GET"])
def datosGeneralesVerificacion():
    if request.method == "POST":
        id = request.form['id']
        puntoCompra = request.form['puntoCompra']
        verificador = request.form['verificador']
        digitador = request.form['digitador']
        proveedor = request.form['proveedor']
        cuadrilla = request.form['cuadrilla']
        
        
        po = request.form['po']
        nboleta = request.form['nboleta']
        bahia = request.form['bahia']

        print(proveedor)
        cur = mysql.connection.cursor()
        cur.execute(
            'SELECT Id_Proveedor from tb_proveedor where NombreProveedor = %s', [proveedor])
        idprov = cur.fetchone()
        cur = mysql.connection.cursor()

        

        # BUSCAR EL JEFE
        print("NOmbre cuad,:",cuadrilla)
        if session['punto'] == 'GRANEL/PLANTA: Receipts' and cuadrilla != '':
            print('entroooooooo aqui')
            #ESTA ES DE PLANTA Y CON CUADRILLA
            cur = mysql.connection.cursor()
            cur.execute(
                'SELECT Id_CuadrillaJefe from tb_cuadrilla where NombreJefe = %s', [cuadrilla])
            idcuadrilla = cur.fetchone()
            cur = mysql.connection.cursor()
            print("CUADRILLA: ",idcuadrilla)
            
        else:
            idcuadrilla = ''

        print("aqui se muestra el proveedor:")
        print(idprov)
        print("CUADRILLA: ",idcuadrilla)

        if idprov:
            if idcuadrilla:
                cur.execute('Update tb_verificacion set PO = %s,NoBoleta = %s,IdProveedor = %s,IdVerificador = %s,IdDigitador = %s,IdPuntoCompra = %s,Bahia = %s,IdJefeCuadrilla = %s where Id_Verificacion = %s',
                            (po, nboleta, idprov[0], verificador, digitador, puntoCompra, bahia,idcuadrilla[0], id))
                digitador = cur.fetchall()
                mysql.connection.commit()
            else:
                    if session['punto'] == 'GRANEL/PLANTA: Receipts' and cuadrilla != '':
                        # llamar el id de oddo
                        idOddocuad = conexion.buscarIdCuadrilla(cuadrilla)
                        # INSERTAMOS EL PROVEEDOR DE ODDO EN LA BASE
                        cur = mysql.connection.cursor()
                        cur.execute(
                            "INSERT INTO tb_cuadrilla (Id_CuadrillaJefe,NombreJefe) VALUES (%s,%s)", (idOddocuad,cuadrilla))
                        cuadrillanuevo = cur.fetchone()
                        # LLAMAMOS AL PROVEEDOR DE NOMBRE TAL
                        cur = mysql.connection.cursor()
                        cur.execute(
                            'SELECT Id_CuadrillaJefe from tb_cuadrilla where NombreJefe = %s', [cuadrilla])
                        idcuad = cur.fetchone()
                        cur = mysql.connection.cursor()

                        cur.execute('Update tb_verificacion set PO = %s,NoBoleta = %s,IdProveedor = %s,IdVerificador = %s,IdDigitador = %s,IdPuntoCompra = %s,Bahia = %s,IdJefeCuadrilla = %s where Id_Verificacion = %s',
                                    (po, nboleta, idprov[0], verificador, digitador, puntoCompra, bahia,idcuad[0], id))
                        digitador = cur.fetchall()
                        mysql.connection.commit()
                    else:
                        cur.execute('Update tb_verificacion set PO = %s,NoBoleta = %s,IdProveedor = %s,IdVerificador = %s,IdDigitador = %s,IdPuntoCompra = %s,Bahia = %s,IdJefeCuadrilla = %s where Id_Verificacion = %s',
                                (po, nboleta, idprov[0], verificador, digitador, puntoCompra, bahia,2, id))
                        digitador = cur.fetchall()
                        mysql.connection.commit()
            
        else:
            # VALIDACIONES DE CUADRILLA SI NO ENCONTRAMOS AL PROVEEDOR EN LA BASE DE DATOS
            if idcuadrilla:
                 # llamar el id de oddo
                idOddo = conexion.buscarIdProveedor(proveedor)
                # INSERTAMOS EL PROVEEDOR DE ODDO EN LA BASE
                cur = mysql.connection.cursor()
                cur.execute(
                    "INSERT INTO tb_proveedor (NombreProveedor,IdOddo,IdEstado) VALUES (%s,%s,1)", (proveedor, idOddo))
                proveedornuevo = cur.fetchone()
                # LLAMAMOS AL PROVEEDOR DE NOMBRE TAL
                cur = mysql.connection.cursor()
                cur.execute(
                    'SELECT Id_Proveedor from tb_proveedor where NombreProveedor = %s', [proveedor])
                idprov = cur.fetchone()
                cur = mysql.connection.cursor()

                


                cur.execute('Update tb_verificacion set PO = %s,NoBoleta = %s,IdProveedor = %s,IdVerificador = %s,IdDigitador = %s,IdPuntoCompra = %s,Bahia = %s,IdJefeCuadrilla = %s where Id_Verificacion = %s',
                            (po, nboleta, idprov[0], verificador, digitador, puntoCompra, bahia,idcuadrilla[0], id))
                digitador = cur.fetchall()
                mysql.connection.commit()
            else:
                #SINO ENCONTRAMOS A NINGUNO DE LOS DOS
                if session['punto'] == 'GRANEL/PLANTA: Receipts' and cuadrilla != '':
                        #llamar el id de oddo
                        idOddocuad = conexion.buscarIdCuadrilla(cuadrilla)
                        # INSERTAMOS LA CUADRILLA DE ODDO EN LA BASE
                        cur = mysql.connection.cursor()
                        cur.execute(
                            "INSERT INTO tb_cuadrilla (Id_CuadrillaJefe,NombreJefe) VALUES (%s,%s)", (idOddocuad,cuadrilla))
                        cuadrillanuevo = cur.fetchone()
                        # LLAMAMOS AL PROVEEDOR DE NOMBRE TAL
                        cur = mysql.connection.cursor()
                        cur.execute(
                            'SELECT Id_CuadrillaJefe from tb_cuadrilla where NombreJefe = %s', [cuadrilla])
                        idcuad = cur.fetchone()
                        cur = mysql.connection.cursor()


                        idOddo = conexion.buscarIdProveedor(proveedor)
                        # INSERTAMOS EL PROVEEDOR DE ODDO EN LA BASE
                        cur = mysql.connection.cursor()
                        cur.execute(
                            "INSERT INTO tb_proveedor (NombreProveedor,IdOddo,IdEstado) VALUES (%s,%s,1)", (proveedor, idOddo))
                        proveedornuevo = cur.fetchone()
                        # LLAMAMOS AL PROVEEDOR DE NOMBRE TAL
                        cur = mysql.connection.cursor()
                        cur.execute(
                            'SELECT Id_Proveedor from tb_proveedor where NombreProveedor = %s', [proveedor])
                        idprov = cur.fetchone()
                        cur = mysql.connection.cursor()

                        cur.execute('Update tb_verificacion set PO = %s,NoBoleta = %s,IdProveedor = %s,IdVerificador = %s,IdDigitador = %s,IdPuntoCompra = %s,Bahia = %s,IdJefeCuadrilla = %s where Id_Verificacion = %s',
                                    (po, nboleta, idprov[0], verificador, digitador, puntoCompra, bahia,idcuad[0], id))
                        digitador = cur.fetchall()
                        mysql.connection.commit()
                else:
                    idOddo = conexion.buscarIdProveedor(proveedor)
                    # INSERTAMOS EL PROVEEDOR DE ODDO EN LA BASE
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "INSERT INTO tb_proveedor (NombreProveedor,IdOddo,IdEstado) VALUES (%s,%s,1)", (proveedor, idOddo))
                    proveedornuevo = cur.fetchone()
                    # LLAMAMOS AL PROVEEDOR DE NOMBRE TAL
                    cur = mysql.connection.cursor()
                    cur.execute(
                        'SELECT Id_Proveedor from tb_proveedor where NombreProveedor = %s', [proveedor])
                    idprov = cur.fetchone()
                    cur = mysql.connection.cursor()
                    cur.execute('Update tb_verificacion set PO = %s,NoBoleta = %s,IdProveedor = %s,IdVerificador = %s,IdDigitador = %s,IdPuntoCompra = %s,Bahia = %s,IdJefeCuadrilla = %s where Id_Verificacion = %s',
                                (po, nboleta, idprov[0], verificador, digitador, puntoCompra, bahia,2, id))
                    digitador = cur.fetchall()
                    mysql.connection.commit()



            #========================================
           
        return "done"
#GUARDAMOS EL ID CON EL QUE QUEREMOS COMPARAR
# @app.route('/agregarComparacion', methods=["POST", "GET"])
# def agregarComparacion():
#     if request.method == "POST":
#         id = request.form['id']
#         comparacion = request.form['valor']

#         cur = mysql.connection.cursor()

#         cur.execute('Update tb_verificacion set NoEnlace = %s where Id_Verificacion = %s',(comparacion, id))
#         mysql.connection.commit()
        
#     return "HECHO"
# CARGAR LOS PESOS DE LAS VERIFICACIONES
@app.route('/listaPesosJefe', methods=["POST", "GET"])
def listaPesosJefe():
    if request.method == "POST":
        nbol = request.form['nboleta']
        print(nbol)

        # cur = mysql.connection.cursor()
        # cur.execute('SELECT NoBoleta from tb_verificacion where Id_verificacion = %s and IdEstado = 7 GROUP BY IdProveedor,IdDigitador,IdVerificador,Bahia,NoBoleta,PO', (id))
        # nbol = cur.fetchall()
        # mysql.connection.commit()
        # print(matverificador)
        # total de materiales del verificador
        cur = mysql.connection.cursor()
        cur.execute('SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2) as neto FROM tb_detalleverificacion as ver inner join tb_verificacion as v on ver.IdVerificacion = v.Id_Verificacion inner join tb_material as m ON ver.IdMaterial = m.Id_Material inner join tb_usuarios as u on v.IdUsuarioCreacion = u.Id_Usuario Where v.NoBoleta = %s and u.IdCargo = %s and v.IdEstado = 7 Group BY ver.IdMaterial', (nbol, 2))
        matverificador = cur.fetchall()
        mysql.connection.commit()
        print(matverificador)
        print("MAT DEL VERIFICADOR")

        # total de materiales del digitador
        cur = mysql.connection.cursor()
        cur.execute('SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2) as neto FROM tb_detalleverificacion as ver inner join tb_verificacion as v on ver.IdVerificacion = v.Id_Verificacion inner join tb_material as m ON ver.IdMaterial = m.Id_Material inner join tb_usuarios as u on v.IdUsuarioCreacion = u.Id_Usuario Where v.NoBoleta = %s and u.IdCargo = %s and v.IdEstado = 7 Group BY ver.IdMaterial', (nbol, 1))
        matdigitador = cur.fetchall()
        mysql.connection.commit()
        print(matdigitador)
        print("MAT DEL digitador")

        # LLamar los pesos del verificador
        cur = mysql.connection.cursor()
        cur.execute("SELECT dt.Id_DetalleVerificacion,dt.IdVerificacion,m.NombreMaterial,dt.PesoBruto,dt.PesoTara,dt.Destare,dt.PesoNeto FROM tb_detalleverificacion as dt inner join tb_verificacion as ver on dt.IdVerificacion = ver.Id_Verificacion inner join tb_verificacion as v ON dt.IdVerificacion = v.Id_Verificacion inner join tb_material as m ON dt.IdMaterial = m.Id_Material inner join tb_usuarios as u on v.IdUsuarioCreacion = u.Id_Usuario Where v.NoBoleta = %s and u.IdCargo = %s and v.IdEstado = 7", (nbol, 2))
        pesosverificador = cur.fetchall()

        # LLamar los pesos del digitador
        cur = mysql.connection.cursor()
        cur.execute("SELECT dt.Id_DetalleVerificacion,dt.IdVerificacion,m.NombreMaterial,dt.PesoBruto,dt.PesoTara,dt.Destare,dt.PesoNeto FROM tb_detalleverificacion as dt inner join tb_verificacion as ver on dt.IdVerificacion = ver.Id_Verificacion inner join tb_verificacion as v ON dt.IdVerificacion = v.Id_Verificacion inner join tb_material as m ON dt.IdMaterial = m.Id_Material inner join tb_usuarios as u on v.IdUsuarioCreacion = u.Id_Usuario Where v.NoBoleta = %s and u.IdCargo = %s and v.IdEstado = 7", (nbol, 1))
        pesosdigitador = cur.fetchall()

        idver = pesosverificador[0][1]
        print("idveeeeeeeeeeeeeer")
        print(idver)

        iddig = pesosdigitador[0][1]
        print("iddigiii")
        print(iddig)

        cur = mysql.connection.cursor()
        cur.execute("SELECT SUM(PesoBruto) FROM tb_detalleverificacion WHERE IdVerificacion = %s ", [idver])
        sumaBruto1 = cur.fetchone()
        if sumaBruto1[0]:
            sumaBrutover = round(sumaBruto1[0], 2)
        else:
            sumaBrutover = 0.00

        print(sumaBrutover)
        #  SUMA DE LA COLUMNA PESOS TARA
        cur = mysql.connection.cursor()
        cur.execute(
                    "SELECT SUM(PesoTara) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [idver])
        sumaTara1 = cur.fetchone()
        if sumaTara1[0]:
            sumaTaraver = round(sumaTara1[0], 2)
        else:
            sumaTaraver = 0.00

        #  SUMA DE LA COLUMNA PESOS DESTARE
        cur = mysql.connection.cursor()
        cur.execute(
                    "SELECT SUM(Destare) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
        sumaDestare1 = cur.fetchone()
        mysql.connection.commit()
        if sumaDestare1[0]:
            sumaDestarever = round(sumaDestare1[0], 2)
        else:
            sumaDestarever = 0.00

        #  SUMA DE LA COLUMNA PESOS NETO
        cur = mysql.connection.cursor()
        cur.execute(
                    "SELECT SUM(PesoNeto) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [idver])
        sumaNeto1 = cur.fetchone()
        mysql.connection.commit()
        if sumaNeto1[0]:
            sumaNetover = round(sumaNeto1[0], 2)
        else:
            sumaNetover = 0.00

        # HACEMOS LOA SUMA DE CADA COLUMNA PARA LOS DIGITADORES
        #  SUMA DE LA COLUMNA PESOS BRUTOS
        cur = mysql.connection.cursor()
        cur.execute(
                    "SELECT SUM(PesoBruto) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [iddig])
        sumaBruto2 = cur.fetchone()
        if sumaBruto2[0]:
            sumaBrutodig = round(sumaBruto2[0], 2)
        else:
            sumaBrutodig = 0.00

        #  SUMA DE LA COLUMNA PESOS TARA
        cur = mysql.connection.cursor()
        cur.execute(
                    "SELECT SUM(PesoTara) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [iddig])
        sumaTara2 = cur.fetchone()
        if sumaTara2[0]:
                    sumaTaradigi = round(sumaTara2[0], 2)
        else:
            sumaTaradigi = 0.00

        #  SUMA DE LA COLUMNA PESOS DESTARE
        cur = mysql.connection.cursor()
        cur.execute(
                    "SELECT SUM(Destare) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
        sumaDestare2 = cur.fetchone()
        mysql.connection.commit()
        if sumaDestare2[0]:
            sumaDestaredigi = round(sumaDestare2[0], 2)
        else:
            sumaDestaredigi = 0.00

        #  SUMA DE LA COLUMNA PESOS NETO
        cur = mysql.connection.cursor()
        cur.execute(
                    "SELECT SUM(PesoNeto) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [iddig])
        sumaNeto2 = cur.fetchone()
        mysql.connection.commit()
        if sumaNeto2[0]:
            sumaNetodigi = round(sumaNeto2[0], 2)
        else:
            sumaNetodigi = 0.00

        return render_template('tablas/tabla-pesoscomparacion.html',varios = "",TotalLinea = pesosdigitador, id=0, pesosdigitador=matdigitador, ids=0, pesosverificador=matverificador, sumaBrutover=sumaBrutover, sumaTaraver=sumaTaraver, sumaDestarever=sumaDestarever, sumaNetover=sumaNetover, sumaBrutodig=sumaBrutodig, sumaTaradigi=sumaTaradigi, sumaDestaredigi=sumaDestaredigi, sumaNetodigi=sumaNetodigi)   

#DESBLOQUEO DE VERIFICACIONES
@app.route('/desbloquear', methods=["POST", "GET"])
def desbloquear():
    if request.method == "POST":
        nboleta = request.form['nboleta']

        cur = mysql.connection.cursor()
        cur.execute(
                    "SELECT Id_Verificacion FROM tb_verificacion WHERE NoBoleta = %s and IdEstado = 7 group by NoBoleta limit 1", [nboleta])
        idver = cur.fetchone()


        cur = mysql.connection.cursor()
        cur.execute(
                    'Update tb_verificacion set IdEstado = 3 Where NoBoleta = %s and (IdEstado = 7 or IdEstado = 8)', [nboleta])
        mysql.connection.commit()

        

        
        fecha = capturarHora()
        # fechacreacion = datetime.date(fecha)
        #print(idver)
        cur = mysql.connection.cursor()
        cur.execute("INSERT INTO tb_desbloqueos (IdVerificacion,Fecha) VALUES (%s,%s)",
                    ( idver[0],fecha))
        mysql.connection.commit()
        return "Done"

# AQIUI VA LA FUNCION DE LISTA DE VARIOS IDS
@app.route('/listaVariosPesos', methods=["POST", "GET"])
def listaVariosPesos():
    if request.method == "POST":
    
        id1 = request.form['id']
        
        ids = json.loads(id1)
        lista_total_ver = []
        lista_total_dig = []
        # print(id1['registrosSeleccionados'][0])
        # print(ids['registrosSeleccionados'])
            
        sumaBrutover = 0
        for id in ids['registrosSeleccionados']:
            print("REGISTROS SELECCIONADOS:",id)
            cur = mysql.connection.cursor()
            cur.execute('SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2) as neto,round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_verificacion as v on ver.IdVerificacion = v.Id_Verificacion inner join tb_material as m ON ver.IdMaterial = m.Id_Material inner join tb_usuarios as u on v.IdUsuarioCreacion = u.Id_Usuario Where v.Id_Verificacion = %s Group BY ver.IdMaterial', [id])
            matverificador = cur.fetchall()
            mysql.connection.commit()
            print(matverificador)
            print("MAT VERIFICADOR")
            lista_total_ver.append(matverificador)
            # MANDAMOS A TRAER LOS PESOS DEL VERIFICADOR PARA SACAR EL TOTAL DE PESADAS DE ESE REGISTRO
            # LLamar los pesos del digitador
            # cur = mysql.connection.cursor()
            # cur.execute("SELECT dt.Id_DetalleVerificacion,dt.IdVerificacion,m.NombreMaterial,dt.PesoBruto,dt.PesoTara,dt.Destare,dt.PesoNeto FROM tb_detalleverificacion as dt inner join tb_verificacion as ver on dt.IdVerificacion = ver.Id_Verificacion inner join tb_verificacion as v ON dt.IdVerificacion = v.Id_Verificacion inner join tb_material as m ON dt.IdMaterial = m.Id_Material inner join tb_usuarios as u on v.IdUsuarioCreacion = u.Id_Usuario Where v.Id_Verificacion = %s and u.IdCargo = %s", (id, 1))
            # pesosdigitador = cur.fetchall()

        
        print("Materiales resultantes ver: ",lista_total_ver)
        total_fila = len(lista_total_ver)
        suma_elementos = {}

        for sublista in lista_total_ver:
            for elemento in sublista:
                material = elemento[0]
                if material not in suma_elementos:
                    suma_elementos[material] = list(elemento[1:])
                else:
                    suma_elementos[material] = [round(x + y, 2) for x, y in zip(suma_elementos[material], elemento[1:])]

        # Convertir el diccionario a una lista de tuplas en el formato deseado
        resultado = [(material,) + tuple(valores) for material, valores in suma_elementos.items()]

        print(resultado)

        # Obtener la cantidad de columnas
        num_columnas = len(resultado[0]) - 1

        # Sumar las columnas
        suma_columnas = [sum(columna) for columna in zip(*[sublista[1:] for sublista in resultado])]


        sumaBrutover = round(suma_columnas[0],2)
        sumaTaraver = round(suma_columnas[1],2)
        sumaDestarever = round(suma_columnas[3],2)
        sumaNetover = round(suma_columnas[2],2)
        print("========================\n",total_fila)
        lista = []
        lista.append(ids['registrosSeleccionados'])
        #print("AQUI VA EL VALOR DE LA LISTA ",tamano_lista_total = sum(len(sublista) for sublista in lista_total_ver))
        return render_template('tablas/tabla-pesoscomparacion.html',varios = "si",TotalLinea = sum(len(sublista) for sublista in lista_total_ver),  id=lista, pesosdigitador=resultado, ids=ids, pesosverificador=resultado, sumaBrutover=sumaBrutover, sumaTaraver=sumaTaraver, sumaDestarever=sumaDestarever, sumaNetover=sumaNetover, sumaBrutodig=sumaBrutover, sumaTaradigi=sumaTaraver, sumaDestaredigi=sumaDestarever, sumaNetodigi=sumaNetover)


#===========================================


@app.route('/listaPesos', methods=["POST", "GET"])
def listaPesos():
    if request.method == "POST":
        id = request.form['id']
        print(id)
        if session['cargo'] != 5:
            if id != "":
                # LLamar la verificacion de ese proveedor
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT dt.Id_DetalleVerificacion,dt.IdVerificacion,m.NombreMaterial,dt.PesoBruto,dt.PesoTara,dt.Destare,dt.PesoNeto FROM tb_detalleverificacion as dt inner join tb_material as m ON dt.IdMaterial = m.Id_Material Where dt.IdVerificacion = %s", [id])
                pesos = cur.fetchall()
                print(pesos)
                # HACEMOS LOA SUMA DE CADA COLUMNA
                #  SUMA DE LA COLUMNA PESOS BRUTOS
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(PesoBruto) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                sumaBruto1 = cur.fetchone()
                if sumaBruto1[0]:
                    sumaBruto = round(sumaBruto1[0], 2)
                else:
                    sumaBruto = 0.00

                print(sumaBruto)
                #  SUMA DE LA COLUMNA PESOS TARA
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(PesoTara) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                sumaTara1 = cur.fetchone()
                if sumaTara1[0]:
                    sumaTara = round(sumaTara1[0], 2)
                else:
                    sumaTara = 0.00

                #  SUMA DE LA COLUMNA PESOS DESTARE
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(Destare) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                sumaDestare1 = cur.fetchone()
                mysql.connection.commit()
                if sumaDestare1[0]:
                    sumaDestare = round(sumaDestare1[0], 2)
                else:
                    sumaDestare = 0.00

                #  SUMA DE LA COLUMNA PESOS NETO
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(PesoNeto) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                sumaNeto1 = cur.fetchone()
                mysql.connection.commit()
                if sumaNeto1[0]:
                    sumaNeto = round(sumaNeto1[0], 2)
                else:
                    sumaNeto = 0.00
                print(pesos)
                # total de materiales
                cur = mysql.connection.cursor()
                cur.execute(
                    'SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2) as neto,round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s Group BY ver.IdMaterial', [id])
                mat = cur.fetchall()
                mysql.connection.commit()
                print("material:", mat)
                return render_template('tablas/tabla-pesos.html', mat=mat, pesos=pesos, sumaBruto=sumaBruto, sumaTara=sumaTara, sumaDestare=sumaDestare, sumaNeto=sumaNeto)
            else:
                pesos = ""
                return render_template('tablas/tabla-pesos.html', pesos=pesos)
        else:
            if id != "":
                print("id: ",id)
                # total de materiales del verificador
                cur = mysql.connection.cursor()
                cur.execute('SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2),round(SUM(ver.Destare),2) as neto FROM tb_detalleverificacion as ver inner join tb_verificacion as v on ver.IdVerificacion = v.Id_Verificacion inner join tb_material as m ON ver.IdMaterial = m.Id_Material inner join tb_usuarios as u on v.IdUsuarioCreacion = u.Id_Usuario Where v.PO = %s and u.IdCargo = %s Group BY ver.IdMaterial', (id, 2))
                matverificador = cur.fetchall()
                mysql.connection.commit()
                print(matverificador)
                print("MAT DEL VERIFICADOR")

                # total de materiales del digitador
                cur = mysql.connection.cursor()
                cur.execute('SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2),round(SUM(ver.Destare),2) as neto FROM tb_detalleverificacion as ver inner join tb_verificacion as v on ver.IdVerificacion = v.Id_Verificacion inner join tb_material as m ON ver.IdMaterial = m.Id_Material inner join tb_usuarios as u on v.IdUsuarioCreacion = u.Id_Usuario Where v.PO = %s and u.IdCargo = %s Group BY ver.IdMaterial', (id, 1))
                matdigitador = cur.fetchall()
                mysql.connection.commit()
                print(matdigitador)
                print("MAT DEL digitador")

                # LLamar los pesos del verificador
                cur = mysql.connection.cursor()
                cur.execute("SELECT dt.Id_DetalleVerificacion,dt.IdVerificacion,m.NombreMaterial,dt.PesoBruto,dt.PesoTara,dt.Destare,dt.PesoNeto FROM tb_detalleverificacion as dt inner join tb_verificacion as ver on dt.IdVerificacion = ver.Id_Verificacion inner join tb_verificacion as v ON dt.IdVerificacion = v.Id_Verificacion inner join tb_material as m ON dt.IdMaterial = m.Id_Material inner join tb_usuarios as u on v.IdUsuarioCreacion = u.Id_Usuario Where v.PO = %s and u.IdCargo = %s", (id, 2))
                pesosverificador = cur.fetchall()

                # LLamar los pesos del digitador
                cur = mysql.connection.cursor()
                cur.execute("SELECT dt.Id_DetalleVerificacion,dt.IdVerificacion,m.NombreMaterial,dt.PesoBruto,dt.PesoTara,dt.Destare,dt.PesoNeto FROM tb_detalleverificacion as dt inner join tb_verificacion as ver on dt.IdVerificacion = ver.Id_Verificacion inner join tb_verificacion as v ON dt.IdVerificacion = v.Id_Verificacion inner join tb_material as m ON dt.IdMaterial = m.Id_Material inner join tb_usuarios as u on v.IdUsuarioCreacion = u.Id_Usuario Where v.PO = %s and u.IdCargo = %s", (id, 1))
                pesosdigitador = cur.fetchall()
                # LISTA DE IDS CON LA MISMA PO
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT dt.IdVerificacion FROM tb_detalleverificacion as dt inner join tb_verificacion as ver on dt.IdVerificacion = ver.Id_Verificacion inner join tb_verificacion as v ON dt.IdVerificacion = v.Id_Verificacion inner join tb_material as m ON dt.IdMaterial = m.Id_Material Where v.PO = %s ORDER BY v.Id_Verificacion ASC", [id])
                ids = cur.fetchall()
                # SELECCIONAR LAS VERIFICACIONES QUE TRAEN EL MISMO PO
                # cur = mysql.connection.cursor()
                # cur.execute("SELECT PO FROM tb_verificacion Where Id_Verificacion = %s",[pesos[1]])
                # POs = cur.fetchall()
                # print(POs)

                # HACEMOS LOA SUMA DE CADA COLUMNA PARA LOS VERIFICADORES
                #  SUMA DE LA COLUMNA PESOS BRUTOS
                idver = pesosverificador[0][1]
                print("idveeeeeeeeeeeeeer")
                print(idver)

                iddig = pesosdigitador[0][1]
                print("iddigiii")
                print(iddig)
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(PesoBruto) FROM tb_detalleverificacion WHERE IdVerificacion = %s ", [idver])
                sumaBruto1 = cur.fetchone()
                if sumaBruto1[0]:
                    sumaBrutover = round(sumaBruto1[0], 2)
                else:
                    sumaBrutover = 0.00

                print(sumaBrutover)
                #  SUMA DE LA COLUMNA PESOS TARA
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(PesoTara) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [idver])
                sumaTara1 = cur.fetchone()
                if sumaTara1[0]:
                    sumaTaraver = round(sumaTara1[0], 2)
                else:
                    sumaTaraver = 0.00

                #  SUMA DE LA COLUMNA PESOS DESTARE
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(Destare) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [idver])
                sumaDestare1 = cur.fetchone()
                mysql.connection.commit()
                print("DESTAREEEEEE: ",sumaDestare1)
                if sumaDestare1[0]:
                    sumaDestarever = round(sumaDestare1[0], 2)
                else:
                    sumaDestarever = 0.00

                #  SUMA DE LA COLUMNA PESOS NETO
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(PesoNeto) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [idver])
                sumaNeto1 = cur.fetchone()
                mysql.connection.commit()
                if sumaNeto1[0]:
                    sumaNetover = round(sumaNeto1[0], 2)
                else:
                    sumaNetover = 0.00

                # HACEMOS LOA SUMA DE CADA COLUMNA PARA LOS DIGITADORES
                #  SUMA DE LA COLUMNA PESOS BRUTOS
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(PesoBruto) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [iddig])
                sumaBruto2 = cur.fetchone()
                if sumaBruto2[0]:
                    sumaBrutodig = round(sumaBruto2[0], 2)
                else:
                    sumaBrutodig = 0.00

                #  SUMA DE LA COLUMNA PESOS TARA
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(PesoTara) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [iddig])
                sumaTara2 = cur.fetchone()
                if sumaTara2[0]:
                    sumaTaradigi = round(sumaTara2[0], 2)
                else:
                    sumaTaradigi = 0.00

                #  SUMA DE LA COLUMNA PESOS DESTARE
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(Destare) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [iddig])
                sumaDestare2 = cur.fetchone()
                mysql.connection.commit()
                if sumaDestare2[0]:
                    sumaDestaredigi = round(sumaDestare2[0], 2)
                else:
                    sumaDestaredigi = 0.00

                #  SUMA DE LA COLUMNA PESOS NETO
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(PesoNeto) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [iddig])
                sumaNeto2 = cur.fetchone()
                mysql.connection.commit()
                if sumaNeto2[0]:
                    sumaNetodigi = round(sumaNeto2[0], 2)
                else:
                    sumaNetodigi = 0.00
                print(pesosverificador)
                print(pesosdigitador)
                return render_template('tablas/tabla-pesoscomparacion.html',varios = "",TotalLinea = pesosdigitador,  id=ids[0], pesosdigitador=matdigitador, ids=ids, pesosverificador=matverificador, sumaBrutover=sumaBrutover, sumaTaraver=sumaTaraver, sumaDestarever=sumaDestarever, sumaNetover=sumaNetover, sumaBrutodig=sumaBrutodig, sumaTaradigi=sumaTaradigi, sumaDestaredigi=sumaDestaredigi, sumaNetodigi=sumaNetodigi)
            else:
                pesos = ""
                return render_template('tablas/tabla-pesoscomparacion.html',varios = "", TotalLinea = '', pesos=pesosdigitador)

    else:
        return "No"


# CARGAR LOS PESOS DE LAS VERIFICACIONES
@app.route('/listaPesosAdmin', methods=["POST", "GET"])
def listaPesosAdmin():
    if request.method == "POST":
        id = request.form['id']
        print(id)
        if session['cargo'] != 5:
            if id != "":
                tarasId = []
                # LLamar la verificacion de ese proveedor
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT dt.Id_DetalleVerificacion,dt.IdVerificacion,m.NombreMaterial,dt.PesoBruto,dt.PesoTara,dt.Destare,dt.PesoNeto FROM tb_detalleverificacion as dt inner join tb_material as m ON dt.IdMaterial = m.Id_Material Where dt.IdVerificacion = %s", [id])
                pesos = cur.fetchall()
                print(pesos)
                # HACEMOS LOA SUMA DE CADA COLUMNA
                #  SUMA DE LA COLUMNA PESOS BRUTOS
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(PesoBruto) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                sumaBruto1 = cur.fetchone()
                if sumaBruto1[0]:
                    sumaBruto = round(sumaBruto1[0], 2)
                else:
                    sumaBruto = 0.00

                print(sumaBruto)
                #  SUMA DE LA COLUMNA PESOS TARA
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(PesoTara) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                sumaTara1 = cur.fetchone()
                if sumaTara1[0]:
                    sumaTara = round(sumaTara1[0], 2)
                else:
                    sumaTara = 0.00

                #  SUMA DE LA COLUMNA PESOS DESTARE
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(Destare) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                sumaDestare1 = cur.fetchone()
                mysql.connection.commit()
                if sumaDestare1[0]:
                    sumaDestare = round(sumaDestare1[0], 2)
                else:
                    sumaDestare = 0.00

                #  SUMA DE LA COLUMNA PESOS NETO
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(PesoNeto) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                sumaNeto1 = cur.fetchone()
                mysql.connection.commit()
                if sumaNeto1[0]:
                    sumaNeto = round(sumaNeto1[0], 2)
                else:
                    sumaNeto = 0.00
                print(pesos)
                for peso in pesos:
                    #  SUMA DE LA COLUMNA PESOS DESTARE
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "SELECT * FROM tb_detalletara WHERE IdDetalleVerificacion = %s", [peso[0]])
                    tara = cur.fetchone()
                    mysql.connection.commit()
                    if tara:
                        tarasId.append(peso[0])
                    print("pesos: ", tarasId)

                return render_template('tablas/tabla-pesos-admin.html', tarasId=tarasId, pesos=pesos, sumaBruto=sumaBruto, sumaTara=sumaTara, sumaDestare=sumaDestare, sumaNeto=sumaNeto)
            else:
                pesos = ""
                return render_template('tablas/tabla-pesos-admin.html', pesos=pesos)
        else:
            if id != "":
                print("VALIDADOR")

                # total de materiales del verificador
                cur = mysql.connection.cursor()
                cur.execute('SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2) as neto FROM tb_detalleverificacion as ver inner join tb_verificacion as v on ver.IdVerificacion = v.Id_Verificacion inner join tb_material as m ON ver.IdMaterial = m.Id_Material inner join tb_usuarios as u on v.IdUsuarioCreacion = u.Id_Usuario Where v.PO = %s and u.IdCargo = %s  Group BY ver.IdMaterial', (id, 2))
                matverificador = cur.fetchall()
                mysql.connection.commit()
                print(matverificador)
                print("MAT DEL VERIFICADOR")

                # total de materiales del digitador
                cur = mysql.connection.cursor()
                cur.execute('SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2) as neto FROM tb_detalleverificacion as ver inner join tb_verificacion as v on ver.IdVerificacion = v.Id_Verificacion inner join tb_material as m ON ver.IdMaterial = m.Id_Material inner join tb_usuarios as u on v.IdUsuarioCreacion = u.Id_Usuario Where v.PO = %s and u.IdCargo = %s  Group BY ver.IdMaterial', (id, 1))
                matdigitador = cur.fetchall()
                mysql.connection.commit()
                print(matdigitador)
                print("MAT DEL digitador")

                # LLamar los pesos del verificador
                cur = mysql.connection.cursor()
                cur.execute("SELECT dt.Id_DetalleVerificacion,dt.IdVerificacion,m.NombreMaterial,dt.PesoBruto,dt.PesoTara,dt.Destare,dt.PesoNeto FROM tb_detalleverificacion as dt inner join tb_verificacion as ver on dt.IdVerificacion = ver.Id_Verificacion inner join tb_verificacion as v ON dt.IdVerificacion = v.Id_Verificacion inner join tb_material as m ON dt.IdMaterial = m.Id_Material inner join tb_usuarios as u on v.IdUsuarioCreacion = u.Id_Usuario Where v.PO = %s and u.IdCargo = %s ", (id, 2))
                pesosverificador = cur.fetchall()

                # LLamar los pesos del digitador
                cur = mysql.connection.cursor()
                cur.execute("SELECT dt.Id_DetalleVerificacion,dt.IdVerificacion,m.NombreMaterial,dt.PesoBruto,dt.PesoTara,dt.Destare,dt.PesoNeto FROM tb_detalleverificacion as dt inner join tb_verificacion as ver on dt.IdVerificacion = ver.Id_Verificacion inner join tb_verificacion as v ON dt.IdVerificacion = v.Id_Verificacion inner join tb_material as m ON dt.IdMaterial = m.Id_Material inner join tb_usuarios as u on v.IdUsuarioCreacion = u.Id_Usuario Where v.PO = %s and u.IdCargo = %s ", (id, 1))
                pesosdigitador = cur.fetchall()
                # LISTA DE IDS CON LA MISMA PO
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT dt.IdVerificacion FROM tb_detalleverificacion as dt inner join tb_verificacion as ver on dt.IdVerificacion = ver.Id_Verificacion inner join tb_verificacion as v ON dt.IdVerificacion = v.Id_Verificacion inner join tb_material as m ON dt.IdMaterial = m.Id_Material Where v.PO = %s  ORDER BY v.Id_Verificacion ASC", [id])
                ids = cur.fetchall()
                # SELECCIONAR LAS VERIFICACIONES QUE TRAEN EL MISMO PO
                # cur = mysql.connection.cursor()
                # cur.execute("SELECT PO FROM tb_verificacion Where Id_Verificacion = %s",[pesos[1]])
                # POs = cur.fetchall()
                # print(POs)

                # HACEMOS LOA SUMA DE CADA COLUMNA PARA LOS VERIFICADORES
                #  SUMA DE LA COLUMNA PESOS BRUTOS
                idver = pesosverificador[0][1]
                print("idveeeeeeeeeeeeeer")
                print(idver)

                iddig = pesosdigitador[0][1]
                print("iddigiii")
                print(iddig)
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(PesoBruto) FROM tb_detalleverificacion WHERE IdVerificacion = %s ", [idver])
                sumaBruto1 = cur.fetchone()
                if sumaBruto1[0]:
                    sumaBrutover = round(sumaBruto1[0], 2)
                else:
                    sumaBrutover = 0.00

                print(sumaBrutover)
                #  SUMA DE LA COLUMNA PESOS TARA
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(PesoTara) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [idver])
                sumaTara1 = cur.fetchone()
                if sumaTara1[0]:
                    sumaTaraver = round(sumaTara1[0], 2)
                else:
                    sumaTaraver = 0.00

                #  SUMA DE LA COLUMNA PESOS DESTARE
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(Destare) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                sumaDestare1 = cur.fetchone()
                mysql.connection.commit()
                if sumaDestare1[0]:
                    sumaDestarever = round(sumaDestare1[0], 2)
                else:
                    sumaDestarever = 0.00

                #  SUMA DE LA COLUMNA PESOS NETO
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(PesoNeto) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [idver])
                sumaNeto1 = cur.fetchone()
                mysql.connection.commit()
                if sumaNeto1[0]:
                    sumaNetover = round(sumaNeto1[0], 2)
                else:
                    sumaNetover = 0.00

                # HACEMOS LOA SUMA DE CADA COLUMNA PARA LOS DIGITADORES
                #  SUMA DE LA COLUMNA PESOS BRUTOS
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(PesoBruto) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [iddig])
                sumaBruto2 = cur.fetchone()
                if sumaBruto2[0]:
                    sumaBrutodig = round(sumaBruto2[0], 2)
                else:
                    sumaBrutodig = 0.00

                #  SUMA DE LA COLUMNA PESOS TARA
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(PesoTara) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [iddig])
                sumaTara2 = cur.fetchone()
                if sumaTara2[0]:
                    sumaTaradigi = round(sumaTara2[0], 2)
                else:
                    sumaTaradigi = 0.00

                #  SUMA DE LA COLUMNA PESOS DESTARE
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(Destare) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                sumaDestare2 = cur.fetchone()
                mysql.connection.commit()
                if sumaDestare2[0]:
                    sumaDestaredigi = round(sumaDestare2[0], 2)
                else:
                    sumaDestaredigi = 0.00

                #  SUMA DE LA COLUMNA PESOS NETO
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(PesoNeto) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [iddig])
                sumaNeto2 = cur.fetchone()
                mysql.connection.commit()
                if sumaNeto2[0]:
                    sumaNetodigi = round(sumaNeto2[0], 2)
                else:
                    sumaNetodigi = 0.00
                print(pesosverificador)
                print(pesosdigitador)
                return render_template('tablas/tabla-pesos-admin.html', id=ids[0], pesosdigitador=matdigitador, ids=ids, pesosverificador=matverificador, sumaBrutover=sumaBrutover, sumaTaraver=sumaTaraver, sumaDestarever=sumaDestarever, sumaNetover=sumaNetover, sumaBrutodig=sumaBrutodig, sumaTaradigi=sumaTaradigi, sumaDestaredigi=sumaDestaredigi, sumaNetodigi=sumaNetodigi)
            else:
                pesos = ""
                return render_template('tablas/tabla-pesos-admin.html', pesos=pesosdigitador)

    else:
        return "No"

# CARGAR LOS PESOS GENERALES DE LAS VERIFICACIONES


@app.route('/vertara', methods=["POST", "GET"])
def vertara():
    if request.method == "POST":
        id = request.form['id']
        cur = mysql.connection.cursor()
        cur.execute(
            "SELECT dt.Id_DetalleVerificacion,dt.IdVerificacion,m.NombreMaterial,dt.PesoBruto,dt.PesoTara,dt.Destare,dt.PesoNeto FROM tb_detalleverificacion as dt inner join tb_material as m ON dt.IdMaterial = m.Id_Material Where dt.IdVerificacion = %s", [id])
        pesos = cur.fetchall()

        for peso in pesos:
            #  SUMA DE LA COLUMNA PESOS DESTARE
            cur = mysql.connection.cursor()
            cur.execute(
                "SELECT * FROM tb_detalletara WHERE IdDetalleVerificacion = %s", [peso[0]])
            tara = cur.fetchone()
            mysql.connection.commit()
            if tara:
                return pesos[4]
            else:
                return ""


# CARGAR LOS PESOS GENERALES DE LAS VERIFICACIONES
@app.route('/generalPesos', methods=["POST", "GET"])
def generalPesos():
    if request.method == "POST":
        id = request.form['id']
        if id != "":
            # total de materiales
            cur = mysql.connection.cursor()
            cur.execute(
                'SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2)  as neto,round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s Group BY ver.IdMaterial', [id])
            mat = cur.fetchall()
            mysql.connection.commit()
            return render_template('tablas/tabla-mat-general.html', mat=mat)
        else:
            pesos = ""
            return render_template('tablas/tabla-mat-general.html', mat=mat)
    else:
        return "No"

# CARGAR LOS PESOS GENERALES DE LAS VERIFICACIONES


@app.route('/clasificacionPesos', methods=["POST", "GET"])
def clasificacionPesos():
    if request.method == "POST":
        id = request.form['id']
        # TOTAL DE MATERIALES DE PRIMERA
        cur = mysql.connection.cursor()
        cur.execute(
            'SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2) as neto, round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s and m.TipoMaterial = "Primera" Group BY m.TipoMaterial', [id])
        primera = cur.fetchall()
        mysql.connection.commit()

        # TOTAL DE MATERIALES DE SEGUNDA
        cur = mysql.connection.cursor()
        cur.execute(
            'SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2),round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s and m.TipoMaterial = "Segunda" Group BY m.TipoMaterial', [id])
        segunda = cur.fetchall()
        mysql.connection.commit()

        # TOTAL DE RECHAZOS
        cur = mysql.connection.cursor()
        cur.execute("SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2),round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s AND m.NombreMaterial like %s Group BY m.TipoMaterial", (id, 'rechazo%'))
        rechazo = cur.fetchall()
        mysql.connection.commit()

        # TOTAL DE JUMBOS
        cur = mysql.connection.cursor()
        cur.execute('SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2),round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s and m.NombreMaterial like %s Group BY m.TipoMaterial', (id, 'jumbo%'))
        jumbo = cur.fetchall()
        mysql.connection.commit()

        if jumbo:
            print("jumbo tiene")
            jumboNuevo = jumbo[0][3]
        else:
            jumboNuevo = 0.0

        print(rechazo)
        if rechazo:
            rechazoNuevo = rechazo[0][3]
        else:
            rechazoNuevo = 0.0

        # if liquido:
            #     liquidoNuevo = liquido[0][3]
            # else:
            #     liquidoNuevo = 0.0

        print(primera)
        if primera and segunda:
            primeraNuevo = primera[0][3]
            segundaNuevo = segunda[0][3]
        elif primera and not segunda:
            primeraNuevo = primera[0][3]
            segundaNuevo = 0
        elif not primera and segunda:
            primeraNuevo = 0
            segundaNuevo = segunda[0][3]
        else:
            primeraNuevo = 1
            segundaNuevo = 1

        return render_template('tablas/tabla-clasificacion.html', segunda=segunda, primera=primera)
    else:
        return "No"

# BUSCAMOS MATERIALES


@app.route('/buscarMaterial', methods=["POST", "GET"])
def buscarMaterial():
    if request.method == "POST":
        material = request.form['material']
        cur = mysql.connection.cursor()
        cur.execute(
            "select * from tb_material Where Id_Estado = 1 AND NombreMaterial like %s limit 6", ['%'+material+'%'])
        materiales = cur.fetchall()
        print(materiales)
        return render_template('otros/material-busqueda.html', materiales=materiales)
    else:
        return "No"

# INSERTAMOS LOS PESOS DEL MATERIALE SELECCIONADO


@app.route('/insertarPesos', methods=["POST", "GET"])
def insertarPesos():
    if request.method == "POST":
        id = request.form['id']
        material = request.form['material']
        destare = request.form['destare']
        pBruto = request.form['pBruto']
        pTara = request.form['pTara']

        # MANDAMOS A LLAMAR EL ID DEL MATERIAL QUE SELECCIONO EL USUARIO
        cur = mysql.connection.cursor()
        cur.execute(
            "select * from tb_material Where Id_Estado = 1 AND NombreMaterial = %s", [material])
        materiales = cur.fetchone()

        # CREAMOS EL DETALLE VERIFICACION AÑADIENDO LOS PESOS QUE EL USUARIO INGRESO
        pNeto = round(float(float(pBruto)-float(pTara)-float(destare)), 2)

        cur = mysql.connection.cursor()
        cur.execute("INSERT INTO tb_detalleverificacion (IdVerificacion,IdMaterial,PesoBruto,PesoTara,PesoNeto,Destare) VALUES (%s,%s,%s,%s,%s,%s)",
                    (id, materiales[0], pBruto, pTara, pNeto, destare))
        mysql.connection.commit()
        # MANDAMOS A LLAMAR TODA LA TABLA DE DETALLE VERIFICACION CON LOS NUEVOS DATOS REGISTRADOS
        cur = mysql.connection.cursor()
        cur.execute(
            "SELECT dt.Id_DetalleVerificacion,dt.IdVerificacion,m.NombreMaterial,dt.PesoBruto,dt.PesoTara,dt.Destare,dt.PesoNeto FROM tb_detalleverificacion as dt inner join tb_material as m ON dt.IdMaterial = m.Id_Material Where dt.IdVerificacion = %s", [id])
        pesos = cur.fetchall()
        mysql.connection.commit()
        # HACEMOS LOA SUMA DE CADA COLUMNA
        #  SUMA DE LA COLUMNA PESOS BRUTOS
        cur = mysql.connection.cursor()
        cur.execute(
            "SELECT SUM(PesoBruto) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
        sumaBruto1 = cur.fetchone()
        mysql.connection.commit()
        sumaBruto = round(sumaBruto1[0], 2)
        #  SUMA DE LA COLUMNA PESOS TARA
        cur = mysql.connection.cursor()
        cur.execute(
            "SELECT SUM(PesoTara) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
        sumaTara1 = cur.fetchone()
        sumaTara = round(sumaTara1[0], 2)
        #  SUMA DE LA COLUMNA PESOS DESTARE
        cur = mysql.connection.cursor()
        cur.execute(
            "SELECT SUM(Destare) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
        sumaDestare1 = cur.fetchone()
        mysql.connection.commit()
        sumaDestare = round(sumaDestare1[0], 2)
        #  SUMA DE LA COLUMNA PESOS NETO
        cur = mysql.connection.cursor()
        cur.execute(
            "SELECT SUM(PesoNeto) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
        sumaNeto1 = cur.fetchone()
        mysql.connection.commit()
        sumaNeto = round(sumaNeto1[0], 2)
        print(sumaNeto)
        # total de materiales
        cur = mysql.connection.cursor()
        cur.execute(
            'SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2) as neto,round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s Group BY ver.IdMaterial', [id])
        mat = cur.fetchall()
        mysql.connection.commit()
        print("material:", mat)
        return render_template('tablas/tabla-pesos.html', mat=mat, pesos=pesos, sumaBruto=sumaBruto, sumaTara=sumaTara, sumaDestare=sumaDestare, sumaNeto=sumaNeto)
    else:
        return "No"

# FINALIZAR VERIFICACION


@app.route('/finalizarVerificacion', methods=["POST", "GET"])
def finalizarVerificacion():
    if request.method == "POST":
        print(session['cargo'])
        if not session['cargo']:
            return "ErrorRed"
        if session['cargo'] != 5:

            if session['cargo'] == 1:
                # ES DIGITADOR
                app.logger.error('COMIENZO DE LA FUNCION VERIFICACION')
                app.logger.error('Cargo: %s', session['cargo'])
                app.logger.error('Uid: %s',session['uid'])
                app.logger.error('userId: %s',session['userId'])
                app.logger.error('Contra: %s',session['pass'])
                id = request.form['id']
                app.logger.info('ID: %s',id)
                print("IDDDDDDDDDD")
                print(id)
                # MANDAMOS A LLAMAR TODA LA TABLA DE DETALLE VERIFICACION CON LOS NUEVOS DATOS REGISTRADOS
                try:
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "SELECT dt.Id_DetalleVerificacion,dt.IdVerificacion,m.NombreMaterial,dt.PesoBruto,dt.PesoTara,dt.Destare,dt.PesoNeto FROM tb_detalleverificacion as dt inner join tb_material as m ON dt.IdMaterial = m.Id_Material Where dt.IdVerificacion = %s", [id])
                    pesos = cur.fetchall()
                    mysql.connection.commit()
                except Exception as e:
                    app.logger.error('Error en la aplicación:', exc_info=e)
                app.logger.error('PESOS: %s',pesos)
                #print("pesos")
                
                if pesos:
                    
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado, p.NombreProveedor,digi.NombreUsuario as digitador,veri.NombreUsuario as verificador,v.Bahia,p.IdOddo,pc.Id_PuntoCompra from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_usuarios as digi on v.IdDigitador = digi.Id_Usuario inner join tb_usuarios as veri on v.IdVerificador = veri.Id_Usuario Where v.IdEstado = 3 AND v.Id_Verificacion = %s", [id])
                    Verificacion = cur.fetchone()
                    mysql.connection.commit()
                    app.logger.info('Verificacion: %s', Verificacion)
                    if Verificacion:
                        try:
                            # HACEMOS LOA SUMA DE CADA COLUMNA
                            #  SUMA DE LA COLUMNA PESOS BRUTOS
                            cur = mysql.connection.cursor()
                            cur.execute(
                                "SELECT SUM(PesoBruto) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                            sumaBruto1 = cur.fetchone()
                            sumaBruto = round(sumaBruto1[0], 2)

                            #  SUMA DE LA COLUMNA PESOS TARA
                            cur = mysql.connection.cursor()
                            cur.execute(
                                "SELECT SUM(PesoTara) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                            sumaTara1 = cur.fetchone()
                            sumaTara = round(sumaTara1[0], 2)
                            #  SUMA DE LA COLUMNA PESOS DESTARE
                            cur = mysql.connection.cursor()
                            cur.execute(
                                "SELECT SUM(Destare) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                            sumaDestare1 = cur.fetchone()
                            mysql.connection.commit()
                            if sumaDestare1:
                                sumaDestare = round(sumaDestare1[0], 2)
                            else:
                                sumaDestare = 0
                            #  SUMA DE LA COLUMNA PESOS NETO
                            cur = mysql.connection.cursor()
                            cur.execute(
                                "SELECT SUM(PesoNeto) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                            sumaNeto1 = cur.fetchone()
                            mysql.connection.commit()
                            sumaNeto = round(sumaNeto1[0], 2)
                            #  FECHA VERIFICACION
                            cur = mysql.connection.cursor()
                            cur.execute(
                                "SELECT Fecha FROM tb_verificacion WHERE Id_Verificacion = %s", [id])
                            fecha2 = cur.fetchone()
                            mysql.connection.commit()
                            fecha = capturarHora()
                            fechacreacion = datetime.date(fecha)
                            # Usuario que lo creó
                            cur = mysql.connection.cursor()
                            cur.execute(
                                "SELECT u.NombreUsuario FROM tb_verificacion as v inner join tb_usuarios as u ON v.IdUsuarioCreacion = u.Id_Usuario WHERE Id_Verificacion = %s", [id])
                            usuario = cur.fetchone()
                            mysql.connection.commit()
                            # # Cambiar el estado de la verificacion ESTADO DE LA VERIFICACION 
                            # CAMBIARLE EL ESTADO SEGUN EL PUNTO DE COMPRA
                            cur = mysql.connection.cursor()
                            cur.execute(
                                'Update tb_verificacion set IdEstado = 8 Where Id_Verificacion = %s', [id])
                            mysql.connection.commit()
                            app.logger.info('CAMBIO DE ESTADO')
                            # =========================================
                            #total de materiales RESUMEN

                            # total de materiales
                            cur = mysql.connection.cursor()
                            cur.execute(
                                'SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2) as neto,round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s Group BY ver.IdMaterial', [id])
                            mat = cur.fetchall()
                            mysql.connection.commit()
                            app.logger.info('Mat: %s', mat)

                            # TOTAL DE MATERIALES DE PRIMERA
                            cur = mysql.connection.cursor()
                            cur.execute(
                                'SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2) as neto, round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s and m.TipoMaterial = "Primera" Group BY m.TipoMaterial', [id])
                            primera = cur.fetchall()
                            mysql.connection.commit()
                            app.logger.info('primera: %s', primera)

                            # TOTAL DE MATERIALES DE SEGUNDA
                            cur = mysql.connection.cursor()
                            cur.execute(
                                'SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2),round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s and m.TipoMaterial = "Segunda" Group BY m.TipoMaterial', [id])
                            segunda = cur.fetchall()
                            mysql.connection.commit()
                            app.logger.info('Segunda: %s', segunda)

                            # TOTAL DE RECHAZOS
                            cur = mysql.connection.cursor()
                            cur.execute("SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2),round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s AND m.NombreMaterial like %s Group BY m.TipoMaterial", (id, 'rechazo%'))
                            rechazo = cur.fetchall()
                            mysql.connection.commit()
                            app.logger.info('Rechazo: %s', rechazo)

                            # TOTAL DE JUMBOS
                            cur = mysql.connection.cursor()
                            cur.execute('SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2),round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s and m.NombreMaterial like %s Group BY m.TipoMaterial', (id, 'jumbo%'))
                            jumbo = cur.fetchall()
                            mysql.connection.commit()
                            app.logger.info('Jumbo: %s', jumbo)

                            # TOTAL DE DEVOLUCION PET
                            cur = mysql.connection.cursor()
                            cur.execute('SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2),round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s and m.NombreMaterial like %s Group BY m.TipoMaterial', (id, 'devolucion material%'))
                            devolucion = cur.fetchall()
                            mysql.connection.commit()
                            app.logger.info('Devolucion: %s', devolucion)

                            #BATERIAS TOTALES
                            cur = mysql.connection.cursor()
                            cur.execute('SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2),round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s and m.NombreMaterial like %s Group BY m.TipoMaterial', (id, 'BATERIA%'))
                            bateria = cur.fetchall()
                            mysql.connection.commit()
                            app.logger.info('Bateria: %s', bateria)

                            # TOTAL LIQUIDO
                            # cur = mysql.connection.cursor()
                            # cur.execute('SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2),round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s and m.NombreMaterial like %s Group BY m.TipoMaterial',(id,'liquido%'))
                            # liquido = cur.fetchall()
                            # mysql.connection.commit()

                            # TOTAL rechazo Pet
                            # cur = mysql.connection.cursor()
                            # cur.execute('SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2),round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s m.NombreMaterial = "LIQUIDO" Group BY m.TipoMaterial',[id])
                            # rechazoPet = cur.fetchall()
                            # mysql.connection.commit()

                            # AQUI SE DEBERIA DE MANDAR A LLAMAR LA FUNCION PARA GENERAR LA ORDEN DE COMPRA EN ODDO
                            # CrearOrdenCompra(proveedorId,puntoCompra,NoBoleta,rechazo,jumbo,liquido,rechazoPet,primera,segunda)
                            # print(jumbo[0][3])
                            if jumbo:
                                #print("jumbo tiene")
                                jumboNuevo = jumbo[0][3]
                            else:
                                jumboNuevo = 0.0

                            bateriaflag = 0
                            if bateria:
                                #print("jumbo tiene")
                                if bateria[0][3] > 200:
                                    bateriaflag = 1
                            else:
                                bateriaflag = 0
                            
                            print("baterriaaaaa:",bateriaflag)
                            app.logger.info('Bateria flag: %s', bateriaflag)
                            # print(rechazo)
                            if rechazo:
                                rechazoNuevo = rechazo[0][3]
                            else:
                                rechazoNuevo = 0.0

                            #print(devolucion)
                            if devolucion:
                                devolucionNuevo = devolucion[0][3]
                            else:
                                devolucionNuevo = 0.0

                            # if liquido:
                            #     liquidoNuevo = liquido[0][3]
                            # else:
                            #     liquidoNuevo = 0.0

                            # print(primera)
                            if primera and segunda:
                                primeraNuevo = primera[0][3]
                                segundaNuevo = segunda[0][3]
                            elif primera and not segunda:
                                primeraNuevo = primera[0][3]
                                segundaNuevo = 0
                            elif not primera and segunda:
                                primeraNuevo = 0
                                segundaNuevo = segunda[0][3]
                            else:
                                primeraNuevo = 1
                                segundaNuevo = 1

                            # print(Verificacion)
                            # print(Verificacion[0][2])
                            # print(Verificacion[0][3])
                            # MANDAR A TRAER LOS MATERIALES DEL VERIFICADOR PARA VER SI SON IGUALES Y GENERAR LA ORDEN DE COMPRA

                        

                            # TENGO LOA MATES DEL DIGITADOR EN LA LISTA mat y los del verificador en matverificador
                            # VERIFICAMOS SI LAS DOS LISTAS DE MATERIALES SON IGUALES

                            # #total de materiales del digitador
                            # cur = mysql.connection.cursor()
                            # cur.execute('SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2) as neto FROM tb_detalleverificacion as ver inner join tb_verificacion as v on ver.IdVerificacion = v.Id_Verificacion inner join tb_material as m ON ver.IdMaterial = m.Id_Material inner join tb_usuarios as u on v.IdUsuarioCreacion = u.Id_Usuario Where v.PO = %s and u.IdCargo = %s and v.IdEstado = 5 Group BY ver.IdMaterial',(id,1))
                            # matdigitador = cur.fetchall()
                            # mysql.connection.commit()
                            # print(matdigitador)
                            # print("MAT DEL digitador")

                            # ======================================================================================================
                            # ======================================================================================================
                            # IdOrden = conexion.CrearOrdenCompra(Verificacion[0][10],Verificacion[0][11],Verificacion[0][3],rechazoNuevo,jumboNuevo,0,0,primeraNuevo,segundaNuevo,session['uid'],session['pass'])

                            # print("ORDEN AQUI")
                            # print(IdOrden)
                            # for material in mat:
                            #     print(material)
                            #     if material[0] == "Rechazo (cobre)" or material[0] == "RECHAZO" or material[0] == "JUMBO" or material[0] == "Rechazo (Aluminio)" or material[0] == "Rechazo (Acero)" or material[0] == "Rechazo (Bronce)" or material[0] == "Rechazo (Cable)" or material[0] == "Rechazo (lata)" :

                            #         print("sosretroll")
                            #     else:
                            #         conexion.IngresarMaterialOrdenCompra(material[0],material[3],IdOrden,session['uid'],session['pass'])
                            # po = conexion.traerPo(IdOrden)
                            # TRAER LA VERIFICACION QUE REALIZO EL VERIFICADOR PARA CAMBIARLE LA PO
                            # cur = mysql.connection.cursor()
                            # cur.execute(
                            #     "SELECT NoEnlace from tb_verificacion Where Id_Verificacion = %s", [id])
                            # nboletaver = cur.fetchone()
                            # mysql.connection.commit()

                            # cur = mysql.connection.cursor()
                            # cur.execute("SELECT Id_Verificacion from tb_verificacion Where Id_Verificacion = %s",[id])
                            # nboletaver = cur.fetchall()
                            # mysql.connection.commit()
                            # print("NOBOLETA: ", Verificacion[0][3])
                            # print('punto de compra: ',Verificacion[0][4])
                            cantBol = 0
                            contadorWhile = 0
                            app.logger.info('WHILE')
                            app.logger.info('Boleta: %s', Verificacion[3])
                            while cantBol != 2:
                                # SELECCIONAMOS CUANTAS VERIFICACIONES TIENEN ESE NUMERO DE BOLETA
                                cur = mysql.connection.cursor()
                                cur.execute("SELECT Count(NoBoleta) from tb_verificacion Where NoBoleta = %s and IdDigitador = %s and IdEstado = 8", (
                                    Verificacion[3], session['userId']))
                                cantidadBoleta = cur.fetchone()
                                mysql.connection.commit()
                                print("cantidad de verificaciones: ",
                                    cantidadBoleta[0])
                                app.logger.info('Cantidad Verificaciones: %s', cantidadBoleta[0])
                                
                                cantBol = cantidadBoleta[0]
                                time.sleep(1)
                                if contadorWhile == 5:

                                    cur = mysql.connection.cursor()

                                    cur.execute('Update tb_verificacion set IdEstado = 7 where NoBoleta = %s and IdEstado = 8',[ Verificacion[3]])
                                    mysql.connection.commit()
                                    app.logger.info('SE ACABO EL TIEMPO')
                                    return "ValoresErroneos"
                                contadorWhile += 1
                            app.logger.info('SALIO DEL WHILE')
                            
                            
                            

                            # total de materiales del verificador
                            cur = mysql.connection.cursor()
                            cur.execute('SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2) as neto,round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_verificacion as v on ver.IdVerificacion = v.Id_Verificacion inner join tb_material as m ON ver.IdMaterial = m.Id_Material inner join tb_usuarios as u on v.IdUsuarioCreacion = u.Id_Usuario Where v.PO = %s and v.NoBoleta = %s and u.IdCargo = %s and v.IdEstado = 8 Group BY ver.IdMaterial', (
                                Verificacion[2], Verificacion[3], 2))
                            matverificador = cur.fetchall()
                            mysql.connection.commit()
                            print(matverificador)
                            print("MAT DEL VERIFICADOR")
                            print(mat)
                            print("MAT DEL DIGITADOR")
                            iguales = 0
                            app.logger.info('OBTUVIMOS AMBOS MATERIALES')
                            if mat == matverificador:
                                iguales = 1
                                if session['punto'] == 'CASETA: Recepciones':
                                    try:
                                        IdOrden = conexion.CrearOrdenCompra(Verificacion[10], Verificacion[11], Verificacion[3],
                                                                    rechazoNuevo, jumboNuevo,devolucionNuevo, 0, 0, 1, 1, session['uid'], session['pass'],'',sumaDestare)
                                        app.logger.info('ORDENDE COMPRA CASETA')
                                        #FUNCION PARA MANDAR A ESCRIBIR LA HORA QUE SE FINALIZA LA CREACION DE LA PO
                                        hi = capturarHora()
                                        fechacreacion = hi.replace(microsecond=0)
                                        cur = mysql.connection.cursor()
                                        cur.execute(
                                                "INSERT INTO tb_conciliacion (IdVerificacion,FechaConciliacion,NoBoleta,IdPuntoCompra) VALUES (%s,%s,%s,%s)", (Verificacion[0], fechacreacion,Verificacion[3],Verificacion[11]))

                                        mysql.connection.commit()

                                    except:
                                        cur = mysql.connection.cursor()
                                        cur.execute(
                                            'Update tb_verificacion set IdEstado = 3 Where NoBoleta = %s and IdEstado = 8', (Verificacion[3]))
                                        mysql.connection.commit()
                                        return "ErrorRed"
                                else:
                                    cur = mysql.connection.cursor()
                                    cur.execute(
                                        "SELECT IdJefeCuadrilla from tb_verificacion Where Id_Verificacion = %s", [id])
                                    jefe = cur.fetchone()
                                    mysql.connection.commit()
                                    print(jefe)
                                    print("JEFE: ",jefe[0])
                                    if jefe[0] != 2:
                                        print("ENTRO IGUAL 2")
                                        try:
                                            IdOrden = conexion.CrearOrdenCompra(Verificacion[10], Verificacion[11], Verificacion[3],
                                                                    rechazoNuevo, jumboNuevo,devolucionNuevo, 0, 0, primeraNuevo, segundaNuevo, session['uid'], session['pass'],jefe[0],sumaDestare)
                                            app.logger.info('ORDENDE COMPRA PLANTA CON JEFE DE CUADRILLA')
                                            #FUNCION PARA MANDAR A ESCRIBIR LA HORA QUE SE FINALIZA LA CREACION DE LA PO
                                            hi = capturarHora()
                                            fechacreacion = hi.replace(microsecond=0)
                                            cur = mysql.connection.cursor()
                                            cur.execute(
                                                "INSERT INTO tb_conciliacion (IdVerificacion,FechaConciliacion,NoBoleta,IdPuntoCompra) VALUES (%s,%s,%s,%s)", (Verificacion[0], fechacreacion,Verificacion[3],Verificacion[11]))

                                            mysql.connection.commit()
                                        except:
                                            cur = mysql.connection.cursor()
                                            cur.execute(
                                                'Update tb_verificacion set IdEstado = 3 Where NoBoleta = %s and IdEstado = 8', (Verificacion[3]))
                                            mysql.connection.commit()
                                            return "ErrorRed"
                                    else:
                                        print("ENTRO DIFERENTE 2")
                                        try:
                                            IdOrden = conexion.CrearOrdenCompra(Verificacion[10], Verificacion[11], Verificacion[3],
                                                                    rechazoNuevo, jumboNuevo,devolucionNuevo, 0, 0, 1, 1, session['uid'], session['pass'],"",sumaDestare)
                                            app.logger.info('ORDENDE COMPRA PLANTA SIN JEFE')
                                            #FUNCION PARA MANDAR A ESCRIBIR LA HORA QUE SE FINALIZA LA CREACION DE LA PO
                                            hi = capturarHora()
                                            fechacreacion = hi.replace(microsecond=0)
                                            cur = mysql.connection.cursor()
                                            cur.execute(
                                                "INSERT INTO tb_conciliacion (IdVerificacion,FechaConciliacion,NoBoleta,IdPuntoCompra) VALUES (%s,%s,%s,%s)", (Verificacion[0], fechacreacion,Verificacion[3],Verificacion[11]))

                                            mysql.connection.commit()
                                        except:
                                            cur = mysql.connection.cursor()
                                            cur.execute(
                                                'Update tb_verificacion set IdEstado = 3 Where NoBoleta = %s and IdEstado = 8', (Verificacion[3]))
                                            mysql.connection.commit()
                                            return "ErrorRed"
                                print("ORDEN AQUI")
                                print(IdOrden)
                                app.logger.info('ORDEN: %s',IdOrden)
                                banderaValidador = 0
                                app.logger.info('INGRESAR MATERIALES A LA ORDEN')
                                for material in mat:

                                    print("Cada Material: ",material)
                                    if material[0] == "Rechazo (cobre)" or material[0] == "RECHAZO" or material[0] == "JUMBO" or material[0] == "Rechazo (Aluminio)" or material[0] == "Rechazo (Lata)" or material[0] == "Rechazo (Acero)" or material[0] == "Rechazo (Bronce)" or material[0] == "Rechazo (Cable)" or material[0] == "Rechazo (lata)" or material[0] == "DEVOLUCION MATERIAL":

                                        print("sosretroll")
                                    else:
                                        print('Material: ',material[0])
                                        print('valor bruto: ',material[3])
                                        app.logger.info('Material: %s',material[0])
                                        app.logger.info('Bruto: %s',material[3])
                                        if bateriaflag == 1:
                                            print('AQUI PASA POR EL VALIDADOR')
                                            app.logger.info('PASA POR EL VALIDADOR')
                                            conexion.IngresarMaterialOrdenCompra(
                                                material[0], material[3], IdOrden, session['uid'], session['pass'])
                                            banderaValidador = 1
                                            app.logger.info('PASA POR EL VALIDADOR')

                                        else:
                                            conexion.IngresarMaterialOrdenCompra(
                                                material[0], material[3], IdOrden, session['uid'], session['pass'])
                                            app.logger.info('NO PASA POR EL VALIDADOR')
                                app.logger.info('CREACION DEL ALBARAS')
                                conexion.CrearAlbaran(IdOrden, session['uid'], session['pass'])
                                        
                                po = conexion.traerPo(IdOrden)
                                app.logger.info('MODIFICAMOS EL ESTADO YA QUE SON IGUALES')
                                cur = mysql.connection.cursor()
                                cur.execute(
                                    'Update tb_verificacion set PO = %s Where NoBoleta = %s and IdEstado = 8', (po[0]['name'],Verificacion[3]))
                                mysql.connection.commit()
                                print("SON IGUALES VALOR: ", iguales)
                                # AQUI SE CAMBIA EL ESTADO SEGUN SU PUNTO DE COMPRA
                                # YA QUE SON CORRECTAS SE LES CAMBIA EL ESTADO
                                # MANDAMOS A LLAMAR EL PUNTO DE COMPRA PARA VALIDAR EL ESTADO
                                
                                if Verificacion[4] == "CASETA: Recepciones" and banderaValidador == 0:
                                    app.logger.info('CASETA SIN VALIDADOR')
                                    cur = mysql.connection.cursor()
                                    cur.execute('Update tb_verificacion set IdEstado = 4 Where NoBoleta = %s and IdEstado = 8 and PO = %s', (Verificacion[3],po[0]['name']))
                                    mysql.connection.commit()
                                    print('entro a caseta')
                                else:
                                    app.logger.info('CASETA CON VALIDADOR Y PLANTA')
                                    cur = mysql.connection.cursor()
                                    cur.execute('Update tb_verificacion set IdEstado = 5 Where NoBoleta = %s and IdEstado = 8 and PO = %s', (Verificacion[3],po[0]['name']))
                                    mysql.connection.commit()
                                # =================================================
                                app.logger.info('==================================================================================================')
                                return "Iguales"
                            else:
                                print('aqui cambiamos')
                                app.logger.info('DIFERENTES')
                                #AQUI ES QUE SE BLOQUEAN
                                cur = mysql.connection.cursor()
                                cur.execute('Update tb_verificacion set IdEstado = 7 Where NoBoleta = %s and IdEstado = 8', [Verificacion[3]])
                                mysql.connection.commit()
                                
                                return "Diferentes"

                            # cur = mysql.connection.cursor()
                            # cur.execute('Update tb_verificacion set PO = %s Where Id_Verificacion = %s',(po[0]['name'],id))
                            # mysql.connection.commit()
                            # return "DOne"
                            # return render_template('otros/factura.html',po = po[0]['name'],segunda = segunda,primera = primera,mat = mat,id = id,usuario = usuario,verificacion = Verificacion, fechaEmision = fecha,fechaCreacion = fechacreacion,pesos = pesos,sumaBruto = sumaBruto,sumaTara = sumaTara,sumaDestare = sumaDestare,sumaNeto = sumaNeto)
                        except Exception as e:
                            cur = mysql.connection.cursor()
                            cur.execute(
                                'Update tb_verificacion set IdEstado = 3 Where NoBoleta = %s and IdEstado = 8', (Verificacion[3]))
                            mysql.connection.commit()
                            return str(e),500
                    else:
                        return "ErrorrRed"
                else:
                    print("vacio")
                    return "vacio"
                # elif session['cargo'] == 3:
                #     #ES ADMINISTRADOR
                #     id = request.form['id']
                #     print(id)
                #     #MANDAMOS A LLAMAR TODA LA TABLA DE DETALLE VERIFICACION CON LOS NUEVOS DATOS REGISTRADOS
                #     cur = mysql.connection.cursor()
                #     cur.execute("SELECT dt.Id_DetalleVerificacion,dt.IdVerificacion,m.NombreMaterial,dt.PesoBruto,dt.PesoTara,dt.Destare,dt.PesoNeto FROM tb_detalleverificacion as dt inner join tb_material as m ON dt.IdMaterial = m.Id_Material Where dt.IdVerificacion = %s",[id])
                #     pesos = cur.fetchall()
                #     mysql.connection.commit()
                #     print("pesos")
                #     if pesos:
                #         cur = mysql.connection.cursor()
                #         cur.execute("select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado, p.NombreProveedor,digi.NombreUsuario as digitador,veri.NombreUsuario as verificador,v.Bahia,p.IdOddo,pc.Id_PuntoCompra from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_usuarios as digi on v.IdDigitador = digi.Id_Usuario inner join tb_usuarios as veri on v.IdVerificador = veri.Id_Usuario Where v.IdEstado = 3 AND v.Id_Verificacion = %s",[id])
                #         Verificacion = cur.fetchall()
                #         mysql.connection.commit()
                #         #HACEMOS LOA SUMA DE CADA COLUMNA
                #         #  SUMA DE LA COLUMNA PESOS BRUTOS
                #         cur = mysql.connection.cursor()
                #         cur.execute("SELECT SUM(PesoBruto) FROM tb_detalleverificacion WHERE IdVerificacion = %s",[id])
                #         sumaBruto1 = cur.fetchone()
                #         sumaBruto = round(sumaBruto1[0],2)
                #         #  SUMA DE LA COLUMNA PESOS TARA
                #         cur = mysql.connection.cursor()
                #         cur.execute("SELECT SUM(PesoTara) FROM tb_detalleverificacion WHERE IdVerificacion = %s",[id])
                #         sumaTara1 = cur.fetchone()
                #         sumaTara = round(sumaTara1[0],2)
                #         #  SUMA DE LA COLUMNA PESOS DESTARE
                #         cur = mysql.connection.cursor()
                #         cur.execute("SELECT SUM(Destare) FROM tb_detalleverificacion WHERE IdVerificacion = %s",[id])
                #         sumaDestare1 = cur.fetchone()
                #         mysql.connection.commit()
                #         sumaDestare = round(sumaDestare1[0],2)
                #         #  SUMA DE LA COLUMNA PESOS NETO
                #         cur = mysql.connection.cursor()
                #         cur.execute("SELECT SUM(PesoNeto) FROM tb_detalleverificacion WHERE IdVerificacion = %s",[id])
                #         sumaNeto1 = cur.fetchone()
                #         mysql.connection.commit()
                #         sumaNeto = round(sumaNeto1[0],2)
                #         #  FECHA VERIFICACION
                #         cur = mysql.connection.cursor()
                #         cur.execute("SELECT Fecha FROM tb_verificacion WHERE Id_Verificacion = %s",[id])
                #         fecha = cur.fetchone()
                #         mysql.connection.commit()
                #         fecha = capturarHora()
                #         fechacreacion = datetime.date(fecha)
                #         #Usuario que lo creó
                #         cur = mysql.connection.cursor()
                #         cur.execute("SELECT u.NombreUsuario FROM `tb_verificacion` as v inner join tb_usuarios as u ON v.IdUsuarioCreacion = u.Id_Usuario WHERE Id_Verificacion = %s",[id])
                #         usuario = cur.fetchone()
                #         mysql.connection.commit()
                #         #Cambiar el estado de la verificacion
                #         cur = mysql.connection.cursor()
                #         cur.execute('Update tb_verificacion set IdEstado = 5 Where Id_Verificacion = %s',[id])
                #         mysql.connection.commit()
                #         #total de materiales RESUMEN

                #         #total de materiales
                #         cur = mysql.connection.cursor()
                #         cur.execute('SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2) as neto,round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s Group BY ver.IdMaterial',[id])
                #         mat = cur.fetchall()
                #         mysql.connection.commit()

                #         #TOTAL DE MATERIALES DE PRIMERA
                #         cur = mysql.connection.cursor()
                #         cur.execute('SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2) as neto, round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s and m.TipoMaterial = "Primera" Group BY m.TipoMaterial',[id])
                #         primera = cur.fetchall()
                #         mysql.connection.commit()

                #         #TOTAL DE MATERIALES DE SEGUNDA
                #         cur = mysql.connection.cursor()
                #         cur.execute('SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2),round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s and m.TipoMaterial = "Segunda" Group BY m.TipoMaterial',[id])
                #         segunda = cur.fetchall()
                #         mysql.connection.commit()

                #         #TOTAL DE RECHAZOS
                #         cur = mysql.connection.cursor()
                #         cur.execute("SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2),round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s AND m.NombreMaterial like %s Group BY m.TipoMaterial",(id,'rechazo%'))
                #         rechazo = cur.fetchall()
                #         mysql.connection.commit()

                #         #TOTAL DE JUMBOS
                #         cur = mysql.connection.cursor()
                #         cur.execute('SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2),round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s and m.NombreMaterial like %s Group BY m.TipoMaterial',(id,'jumbo%'))
                #         jumbo = cur.fetchall()
                #         mysql.connection.commit()

                #         #TOTAL LIQUIDO
                #         # cur = mysql.connection.cursor()
                #         # cur.execute('SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2),round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s and m.NombreMaterial like %s Group BY m.TipoMaterial',(id,'liquido%'))
                #         # liquido = cur.fetchall()
                #         # mysql.connection.commit()

                #         #TOTAL rechazo Pet
                #         # cur = mysql.connection.cursor()
                #         # cur.execute('SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2),round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s m.NombreMaterial = "LIQUIDO" Group BY m.TipoMaterial',[id])
                #         # rechazoPet = cur.fetchall()
                #         # mysql.connection.commit()

                #         #AQUI SE DEBERIA DE MANDAR A LLAMAR LA FUNCION PARA GENERAR LA ORDEN DE COMPRA EN ODDO
                #         #CrearOrdenCompra(proveedorId,puntoCompra,NoBoleta,rechazo,jumbo,liquido,rechazoPet,primera,segunda)
                #         #print(jumbo[0][3])
                #         if jumbo:
                #             print("jumbo tiene")
                #             jumboNuevo = jumbo[0][3]
                #         else:
                #             jumboNuevo = 0.0

                #         print(rechazo)
                #         if rechazo:
                #             rechazoNuevo = rechazo[0][3]
                #         else:
                #             rechazoNuevo = 0.0

                #         # if liquido:
                #         #     liquidoNuevo = liquido[0][3]
                #         # else:
                #         #     liquidoNuevo = 0.0

                #         print(primera)
                #         if primera and segunda:
                #             primeraNuevo = primera[0][3]
                #             segundaNuevo = segunda[0][3]
                #         elif primera and not segunda:
                #             primeraNuevo = primera[0][3]
                #             segundaNuevo = 0
                #         elif not primera and segunda:
                #             primeraNuevo = 0
                #             segundaNuevo = segunda[0][3]
                #         else:
                #             primeraNuevo = 1
                #             segundaNuevo = 1
                #         print(Verificacion)
                #         IdOrden = conexion.CrearOrdenCompra(Verificacion[0][10],Verificacion[0][11],Verificacion[0][3],rechazoNuevo,jumboNuevo,0,0,primeraNuevo,segundaNuevo,session['uid'],session['pass'])

                #         print("ORDEN AQUI")
                #         print(IdOrden)
                #         for material in mat:
                #             print(material)
                #             if material[0] == "Rechazo (cobre)" or material[0] == "RECHAZO" or material[0] == "JUMBO" or material[0] == "Rechazo (Aluminio)" or material[0] == "Rechazo (Acero)" or material[0] == "Rechazo (Bronce)" or material[0] == "Rechazo (Cable)" or material[0] == "Rechazo (lata)" :

                #                 print("sosretroll")
                #             else:
                #                 conexion.IngresarMaterialOrdenCompra(material[0],material[3],IdOrden,session['uid'],session['pass'])
                #         po = conexion.traerPo(IdOrden)
                #         #TRAER LA VERIFICACION QUE REALIZO EL VERIFICADOR PARA CAMBIARLE LA PO
                #         cur = mysql.connection.cursor()
                #         cur.execute("SELECT NoBoleta from tb_verificacion Where Id_Verificacion = %s",[id])
                #         nboletaver = cur.fetchone()
                #         mysql.connection.commit()

                #         # cur = mysql.connection.cursor()
                #         # cur.execute("SELECT Id_Verificacion from tb_verificacion Where Id_Verificacion = %s",[id])
                #         # nboletaver = cur.fetchall()
                #         # mysql.connection.commit()
                #         print("NOBOLETA: ",nboletaver[0])

                #         cantBol = 0
                #         while cantBol != 2:
                #             # SELECCIONAMOS CUANTAS VERIFICACIONES TIENEN ESE NUMERO DE BOLETA
                #             cur = mysql.connection.cursor()
                #             cur.execute("SELECT count(NoBoleta) from tb_verificacion Where NoBoleta = %s and IdDigitador = %s",(nboletaver[0],session['userId']))
                #             cantidadBoleta = cur.fetchone()
                #             mysql.connection.commit()
                #             print("cantidad de verificaciones: ",cantidadBoleta[0])
                #             cantBol = cantidadBoleta[0]
                #             time.sleep(1)

                #         cur = mysql.connection.cursor()
                #         cur.execute('Update tb_verificacion set PO = %s Where NoBoleta = %s',(po[0]['name'],nboletaver[0]))
                #         mysql.connection.commit()

                #         # cur = mysql.connection.cursor()
                #         # cur.execute('Update tb_verificacion set PO = %s Where Id_Verificacion = %s',(po[0]['name'],id))
                #         # mysql.connection.commit()

                #         return render_template('otros/factura.html',po = po[0]['name'],segunda = segunda,primera = primera,mat = mat,id = id,usuario = usuario,verificacion = Verificacion, fechaEmision = fecha,fechaCreacion = fechacreacion,pesos = pesos,sumaBruto = sumaBruto,sumaTara = sumaTara,sumaDestare = sumaDestare,sumaNeto = sumaNeto)
                #     else:
                #         print("vacio")
                #         return "vacio"

            else:
                id = request.form['id']

                # MANDAMOS A LLAMAR TODA LA TABLA DE DETALLE VERIFICACION CON LOS NUEVOS DATOS REGISTRADOS
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT dt.Id_DetalleVerificacion,dt.IdVerificacion,m.NombreMaterial,dt.PesoBruto,dt.PesoTara,dt.Destare,dt.PesoNeto FROM tb_detalleverificacion as dt inner join tb_material as m ON dt.IdMaterial = m.Id_Material Where dt.IdVerificacion = %s", [id])
                pesos = cur.fetchall()
                mysql.connection.commit()
                print("pesos")
                if pesos:
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado, p.NombreProveedor,digi.NombreUsuario as digitador,veri.NombreUsuario as verificador,v.Bahia,p.IdOddo,pc.Id_PuntoCompra from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_usuarios as digi on v.IdDigitador = digi.Id_Usuario inner join tb_usuarios as veri on v.IdVerificador = veri.Id_Usuario Where v.IdEstado = 3 AND v.Id_Verificacion = %s", [id])
                    Verificacion = cur.fetchall()
                    mysql.connection.commit()
                    # HACEMOS LOA SUMA DE CADA COLUMNA
                    #  SUMA DE LA COLUMNA PESOS BRUTOS
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "SELECT SUM(PesoBruto) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                    sumaBruto1 = cur.fetchone()
                    sumaBruto = round(sumaBruto1[0], 2)
                    #  SUMA DE LA COLUMNA PESOS TARA
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "SELECT SUM(PesoTara) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                    sumaTara1 = cur.fetchone()
                    sumaTara = round(sumaTara1[0], 2)
                    #  SUMA DE LA COLUMNA PESOS DESTARE
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "SELECT SUM(Destare) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                    sumaDestare1 = cur.fetchone()
                    mysql.connection.commit()
                    sumaDestare = round(sumaDestare1[0], 2)
                    #  SUMA DE LA COLUMNA PESOS NETO
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "SELECT SUM(PesoNeto) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                    sumaNeto1 = cur.fetchone()
                    mysql.connection.commit()
                    sumaNeto = round(sumaNeto1[0], 2)
                    #  FECHA VERIFICACION
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "SELECT Fecha FROM tb_verificacion WHERE Id_Verificacion = %s", [id])
                    fecha = cur.fetchone()
                    mysql.connection.commit()
                    fecha1 = capturarHora()
                    fechacreacion = datetime.date(fecha1)
                    # Usuario que lo creó
                    cur = mysql.connection.cursor()
                    cur.execute(
                        "SELECT u.NombreUsuario FROM `tb_verificacion` as v inner join tb_usuarios as u ON v.IdUsuarioCreacion = u.Id_Usuario WHERE Id_Verificacion = %s", [id])
                    usuario = cur.fetchone()
                    mysql.connection.commit()
                    # Cambiar el estado de la verificacion
                    cur = mysql.connection.cursor()
                    cur.execute(
                        'Update tb_verificacion set IdEstado = 8 Where Id_Verificacion = %s', [id])
                    mysql.connection.commit()

                    # =========================
                    # total de materiales
                    cur = mysql.connection.cursor()
                    cur.execute(
                        'SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2) as neto,round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s Group BY ver.IdMaterial', [id])
                    mat = cur.fetchall()
                    mysql.connection.commit()

                    # TOTAL DE MATERIALES DE PRIMERA
                    cur = mysql.connection.cursor()
                    cur.execute(
                        'SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2) as neto, round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s and m.TipoMaterial = "Primera" Group BY m.TipoMaterial', [id])
                    primera = cur.fetchall()
                    mysql.connection.commit()

                    # TOTAL DE MATERIALES DE SEGUNDA
                    cur = mysql.connection.cursor()
                    cur.execute(
                        'SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2),round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s and m.TipoMaterial = "Segunda" Group BY m.TipoMaterial', [id])
                    segunda = cur.fetchall()
                    mysql.connection.commit()

                    # TOTAL DE RECHAZOS
                    cur = mysql.connection.cursor()
                    cur.execute("SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2),round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s AND m.NombreMaterial like %s Group BY m.TipoMaterial", (id, 'rechazo%'))
                    rechazo = cur.fetchall()
                    mysql.connection.commit()

                    # TOTAL DE JUMBOS
                    cur = mysql.connection.cursor()
                    cur.execute('SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2),round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s and m.NombreMaterial like %s Group BY m.TipoMaterial', (id, 'jumbo%'))
                    jumbo = cur.fetchall()
                    mysql.connection.commit()

                    # TOTAL LIQUIDO
                    cur = mysql.connection.cursor()
                    cur.execute('SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2),round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s and m.NombreMaterial like %s Group BY m.TipoMaterial', (id, 'liquido%'))
                    liquido = cur.fetchall()
                    mysql.connection.commit()

                    # TOTAL rechazo Pet
                    # cur = mysql.connection.cursor()
                    # cur.execute('SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2),round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s m.NombreMaterial = "LIQUIDO" Group BY m.TipoMaterial',[id])
                    # rechazoPet = cur.fetchall()
                    # mysql.connection.commit()

                    # AQUI SE DEBERIA DE MANDAR A LLAMAR LA FUNCION PARA GENERAR LA ORDEN DE COMPRA EN ODDO
                    # CrearOrdenCompra(proveedorId,puntoCompra,NoBoleta,rechazo,jumbo,liquido,rechazoPet,primera,segunda)
                    # print(jumbo[0][3])
                    # if jumbo:
                    #     print("jumbo tiene")
                    #     jumboNuevo = jumbo[0][3]
                    # else:
                    #     jumboNuevo = 0.0

                    # print(rechazo)
                    # if rechazo:
                    #     rechazoNuevo = rechazo[0][3]
                    # else:
                    #     rechazoNuevo = 0.0
                    # if liquido:
                    #     liquidoNuevo = liquido[0][3]
                    # else:
                    #     liquidoNuevo = 0.0

                    # print(primera)
                    # if primera and segunda:
                    #     primeraNuevo = primera[0][3]
                    #     segundaNuevo = segunda[0][3]
                    # elif primera and not segunda:
                    #     primeraNuevo = primera[0][3]
                    #     segundaNuevo = 0
                    # elif not primera and segunda:
                    #     primeraNuevo = 0
                    #     segundaNuevo = segunda[0][3]
                    # else:
                    #     primeraNuevo = 1
                    #     segundaNuevo = 1

                    # IdOrden = conexion.CrearOrdenCompra(Verificacion[0][10],Verificacion[0][11],Verificacion[0][3],rechazoNuevo,jumboNuevo,liquidoNuevo,0,primeraNuevo,segundaNuevo)

                    # print("ORDEN AQUI")
                    # print(IdOrden)
                    return "Hecho"
                    return render_template('otros/factura.html', segunda="", primera="", mat=mat, id=id, usuario=usuario, verificacion=Verificacion, fechaEmision=fecha, fechaCreacion=fechacreacion, pesos=pesos, sumaBruto=sumaBruto, sumaTara=sumaTara, sumaDestare=sumaDestare, sumaNeto=sumaNeto)
                else:
                    print("vacio")
                    return "vacio"
        else:
            print("VALIDADOR")
            id = request.form['id']
            tipoMaterial = request.form['tipo']
            peso = request.form['peso']
            var1 = request.form['variacion1']
            var2 = request.form['variacion2']

            cur = mysql.connection.cursor()
            cur.execute(
                "INSERT INTO tb_validacion (IdVerificacion,TipoMaterial,PesoBascula,Variacion1,Variacion2) VALUES (%s,%s,%s,%s,%s)", (id, tipoMaterial, peso,var1,var2))

            mysql.connection.commit()

            # MANDAMOS A LLAMAR TODA LA TABLA DE DETALLE VERIFICACION CON LOS NUEVOS DATOS REGISTRADOS
            cur = mysql.connection.cursor()
            cur.execute(
                "SELECT dt.Id_DetalleVerificacion,dt.IdVerificacion,m.NombreMaterial,dt.PesoBruto,dt.PesoTara,dt.Destare,dt.PesoNeto FROM tb_detalleverificacion as dt inner join tb_material as m ON dt.IdMaterial = m.Id_Material Where dt.IdVerificacion = %s", [id])
            pesos = cur.fetchall()
            mysql.connection.commit()
            print("pesos")
            print(pesos)
            if pesos:
                cur = mysql.connection.cursor()
                cur.execute(
                    "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado, p.NombreProveedor,digi.NombreUsuario as digitador,veri.NombreUsuario as verificador,v.Bahia from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_usuarios as digi on v.IdDigitador = digi.Id_Usuario inner join tb_usuarios as veri on v.IdVerificador = veri.Id_Usuario Where v.IdEstado = 5 AND v.Id_Verificacion = %s", [id])
                Verificacion = cur.fetchall()
                mysql.connection.commit()
                print("VERIFICACION: ",Verificacion)
                # HACEMOS LOA SUMA DE CADA COLUMNA
                #  SUMA DE LA COLUMNA PESOS BRUTOS
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(PesoBruto) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                sumaBruto1 = cur.fetchone()
                sumaBruto = round(sumaBruto1[0], 2)
                #  SUMA DE LA COLUMNA PESOS TARA
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(PesoTara) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                sumaTara1 = cur.fetchone()
                sumaTara = round(sumaTara1[0], 2)
                #  SUMA DE LA COLUMNA PESOS DESTARE
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(Destare) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                sumaDestare1 = cur.fetchone()
                mysql.connection.commit()
                sumaDestare = round(sumaDestare1[0], 2)
                #  SUMA DE LA COLUMNA PESOS NETO
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(PesoNeto) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                sumaNeto1 = cur.fetchone()
                mysql.connection.commit()
                sumaNeto = round(sumaNeto1[0], 2)
                #  FECHA VERIFICACION
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT Fecha FROM tb_verificacion WHERE Id_Verificacion = %s", [id])
                fecha = cur.fetchone()
                mysql.connection.commit()
                fecha1 = capturarHora()
                fechacreacion = datetime.date(fecha1)
                # Usuario que lo creó
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT u.NombreUsuario FROM `tb_verificacion` as v inner join tb_usuarios as u ON v.IdUsuarioCreacion = u.Id_Usuario WHERE Id_Verificacion = %s", [id])
                usuario = cur.fetchone()
                mysql.connection.commit()
                # Cambiar el estado de la verificacion
                cur = mysql.connection.cursor()
                cur.execute(
                    'Update tb_verificacion set IdEstado = 9 Where Id_Verificacion = %s', [id])
                mysql.connection.commit()

                # Cambiar el estado de la verificacion
                cur = mysql.connection.cursor()
                cur.execute('Update tb_verificacion set IdEstado = 9 Where PO = %s', [
                            Verificacion[0][2]])
                mysql.connection.commit()

                # total de materiales
                cur = mysql.connection.cursor()
                cur.execute(
                    'SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2) as neto,round(SUM(ver.Destare),2) as destare FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s Group BY ver.IdMaterial', [id])
                mat = cur.fetchall()
                mysql.connection.commit()
                print(pesos)
                print("factura aquiii")
                print(len(pesos))
                return "Hecho"
                # return render_template('otros/factura.html',mat = mat,id = id,usuario = usuario,verificacion = Verificacion, fechaEmision = fecha,fechaCreacion = fechacreacion,pesos = pesos,sumaBruto = sumaBruto,sumaTara = sumaTara,sumaDestare = sumaDestare,sumaNeto = sumaNeto)
            else:
                print("vacio")
                return "vacio"

    else:
        return "No"


@app.route('/finalizarVerificacionVarios', methods=["POST", "GET"])
def finalizarVerificacionVarios():
    if request.method == "POST":
            print("VALIDADOR")
            id = request.form['id']
            tipoMaterial = request.form['tipo']
            peso = request.form['peso']
            var1 = request.form['variacion1']
            var2 = request.form['variacion2']

            array = ast.literal_eval(id)
            for id1 in array:           
                cur = mysql.connection.cursor()
                cur.execute(
                    "INSERT INTO tb_validacion (IdVerificacion,TipoMaterial,PesoBascula,Variacion1,Variacion2) VALUES (%s,%s,%s,%s,%s)", (id1, tipoMaterial, peso,var1,var2))

                mysql.connection.commit()
                cur = mysql.connection.cursor()
                cur.execute(
                    "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado, p.NombreProveedor,digi.NombreUsuario as digitador,veri.NombreUsuario as verificador,v.Bahia from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_usuarios as digi on v.IdDigitador = digi.Id_Usuario inner join tb_usuarios as veri on v.IdVerificador = veri.Id_Usuario Where v.Id_Verificacion = %s", [id1])
                Verificacion = cur.fetchone()
                mysql.connection.commit()


                print('IDDDD',id1)
                # Cambiar el estado de la verificacion
                cur = mysql.connection.cursor()
                cur.execute(
                    'Update tb_verificacion set IdEstado = 9 Where NoBoleta = %s', [Verificacion[3]])
                mysql.connection.commit()
            return "Hecho"




# FINALIZAR VERIFICACION
@app.route('/finalizarVerificacionmal', methods=["POST", "GET"])
def finalizarVerificacionmal():
    if request.method == "POST":
        if session['cargo'] != 5:
            id = request.form['id']

            # MANDAMOS A LLAMAR TODA LA TABLA DE DETALLE VERIFICACION CON LOS NUEVOS DATOS REGISTRADOS
            cur = mysql.connection.cursor()
            cur.execute(
                "SELECT dt.Id_DetalleVerificacion,dt.IdVerificacion,m.NombreMaterial,dt.PesoBruto,dt.PesoTara,dt.Destare,dt.PesoNeto FROM tb_detalleverificacion as dt inner join tb_material as m ON dt.IdMaterial = m.Id_Material Where dt.IdVerificacion = %s", [id])
            pesos = cur.fetchall()
            mysql.connection.commit()
            print("pesos")
            if pesos:
                cur = mysql.connection.cursor()
                cur.execute(
                    "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado, p.NombreProveedor,digi.NombreUsuario as digitador,veri.NombreUsuario as verificador,v.Bahia from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_usuarios as digi on v.IdDigitador = digi.Id_Usuario inner join tb_usuarios as veri on v.IdVerificador = veri.Id_Usuario Where v.IdEstado = 3 AND v.Id_Verificacion = %s", [id])
                Verificacion = cur.fetchall()
                mysql.connection.commit()
                # HACEMOS LOA SUMA DE CADA COLUMNA
                #  SUMA DE LA COLUMNA PESOS BRUTOS
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(PesoBruto) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                sumaBruto1 = cur.fetchone()
                sumaBruto = round(sumaBruto1[0], 2)
                #  SUMA DE LA COLUMNA PESOS TARA
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(PesoTara) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                sumaTara1 = cur.fetchone()
                sumaTara = round(sumaTara1[0], 2)
                #  SUMA DE LA COLUMNA PESOS DESTARE
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(Destare) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                sumaDestare1 = cur.fetchone()
                mysql.connection.commit()
                sumaDestare = round(sumaDestare1[0], 2)
                #  SUMA DE LA COLUMNA PESOS NETO
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(PesoNeto) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                sumaNeto1 = cur.fetchone()
                mysql.connection.commit()
                sumaNeto = round(sumaNeto1[0], 2)
                #  FECHA VERIFICACION
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT Fecha FROM tb_verificacion WHERE Id_Verificacion = %s", [id])
                fecha = cur.fetchone()
                mysql.connection.commit()
                fecha1 = capturarHora()
                fechacreacion = datetime.date(fecha1)
                # Usuario que lo creó
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT u.NombreUsuario FROM `tb_verificacion` as v inner join tb_usuarios as u ON v.IdUsuarioCreacion = u.Id_Usuario WHERE Id_Verificacion = %s", [id])
                usuario = cur.fetchone()
                mysql.connection.commit()
                # Cambiar el estado de la verificacion
                cur = mysql.connection.cursor()
                cur.execute(
                    'Update tb_verificacion set IdEstado = 6 Where Id_Verificacion = %s', [id])
                mysql.connection.commit()
                # total de materiales
                cur = mysql.connection.cursor()
                cur.execute(
                    'SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2) as neto FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s Group BY ver.IdMaterial', [id])
                mat = cur.fetchall()
                mysql.connection.commit()
                print(pesos)
                print("factura aquiii")
                print(len(pesos))
                return render_template('otros/factura.html', mat=mat, id=id, usuario=usuario, verificacion=Verificacion, fechaEmision=fecha, fechaCreacion=fechacreacion, pesos=pesos, sumaBruto=sumaBruto, sumaTara=sumaTara, sumaDestare=sumaDestare, sumaNeto=sumaNeto)
            else:
                print("vacio")
                return "vacio"
        else:

            print("VALIDADOR")
            id = request.form['id']
            print(id)
            tipoMaterial = request.form['tipo']
            peso = request.form['peso']
            var1 = request.form['variacion1']
            var2 = request.form['variacion2']

            cur = mysql.connection.cursor()
            cur.execute(
                "INSERT INTO tb_validacion (IdVerificacion,TipoMaterial,PesoBascula,Variacion1,Variacion2) VALUES (%s,%s,%s,%s,%s)", (id, tipoMaterial, peso,var1,var2))

            mysql.connection.commit()
            # MANDAMOS A LLAMAR TODA LA TABLA DE DETALLE VERIFICACION CON LOS NUEVOS DATOS REGISTRADOS
            cur = mysql.connection.cursor()
            cur.execute(
                "SELECT dt.Id_DetalleVerificacion,dt.IdVerificacion,m.NombreMaterial,dt.PesoBruto,dt.PesoTara,dt.Destare,dt.PesoNeto FROM tb_detalleverificacion as dt inner join tb_material as m ON dt.IdMaterial = m.Id_Material Where dt.IdVerificacion = %s", [id])
            pesos = cur.fetchall()
            mysql.connection.commit()
            print("pesos")
            print(pesos)
            if pesos:
                cur = mysql.connection.cursor()
                cur.execute(
                    "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado, p.NombreProveedor,digi.NombreUsuario as digitador,veri.NombreUsuario as verificador,v.Bahia from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_usuarios as digi on v.IdDigitador = digi.Id_Usuario inner join tb_usuarios as veri on v.IdVerificador = veri.Id_Usuario Where v.Id_Verificacion = %s", [id])
                Verificacion = cur.fetchone()
                mysql.connection.commit()
                print("VERIFICACION",Verificacion)
                po = Verificacion[2]
                print(po)
                # MANDAR EL COMENTARIO A ODOO
                


                # HACEMOS LOA SUMA DE CADA COLUMNA
                #  SUMA DE LA COLUMNA PESOS BRUTOS
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(PesoBruto) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                sumaBruto1 = cur.fetchone()
                sumaBruto = round(sumaBruto1[0], 2)
                #  SUMA DE LA COLUMNA PESOS TARA
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(PesoTara) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                sumaTara1 = cur.fetchone()
                sumaTara = round(sumaTara1[0], 2)
                #  SUMA DE LA COLUMNA PESOS DESTARE
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(Destare) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                sumaDestare1 = cur.fetchone()
                mysql.connection.commit()
                sumaDestare = round(sumaDestare1[0], 2)
                #  SUMA DE LA COLUMNA PESOS NETO
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT SUM(PesoNeto) FROM tb_detalleverificacion WHERE IdVerificacion = %s", [id])
                sumaNeto1 = cur.fetchone()
                mysql.connection.commit()
                sumaNeto = round(sumaNeto1[0], 2)
                #  FECHA VERIFICACION
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT Fecha FROM tb_verificacion WHERE Id_Verificacion = %s", [id])
                fecha = cur.fetchone()
                mysql.connection.commit()
                fecha1 = capturarHora()
                fechacreacion = datetime.date(fecha1)
                # Usuario que lo creó
                cur = mysql.connection.cursor()
                cur.execute(
                    "SELECT u.NombreUsuario FROM `tb_verificacion` as v inner join tb_usuarios as u ON v.IdUsuarioCreacion = u.Id_Usuario WHERE Id_Verificacion = %s", [id])
                usuario = cur.fetchone()
                mysql.connection.commit()
                # Cambiar el estado de la verificacion
                cur = mysql.connection.cursor()
                cur.execute(
                    'Update tb_verificacion set IdEstado = 6 Where Id_Verificacion = %s', [id])
                mysql.connection.commit()

                # Cambiar el estado de la verificacion
                cur = mysql.connection.cursor()
                cur.execute('Update tb_verificacion set IdEstado = 6 Where PO = %s', [
                            Verificacion[2]])
                mysql.connection.commit()

                # total de materiales
                cur = mysql.connection.cursor()
                cur.execute(
                    'SELECT m.NombreMaterial,round(sum(ver.PesoBruto),2) as bruto,round(sum(ver.PesoTara),2) as tara,round(SUM(ver.PesoNeto),2) as neto FROM tb_detalleverificacion as ver inner join tb_material as m ON ver.IdMaterial = m.Id_Material WHERE ver.IdVerificacion = %s Group BY ver.IdMaterial', [id])
                mat = cur.fetchall()
                mysql.connection.commit()
                print(pesos)
                print("factura aquiii")
                print(len(pesos))
                return render_template('otros/factura.html', mat=mat, id=id, usuario=usuario, verificacion=Verificacion, fechaEmision=fecha, fechaCreacion=fechacreacion, pesos=pesos, sumaBruto=sumaBruto, sumaTara=sumaTara, sumaDestare=sumaDestare, sumaNeto=sumaNeto)
            else:
                print("vacio")
                return "vacio"

    else:
        return "No"
    
# FINALIZAR VERIFICACION
@app.route('/finalizarVerificacionmalVarios', methods=["POST", "GET"])
def finalizarVerificacionmalVarios():
    if request.method == "POST":
        
            print("VALIDADOR")
            id = request.form['id']
            print(id)
            tipoMaterial = request.form['tipo']
            peso = request.form['peso']
            var1 = request.form['variacion1']
            var2 = request.form['variacion2']
            array = ast.literal_eval(id)
            for id1 in array:           
                
                cur = mysql.connection.cursor()
                cur.execute(
                    "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado, p.NombreProveedor,digi.NombreUsuario as digitador,veri.NombreUsuario as verificador,v.Bahia from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_usuarios as digi on v.IdDigitador = digi.Id_Usuario inner join tb_usuarios as veri on v.IdVerificador = veri.Id_Usuario Where v.Id_Verificacion = %s", [id1])
                Verificacion = cur.fetchone()
                mysql.connection.commit()


                print('IDDDD',id1)
                # Cambiar el estado de la verificacion
                cur = mysql.connection.cursor()
                cur.execute(
                    'Update tb_verificacion set IdEstado = 6 Where NoBoleta = %s', [Verificacion[3]])
                mysql.connection.commit()

                
            return "done"
           

    else:
        return "No"



# MODULO DE ERRORES
@app.route('/administracionErrores')
def administracionErrores():
    try:
        if session['userrole'] == 1:
            # NECESITAMOS LA LISTA DE VERIFICADORES, DIGITADORES
            # CONSULTA PARA LOS VERIFICADORES
            cur = mysql.connection.cursor()
            cur.execute(
                "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 2")
            verificador = cur.fetchall()
            # CONSULTA PARA LOS DIGITADOR
            cur = mysql.connection.cursor()
            cur.execute(
                "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 1")
            digitador = cur.fetchall()
            return render_template('AdminErrores.html', verificadores=verificador, digitadores=digitador)
        else:
            return render_template('otros/error.html')
    except:
        return render_template('otros/error.html')



# MODULO DE ADMINISTRACION
@app.route('/administracion')
def administracion():
    try:
        if session['userrole'] == 1:
            # NECESITAMOS LA LISTA DE VERIFICADORES, DIGITADORES
            # CONSULTA PARA LOS VERIFICADORES
            cur = mysql.connection.cursor()
            cur.execute(
                "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 2")
            verificador = cur.fetchall()
            # CONSULTA PARA LOS DIGITADOR
            cur = mysql.connection.cursor()
            cur.execute(
                "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 1")
            digitador = cur.fetchall()
            return render_template('administracion.html', verificadores=verificador, digitadores=digitador)
        else:
            return render_template('otros/error.html')
    except:
        return render_template('otros/error.html')


# VALOR DE LA TABLA DE ADMINISTRACION
@app.route('/valorTablaAdmin', methods=["POST", "GET"])
def valorTablaAdmin():
    if request.method == "POST":
        opc = request.form['valor']
        print(opc)
        # OPCIONES DE FILTRO
        if opc == '1':
            po = request.form['po']
            if po != "":

                cur = mysql.connection.cursor()
                cur.execute(
                    "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.NoBoleta like %s and v.IdEstado = 6 and v.PO != '--' GROUP BY v.PO", [po+'%'])
                verificaciones = cur.fetchall()
                mysql.connection.commit()
                print(verificaciones)
                return render_template('tablas/tabla-filtracion.html', opc=opc, verificaciones=verificaciones)
            else:
                cur = mysql.connection.cursor()
                cur.execute("select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.IdEstado = 6 and v.PO != '--' GROUP BY v.PO")
                verificaciones = cur.fetchall()
                mysql.connection.commit()
                print(verificaciones)
                return render_template('tablas/tabla-filtracion.html', opc=opc, verificaciones=verificaciones)
        elif opc == '2':
            po = request.form['po']
            if po != "":
                cur = mysql.connection.cursor()
                cur.execute(
                    "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.NoBoleta like %s and v.IdEstado = 4 and v.PO != '--' GROUP BY v.PO", [po+'%'])
                verificaciones = cur.fetchall()
                mysql.connection.commit()
                print(verificaciones)
                return render_template('tablas/tabla-filtracion.html', opc=opc, verificaciones=verificaciones)
            else:
                print("validaddas")
                cur = mysql.connection.cursor()
                cur.execute("select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.IdEstado = 4 and v.PO != '--' GROUP BY v.PO")
                verificaciones = cur.fetchall()
                mysql.connection.commit()
                print(verificaciones)
                return render_template('tablas/tabla-filtracion.html', opc=opc, verificaciones=verificaciones)
        elif opc == '3':
            po = request.form['po']
            if po != "":
                cur = mysql.connection.cursor()
                cur.execute(
                    "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.NoBoleta like %s and v.IdEstado = 3 and v.PO != '--' GROUP BY v.PO", [po+'%'])
                verificaciones = cur.fetchall()
                mysql.connection.commit()
                print(verificaciones)
                return render_template('tablas/tabla-filtracion.html', opc=opc, verificaciones=verificaciones)
            else:
                cur = mysql.connection.cursor()
                cur.execute("select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.IdEstado = 3 and v.PO != '--' GROUP BY v.PO")
                verificaciones = cur.fetchall()
                mysql.connection.commit()
                print(verificaciones)
                return render_template('tablas/tabla-filtracion.html', opc=opc, verificaciones=verificaciones)
        if opc == "Usuarios":
            usuario = request.form['proveedor']
            # SELECCIONAR EL ID DEL PROVEEDOR
            cur = mysql.connection.cursor()
            cur.execute(
                "select * from tb_usuarios Where NombreUsuario = %s", [usuario])
            proveedornuevo = cur.fetchone()
            # LLAMAMOS LAS VERIFICACIONES DEL USUARIO
            cur = mysql.connection.cursor()
            cur.execute("select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado Where v.IdUsuarioCreacion = %s and v.PO != '--' GROUP BY v.PO", [
                        proveedornuevo[0]])
            verificaciones = cur.fetchall()
            mysql.connection.commit()
            print(verificaciones)
            return render_template('tablas/tabla-filtracion.html', opc=opc, verificaciones=verificaciones)
            return "usu"
        elif opc == "Proveedores":
            proveedor = request.form['proveedor']
            # SELECCIONAR EL ID DEL PROVEEDOR
            cur = mysql.connection.cursor()
            cur.execute(
                "select * from tb_proveedor Where NombreProveedor = %s", [proveedor])
            proveedornuevo = cur.fetchone()
            # LLAMAMOS LAS VERIFICACIONES DEL USUARIO
            cur = mysql.connection.cursor()
            cur.execute("select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado Where v.IdProveedor = %s and v.PO != '--' GROUP BY v.PO", [
                        proveedornuevo[0]])
            verificaciones = cur.fetchall()
            mysql.connection.commit()
            return render_template('tablas/tabla-filtracion.html', opc=opc, verificaciones=verificaciones)
        elif opc == "verificaciones (Pendientes)":
            cur = mysql.connection.cursor()
            cur.execute("select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado Where v.IdEstado = 3 and v.PO != '--' GROUP BY v.PO")
            verificaciones = cur.fetchall()
            mysql.connection.commit()
            print(verificaciones)
            return render_template('tablas/tabla-filtracion.html', opc=opc, verificaciones=verificaciones)
        elif opc == "verificaciones (Terminadas)":
            cur = mysql.connection.cursor()
            cur.execute("select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado Where v.IdEstado = 4 and v.PO != '--' GROUP BY v.PO")
            verificaciones = cur.fetchall()
            mysql.connection.commit()
            print(verificaciones)
            return render_template('tablas/tabla-filtracion.html', opc=opc, verificaciones=verificaciones)
        elif opc == "po":
            po = request.form['po']
            cur = mysql.connection.cursor()
            cur.execute(
                "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.NoBoleta like %s and v.IdEstado != 3 and v.PO != '--' GROUP BY v.PO,v.NoBoleta", [po+'%'])
            verificaciones = cur.fetchall()
            mysql.connection.commit()
            print(verificaciones)
            return render_template('tablas/tabla-filtracion.html', opc=opc, verificaciones=verificaciones)
        return render_template('tablas/tabla-filtracion.html', opc=opc)

# VALOR TABLA DE ERRORES
@app.route('/traerBoletasSolas', methods=["POST", "GET"])
def traerBoletasSolas():
    if request.method == "POST":
        opc = request.form['valor']
        # OPCIONES DE FILTRO

        po = request.form['po']
        cur = mysql.connection.cursor()
        cur.execute(
            "select * from tb_verificacion where NoBoleta like %s and IdProveedor is Null", [po+'%'])
        verificaciones = cur.fetchall()
        mysql.connection.commit()
        return render_template('tablas/tabla-vacias.html', verificaciones=verificaciones)
        
       
    return render_template('tablas/tabla-filtracionusu.html', opc=opc)

@app.route('/eliminarVacia', methods=["POST", "GET"])
def eliminarVacia():
    if request.method == "POST":
        id = request.form['id']
        # OPCIONES DE FILTRO

        cur = mysql.connection.cursor()
        cur.execute(
            "DELETE from tb_verificacion where Id_Verificacion = %s", [id])
        mysql.connection.commit()
        return "ELIMINADO"
        
       
    return render_template('tablas/tabla-filtracionusu.html')

# CAMBIAMOS EL ESTADOP
@app.route('/cambiarEstado', methods=["POST", "GET"])
def cambiarEstado():
    if request.method == "POST":
        id = request.form['id']
        # OPCIONES DE FILTRO

        cur = mysql.connection.cursor()
        cur.execute(
            "Update tb_verificacion set IdEstado = 3 Where Id_Verificacion = %s", [id])
        mysql.connection.commit()
        return "Cambiado"

# CAMBIAMOS EL ESTADOP
@app.route('/eliminarVer', methods=["POST", "GET"])
def eliminarVer():
    if request.method == "POST":
        id = request.form['id']
        # OPCIONES DE FILTRO

        cur = mysql.connection.cursor()
        cur.execute(
            "DELETE from tb_verificacion Where Id_Verificacion = %s and PO == '--'", [id])
        mysql.connection.commit()
        return "Cambiado"

# ANULAMOS LA VERIFICACION CON NUMERO DE BOLETA
@app.route('/AnularValidacion', methods=["POST", "GET"])
def AnularValidacion():
    if request.method == "POST":
        id = request.form['id']
        # OPCIONES DE FILTRO

        cur = mysql.connection.cursor()
        cur.execute(
            "Update tb_verificacion set IdEstado = 6 Where Id_Verificacion = %s", [id])
        mysql.connection.commit()
        return "Cambiado"



    
# HABILITAMOS LA VALIDACION NUEVAMENTE AL VALIDADOR
@app.route('/HabilitarValidacion', methods=["POST", "GET"])
def HabilitarValidacion():
    if request.method == "POST":
        id = request.form['id']
        # OPCIONES DE FILTRO
        print(id)
        cur = mysql.connection.cursor()
        cur.execute(
            "SELECT Id_Verificacion FROM tb_verificacion WHERE NoBoleta = %s and IdEstado = 6 and IdVerificador = IdUsuarioCreacion", [id])
        Id = cur.fetchone()
        mysql.connection.commit()
        print(Id[0])
        cur = mysql.connection.cursor()
        cur.execute(
            "DELETE from tb_validacion Where IdVerificacion = %s", [Id[0]])
        mysql.connection.commit()

        cur = mysql.connection.cursor()
        cur.execute(
            "Update tb_verificacion set IdEstado = 5 Where NoBoleta = %s and IdEstado = 6", [id])
        mysql.connection.commit()
        return "Cambiado"




# VALOR DE LA TABLA DE ADMINISTRACION


@app.route('/valorTablaUser', methods=["POST", "GET"])
def valorTablaUser():
    if request.method == "POST":
        opc = request.form['valor']
        print(opc)
        # OPCIONES DE FILTRO
        if opc == '1':
            po = request.form['po']
            if po != "":

                cur = mysql.connection.cursor()
                cur.execute(
                    "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.NoBoleta like %s and v.IdEstado = 6 GROUP BY v.PO", [po+'%'])
                verificaciones = cur.fetchall()
                mysql.connection.commit()
                print(verificaciones)
                return render_template('tablas/tabla-filtracionusu.html', opc=opc, verificaciones=verificaciones)
            else:
                cur = mysql.connection.cursor()
                cur.execute("select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.IdEstado = 6 GROUP BY v.PO")
                verificaciones = cur.fetchall()
                mysql.connection.commit()
                print(verificaciones)
                return render_template('tablas/tabla-filtracionusu.html', opc=opc, verificaciones=verificaciones)
        elif opc == '2':
            po = request.form['po']
            if po != "":
                cur = mysql.connection.cursor()
                cur.execute(
                    "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.NoBoleta like %s and v.IdEstado = 4 GROUP BY v.PO", [po+'%'])
                verificaciones = cur.fetchall()
                mysql.connection.commit()
                print(verificaciones)
                return render_template('tablas/tabla-filtracionusu.html', opc=opc, verificaciones=verificaciones)
            else:
                print("validaddas")
                cur = mysql.connection.cursor()
                cur.execute("select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.IdEstado = 4 GROUP BY v.PO")
                verificaciones = cur.fetchall()
                mysql.connection.commit()
                print(verificaciones)
                return render_template('tablas/tabla-filtracionusu.html', opc=opc, verificaciones=verificaciones)
        elif opc == '3':
            po = request.form['po']
            if po != "":
                cur = mysql.connection.cursor()
                cur.execute(
                    "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.NoBoleta like %s and v.IdEstado = 3 GROUP BY v.PO", [po+'%'])
                verificaciones = cur.fetchall()
                mysql.connection.commit()
                print(verificaciones)
                return render_template('tablas/tabla-filtracionusu.html', opc=opc, verificaciones=verificaciones)
            else:
                cur = mysql.connection.cursor()
                cur.execute("select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.IdEstado = 3 GROUP BY v.PO")
                verificaciones = cur.fetchall()
                mysql.connection.commit()
                print(verificaciones)
                return render_template('tablas/tabla-filtracionusu.html', opc=opc, verificaciones=verificaciones)
        elif opc == "po":
            po = request.form['po']
            cur = mysql.connection.cursor()
            cur.execute(
                "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.NoBoleta like %s GROUP BY v.PO", [po+'%'])
            verificaciones = cur.fetchall()
            mysql.connection.commit()
            print(verificaciones)
            return render_template('tablas/tabla-filtracionusu.html', opc=opc, verificaciones=verificaciones)
        return render_template('tablas/tabla-filtracionusu.html', opc=opc)


@app.route('/valorTablaHistorialVerificador', methods=["POST", "GET"])
def valorTablaHistorialVerificador():
    if request.method == "POST":
        opc = request.form['valor']
        print(opc)
        # OPCIONES DE FILTRO
        if session['cargo'] == 5:
            #VALIDADOR 
            if opc == "po":
                po = request.form['po']
                cur = mysql.connection.cursor()
                cur.execute(
                    "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.NoBoleta like %s and v.PO != '--' and v.IdEstado = 9 GROUP BY v.PO", [po+'%'])
                verificaciones = cur.fetchall()
                mysql.connection.commit()
                print(verificaciones)
                print('validador')
                return render_template('tablas/tabla-filtracion-validador.html', opc=opc, verificaciones=verificaciones)
        elif session['cargo'] == 1:
            #DIGITADOR
            if opc == "po":
                po = request.form['po']
                cur = mysql.connection.cursor()
                cur.execute(
                    "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.NoBoleta like %s and v.IdDigitador = %s and v.IdUsuarioCreacion = %s", (po+'%',session["userId"],session["userId"]))
                verificaciones = cur.fetchall()
                mysql.connection.commit()
                print(verificaciones)
                print('digitador')
                return render_template('tablas/tabla-filtracion.html', opc=opc, verificaciones=verificaciones)
        elif session['cargo'] == 2:
            #VERIFICADOR
            if opc == "po":
                po = request.form['po']
                cur = mysql.connection.cursor()
                cur.execute(
                    "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.NoBoleta like %s and v.IdVerificador = %s and v.IdUsuarioCreacion = %s", (po+'%',session["userId"],session["userId"]))
                verificaciones = cur.fetchall()
                mysql.connection.commit()
                print(verificaciones)
                print('validador')
                return render_template('tablas/tabla-filtracion.html', opc=opc, verificaciones=verificaciones)
        
        else:
            if opc == "po":
                po = request.form['po']
                cur = mysql.connection.cursor()
                cur.execute(
                    "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.NoBoleta like %s  GROUP BY v.PO", [po+'%'])
                verificaciones = cur.fetchall()
                mysql.connection.commit()
                print(verificaciones)
                return render_template('tablas/tabla-filtracion.html', opc=opc, verificaciones=verificaciones)
        return render_template('tablas/tabla-filtracion-validador.html', opc=opc)

# APARTADO DE AJUSTES

@app.route('/ajustes')
def ajustes():
    try:
        if session['userId']:
            return render_template('ajustes.html')
        else:
            return render_template('otros/error.html')

    except:
        return render_template('otros/error.html')


# APARTADO DE VER PROVEEDORES
@app.route('/verProveedores', methods=["POST", "GET"])
def verProveedores():
    if request.method == "POST":

        # SELECCIONAR EL ID DEL PROVEEDOR
        cur = mysql.connection.cursor()
        cur.execute("select * from tb_proveedor Where IdEstado = 3")
        proveedores = cur.fetchall()

        return render_template('modal/proveedores-modal.html')
# APARTADO DE VER USUARIOS


@app.route('/verUsuarios', methods=["POST", "GET"])
def verUsuarios():
    try:
        if session['userrole'] == 1:
            if request.method == "POST":
                # SELECCIONAR EL ID DEL PROVEEDOR
                cur = mysql.connection.cursor()
                cur.execute("select * from tb_usuarios Where IdEstado = 3")
                proveedores = cur.fetchall()

            return render_template('modal/usuarios-modal.html')
        else:
            return render_template('otros/error.html')
    except:
        return render_template('otros/error.html')


# TRAER TODOS LOS PROVEEDORES
@app.route('/traerProveedores', methods=["POST", "GET"])
def traerProveedores():
    if request.method == "POST":
        proveedor = request.form['proveedor']
        if proveedor == "":
            # SELECCIONAR EL ID DEL PROVEEDOR
            cur = mysql.connection.cursor()
            cur.execute(
                "SELECT p.Id_Proveedor,p.NombreProveedor,e.NombreEstado,p.Cedula FROM tb_proveedor as p inner join tb_estado as e ON p.IdEstado = e.Id_Estado")
            proveedores = cur.fetchall()
        else:
            cur = mysql.connection.cursor()
            cur.execute("SELECT p.Id_Proveedor,p.NombreProveedor,e.NombreEstado,p.Cedula FROM tb_proveedor as p inner join tb_estado as e ON p.IdEstado = e.Id_Estado where p.NombreProveedor like %s", [
                        proveedor+'%'])
            proveedores = cur.fetchall()

        print(proveedores)
        return render_template('tablas/tabla-proveedornuevo.html', prov=proveedores)

# TRAER TODOS LOS PROVEEDORES


@app.route('/traerUsuarios', methods=["POST", "GET"])
def traerUsuarios():
    if request.method == "POST":
        cedula = request.form['usuario']
        if cedula == "":
            # SELECCIONAR EL ID DEL PROVEEDOR
            cur = mysql.connection.cursor()
            cur.execute("SELECT u.Id_Usuario,u.NombreUsuario as Nombre,c.Usuarios as Usuario,r.NombreRol,e.NombreEstado,u.Cedula,car.NombreCargo FROM tb_usuarios as u inner join tb_credenciales as c ON u.IdCredenciales = c.Id_Credenciales inner join tb_roles as r on c.IdRol = r.Id_Rol  inner join tb_estado as e on u.IdEstado = e.Id_Estado inner join tb_cargo as car on u.IdCargo = car.Id_Cargo where IdCargo = 2")
            proveedores = cur.fetchall()
        else:
            cur = mysql.connection.cursor()
            cur.execute(
                "SELECT u.Id_Usuario,u.NombreUsuario as Nombre,c.Usuarios as Usuario,r.NombreRol,e.NombreEstado,u.Cedula,car.NombreCargo FROM tb_usuarios as u inner join tb_credenciales as c ON u.IdCredenciales = c.Id_Credenciales inner join tb_roles as r on c.IdRol = r.Id_Rol  inner join tb_estado as e on u.IdEstado = e.Id_Estado inner join tb_cargo as car on u.IdCargo = car.Id_Cargo where u.Cedula like %s", [cedula+'%'])
            proveedores = cur.fetchall()

        return render_template('tablas/tabla-usuario.html', prov=proveedores)

# INSERTAMOS LOS PESOS DEL MATERIALE SELECCIONADO


@app.route('/insertarProveedor', methods=["POST", "GET"])
def insertarProveedor():
    if request.method == "POST":
        cedula = request.form['cedula']
        proveedor = request.form['proveedor']
        # INSERTAMOS EL NUEVO PROVEEDOR EN LA BASE
        cur = mysql.connection.cursor()
        cur.execute("INSERT INTO tb_proveedor (NombreProveedor,Cedula,IdEstado) VALUES (%s,%s,1)",
                    (proveedor.upper(), cedula))
        mysql.connection.commit()
        return "done"
    else:
        return "No"

# INSERTAMOS EL USUARIO


@app.route('/insertarUsuario', methods=["POST", "GET"])
def insertarUsuario():
    if request.method == "POST":
        flagUSuario = "nuevo"
        cedula = request.form['cedula']
        if flagUSuario == "editar":
            id = request.form['id']
            cur = mysql.connection.cursor()
            cur.execute(
                "SELECT u.Id_Usuario,u.NombreUsuario as Nombre,c.Usuarios as Usuario,r.NombreRol,e.NombreEstado,u.Cedula,car.NombreCargo FROM tb_usuarios as u inner join tb_credenciales as c ON u.IdCredenciales = c.Id_Credenciales inner join tb_roles as r on u.Id_Usuario = r.Id_Rol inner join tb_estado as e on u.IdEstado = e.Id_Estado inner join tb_cargo as car on u.IdCargo = car.Id_Cargo where Id_Usuario = %s", [id])
            usuario = cur.fetchall()

        # TRAEMOS LOS CARGOS
        cur = mysql.connection.cursor()
        cur.execute("SELECT * From tb_cargo")
        cargos = cur.fetchall()
        # TRAEMOS LOS ROLES
        cur = mysql.connection.cursor()
        cur.execute("SELECT * From tb_roles")
        roles = cur.fetchall()

        # TRAEMOS LOS PUNTOS DE VENTA
        cur = mysql.connection.cursor()
        cur.execute("SELECT * From tb_puntocompra")
        punto = cur.fetchall()
        return render_template('modal/usuarionuevo-modal.html', punto=punto, flagUSuario=flagUSuario, info=cedula, cargos=cargos, roles=roles)

# INSERTAMOS EL USUARIO NUEVO


@app.route('/insertarUsuarioNuevo', methods=["POST", "GET"])
def insertarUsuarioNuevo():
    if request.method == "POST":
        flagUSuario = "nuevo"
        cedula = request.form['cedula']

        # TRAEMOS LOS CARGOS
        cur = mysql.connection.cursor()
        cur.execute("SELECT * From tb_cargo")
        cargos = cur.fetchall()
        # TRAEMOS LOS ROLES
        cur = mysql.connection.cursor()
        cur.execute("SELECT * From tb_roles")
        roles = cur.fetchall()
        print(cedula)
        # TRAEMOS LOS PUNTOS DE VENTA
        cur = mysql.connection.cursor()
        cur.execute("SELECT * From tb_puntocompra")
        punto = cur.fetchall()
        return render_template('modal/usuarionuevo-modal.html', punto=punto, flagUSuario=flagUSuario, info=cedula, cargos=cargos, roles=roles)

# TRAER TARAS


@app.route('/traerTara', methods=["POST", "GET"])
def traerTara():
    if request.method == "POST":
        id = request.form['id']
        # TRAEMOS LOS valores de la tara a la tabla
        cur = mysql.connection.cursor()
        cur.execute(
            "SELECT * From tb_detalletara where IdDetalleVerificacion = %s", [id])
        taras = cur.fetchall()
        print("Taras traer aqui: ", taras)
        return render_template('tablas/tabla-verTara.html', taras=taras)


# DETALLE DE TARAS PARA AÑADIR LOS OTROS VALORES DE LA TARA
@app.route('/addTarasExtras', methods=["POST", "GET"])
def addTarasExtras():
    if request.method == "POST":
        id = request.form['id']
        valor = request.form['valor']
        tipo = request.form['tipo']
        print("iddd")
        print(id)
        # INSERTAMOS el valor de la tara al detalle
        cur = mysql.connection.cursor()
        cur.execute(
            "INSERT INTO tb_detalletara (IdDetalleVerificacion,Contenedor,ValorTaraExtra) VALUES (%s,%s,%s)", (id, tipo, valor))
        mysql.connection.commit()
        # TRAEMOS LOS valores de la tara a la tabla
        cur = mysql.connection.cursor()
        cur.execute(
            "SELECT * From tb_detalletara where IdDetalleVerificacion = %s", [id])
        taras = cur.fetchall()

        # total del peso de las taras que existen en el pesaje
        cur = mysql.connection.cursor()
        cur.execute(
            'SELECT PesoTara as tara from tb_detalleverificacion WHERE Id_DetalleVerificacion = %s', [id])
        pesotaraviejo = cur.fetchone()
        mysql.connection.commit()
        print(pesotaraviejo)
        taranueva = round(float(pesotaraviejo[0]) + float(valor), 2)
        print(taranueva)
        cur.execute(
            "UPDATE tb_detalleverificacion set PesoTara = %s Where Id_DetalleVerificacion = %s", (taranueva, id))
        proveedor = cur.fetchall()
        mysql.connection.commit()

        # TRAEMOS LA TARA
        # PARA HACER LAS OPERACIONES
        cur = mysql.connection.cursor()
        cur.execute(
            'SELECT PesoBruto,PesoTara,Destare as tara from tb_detalleverificacion WHERE Id_DetalleVerificacion = %s', [id])
        datos = cur.fetchone()
        mysql.connection.commit()

        # AL MODIFICAR EL PESO BRUTO TAMBIEN SE MODIFICA EL PESO NETO

        pNeto = round(
            float(float(datos[0])-float(datos[1])-float(datos[2])), 2)

        cur = mysql.connection.cursor()
        cur.execute(
            "UPDATE tb_detalleverificacion set PesoNeto = %s Where Id_DetalleVerificacion = %s", (pNeto, id))
        mysql.connection.commit()
        print("pessooooo netooo")
        print(datos[0])
        print(pNeto)
        return render_template('tablas/tabla-taras.html', taras=taras, total=taranueva)

# TRAER PESOS NETOS Y DESTARE PARA MODIFICARLOS


@app.route('/traerPesosNetos', methods=["POST", "GET"])
def traerPesosNetos():
    if request.method == "POST":
        id = request.form['id']
        # TRAEMOS LOS valores de la tara a la tabla
        cur = mysql.connection.cursor()
        cur.execute(
            "SELECT * From tb_detalleverificacion where Id_DetalleVerificacion = %s", [id])
        detalle = cur.fetchall()

        return jsonify(detalle[0][3])


@app.route('/traerDestare', methods=["POST", "GET"])
def traerDestare():
    if request.method == "POST":
        id = request.form['id']
        # TRAEMOS LOS valores de la tara a la tabla
        cur = mysql.connection.cursor()
        cur.execute(
            "SELECT * From tb_detalleverificacion where Id_DetalleVerificacion = %s", [id])
        detalle = cur.fetchall()

        return jsonify(detalle[0][6])


# VER TARA ESPECIFICA
@app.route('/verTaras', methods=["POST", "GET"])
def verTaras():
    if request.method == "POST":
        id = request.form['id']
        # TRAEMOS LOS valores de la tara a la tabla
        cur = mysql.connection.cursor()
        cur.execute(
            "SELECT * From tb_detalletara where IdDetalleVerificacion = %s", [id])
        taras = cur.fetchall()

        # total del peso de las taras que existen en el pesaje
        cur = mysql.connection.cursor()
        cur.execute(
            'SELECT PesoTara as tara from tb_detalleverificacion WHERE Id_DetalleVerificacion = %s', [id])
        pesotaraviejo = cur.fetchone()
        mysql.connection.commit()

        return render_template('tablas/tabla-taras.html', taras=taras, total=pesotaraviejo)

# MODIFICAR EL PESO BRUTO


@app.route('/modificarPesoBruto', methods=["POST", "GET"])
def modificarPesoBruto():
    if request.method == "POST":
        id = request.form['id']
        valor = request.form['valor']
        destare = request.form['destare']
        # MODIFICAMOS EL VALOR DEL PESO BRUTO
        cur = mysql.connection.cursor()
        cur.execute(
            "UPDATE tb_detalleverificacion set PesoBruto = %s, Destare = %s Where Id_DetalleVerificacion = %s", (valor, destare, id))
        proveedor = cur.fetchall()
        mysql.connection.commit()

        # TRAEMOS LA TARA
        # PARA HACER LAS OPERACIONES
        cur = mysql.connection.cursor()
        cur.execute(
            'SELECT PesoTara,Destare as tara from tb_detalleverificacion WHERE Id_DetalleVerificacion = %s', [id])
        datos = cur.fetchone()
        mysql.connection.commit()
        mysql.connection.commit()
        # AL MODIFICAR EL PESO BRUTO TAMBIEN SE MODIFICA EL PESO NETO

        pNeto = round(float(float(valor)-float(datos[0])-float(datos[1])), 2)

        cur = mysql.connection.cursor()
        cur.execute(
            "UPDATE tb_detalleverificacion set PesoNeto = %s Where Id_DetalleVerificacion = %s", (pNeto, id))
        proveedor = cur.fetchall()
        mysql.connection.commit()

        return "done"

# MODIFICAR EL PESO BRUTO


@app.route('/elimLinea', methods=["POST", "GET"])
def elimLinea():
    if request.method == "POST":
        id = request.form['id']
        # MODIFICAMOS EL VALOR DEL PESO BRUTO
        cur = mysql.connection.cursor()
        cur.execute(
            "DELETE from tb_detalleverificacion Where Id_DetalleVerificacion = %s", [id])
        mysql.connection.commit()

        return "done"

# MANDAMOS LOS IDS QUE DESEAMOS CONCILIAR Y UNIFICAR
@app.route('/conciliarId', methods=["POST", "GET"])
def conciliarId():
    if request.method == "POST":
        id1 = request.form['ids']
        ids = json.loads(id1)
        print(ids['registrosSeleccionados'][0])
        lista_total = []

        # CARGAMOS LOS DATOS DE LA VERIFICACION
        cur = mysql.connection.cursor()
        cur.execute(
                    "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado, p.NombreProveedor,digi.NombreUsuario as digitador,veri.NombreUsuario as verificador,v.Bahia from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_usuarios as digi on v.IdDigitador = digi.Id_Usuario inner join tb_usuarios as veri on v.IdVerificador = veri.Id_Usuario Where v.Id_Verificacion = %s", [ids['registrosSeleccionados'][0]])
        verificacion = cur.fetchall()
        mysql.connection.commit()
        
        lista_po = []
        for id in ids['registrosSeleccionados']:
            #VAMOS A TRAER LAS POS QUE POSEEN ESA BOLETA
            cur = mysql.connection.cursor()
            cur.execute(
                        "select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,v.IdEstado, p.NombreProveedor,digi.NombreUsuario as digitador,veri.NombreUsuario as verificador,v.Bahia from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_usuarios as digi on v.IdDigitador = digi.Id_Usuario inner join tb_usuarios as veri on v.IdVerificador = veri.Id_Usuario Where v.Id_Verificacion = %s", [id])
            verificacion1 = cur.fetchone()
            mysql.connection.commit()
            mysql.connection.commit()
            print("LISTA DE POS")
            print(verificacion1)
            lista_po.append(verificacion1[2])
            print(lista_po)
            
        
        # DATOS EXTRAS DEL MODAL
        # ESTAS CONSULTAS SON PARA TRAER LOS PROVEEDORES Y LOS DATOS DE LOS SELECT
        # CONSULTA PARA LOS PUNTOS DE COMPRA
        cur = mysql.connection.cursor()
        cur.execute("select * from tb_puntocompra Where IdEstado = 1")
        punto = cur.fetchall()
        # CONSULTA PARA LOS MATERIALES
        cur = mysql.connection.cursor()
        cur.execute("select * from tb_material Where Id_Estado = 1")
        material = cur.fetchall()
        # CONSULTA PARA LOS VERIFICADORES
        cur = mysql.connection.cursor()
        cur.execute(
            "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 2")
        verificador = cur.fetchall()
        # CONSULTA PARA LOS DIGITADOR
        cur = mysql.connection.cursor()
        cur.execute(
                "select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 1")
        digitador = cur.fetchall()

            # necesito mandar a llamar los dastos del usuario que esta logueado para ponerlo como verificador
        cur = mysql.connection.cursor()
        cur.execute("select cred.Id_Credenciales,u.NombreUsuario from tb_usuarios as u inner join tb_credenciales as cred on u.IdCredenciales = cred.Id_Credenciales Where cred.Id_Credenciales = %s", [
                        session['userId']])
        usuariolog = cur.fetchone()


        # necesito mandar a llamar los dastos del usuario para saber en que punto de venta esta
        cur = mysql.connection.cursor()
        cur.execute("select pc.Id_PuntoCompra,pc.NombrePuntoCompra from tb_usuarios as u inner join tb_puntocompra as pc on u.IdPuesto = pc.Id_PuntoCompra inner join tb_credenciales as cred on u.IdCredenciales = cred.Id_Credenciales where cred.Id_Credenciales = %s", [
                    session['userId']])
        usuariopunto = cur.fetchone()

        print("IDDDDS",ids)
        # LE PASAMOS EL JSON QUE CONTIENE LOS IDS DE LOS REGISTROS QUE VIENEN PARA CONCILIARSE
        return render_template('modal/comparacion-modal.html',po = lista_po,ids = id1,validacion = "",val = "", usuariopunto=usuariopunto, usuariolog=usuariolog, verificacion=verificacion, Punto=punto, Material=material, Verificador=verificador, Digitador=digitador)
# ELIMINAR EL VALOR TARA DETALLE


@app.route('/elimLIneaTara', methods=["POST", "GET"])
def elimLIneaTara():
    if request.method == "POST":
        id = request.form['id']

        # SELECCIONAMOS EL VALOR QUE VAMOS A BORRAR PARA RESTARSELO AL TOTAL
        cur = mysql.connection.cursor()
        cur.execute(
            "SELECT IdDetalleVErificacion,ValorTaraExtra from tb_detalletara Where Id_DetalleTara = %s", [id])
        valorTara = cur.fetchone()
        mysql.connection.commit()

        # total del peso de las taras que existen en el pesaje
        cur = mysql.connection.cursor()
        cur.execute('SELECT PesoTara as tara from tb_detalleverificacion WHERE Id_DetalleVerificacion = %s', [
                    valorTara[0]])
        pesotaraviejo = cur.fetchone()
        mysql.connection.commit()

        taranueva = round(float(pesotaraviejo[0]) - float(valorTara[1]), 2)
        cur.execute("UPDATE tb_detalleverificacion set PesoTara = %s Where Id_DetalleVerificacion = %s",
                    (taranueva, valorTara[0]))
        proveedor = cur.fetchall()
        mysql.connection.commit()

        # TRAEMOS LA TARA
        # PARA HACER LAS OPERACIONES
        cur = mysql.connection.cursor()
        cur.execute('SELECT PesoBruto,PesoTara,Destare as tara from tb_detalleverificacion WHERE Id_DetalleVerificacion = %s', [
                    valorTara[0]])
        datos = cur.fetchone()
        mysql.connection.commit()

        # AL MODIFICAR EL PESO BRUTO TAMBIEN SE MODIFICA EL PESO NETO

        pNeto = round(
            float(float(datos[0])-float(datos[1])-float(datos[2])), 2)

        cur = mysql.connection.cursor()
        cur.execute(
            "UPDATE tb_detalleverificacion set PesoNeto = %s Where Id_DetalleVerificacion = %s", (pNeto, valorTara[0]))
        mysql.connection.commit()

        print("aaaaaaaaaaaaaaa")
        print(datos)
        print(pesotaraviejo)
        print(valorTara[1])
        print(taranueva)

        # MODIFICAMOS EL VALOR DEL PESO BRUTO
        cur = mysql.connection.cursor()
        cur.execute(
            "DELETE from tb_detalletara Where Id_DetalleTara = %s", [id])
        mysql.connection.commit()

        return "done"

# CREACION DE FILTROS PARA LA BUSQUEDA DE VERIFICACIONES


@app.route('/addFiltro', methods=["POST", "GET"])
def addFiltro():
    if request.method == "POST":
        filtro = request.form['filtro']
        # cur = mysql.connection.cursor()
        # cur.execute("select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.PO like %s and v.IdEstado = 6",[po+'%'])
        # verificaciones = cur.fetchall()
        # mysql.connection.commit()
        # primero es tener la consulta base
        data = json.loads(filtro)
        headers = []
        if data:
            consultaBase = 'select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where'
            for clave in data.keys():
                headers.append(clave)
            consulta = ''
            contador = 0

            for value in data.values():
                if headers[contador] == "Fecha":
                    fechas = value
                    fechas = fechas.split('a')
                    print(fechas[0])
                    consulta += 'date(v.'+headers[contador]+') BETWEEN "' + \
                        fechas[0]+'" AND "'+''+fechas[1]+'" AND '
                elif headers[contador] == "IdProveedor":
                    a = 1
                    cur = mysql.connection.cursor()
                    cur.execute(
                        'SELECT Id_Proveedor from tb_proveedor where NombreProveedor = %s', [value])
                    idprov = cur.fetchone()
                    cur = mysql.connection.cursor()

                    print("aqui se muestra el proveedor:")
                    print(idprov)

                    if idprov:
                        cur = mysql.connection.cursor()
                        cur.execute(
                            'SELECT Id_Proveedor from tb_proveedor where NombreProveedor = %s', [value])
                        idprov = cur.fetchone()

                        consulta += 'v.'+headers[contador] + \
                            ' = '+str(idprov[0]) + ' AND '
                    else:
                        # llamar el id de oddo
                        idOddo = conexion.buscarIdProveedor(value)
                        # INSERTAMOS EL PROVEEDOR DE ODDO EN LA BASE
                        cur = mysql.connection.cursor()
                        cur.execute(
                            "INSERT INTO tb_proveedor (NombreProveedor,IdOddo,IdEstado) VALUES (%s,%s,1)", (value, idOddo))
                        proveedornuevo = cur.fetchone()
                        # LLAMAMOS AL PROVEEDOR DE NOMBRE TAL
                        cur = mysql.connection.cursor()
                        cur.execute(
                            'SELECT Id_Proveedor from tb_proveedor where NombreProveedor = %s', [value])
                        idprov = cur.fetchone()

                        consulta += 'v.'+headers[contador] + \
                            ' = '+str(idprov[0]) + ' AND '

                    # consulta += 'date(v.'+headers[contador]+') BETWEEN "'+fechas[0]+'" AND "'+''+fechas[1]+'" AND '
                else:
                    
                    
                    consulta += 'v.'+headers[contador]+' = '+value+' AND '
                contador += 1
                # consultaBase += ' AND '+data
            consulta_total = consultaBase+' '+consulta[:-4]+' GROUP BY v.PO'
            print(consulta_total)
        else:
            consulta_total = 'select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.PO != "--" and v.IdEstado = 9 GROUP BY v.PO'
            
        cur = mysql.connection.cursor()
        cur.execute(" "+consulta_total)
        verificaciones = cur.fetchall()
        mysql.connection.commit()
        return render_template('tablas/tabla-filtracion.html', opc="", verificaciones=verificaciones)

@app.route('/addFiltroVal', methods=["POST", "GET"])
def addFiltroVal():
    if request.method == "POST":
        filtro = request.form['filtro']
        # cur = mysql.connection.cursor()
        # cur.execute("select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.PO like %s and v.IdEstado = 6",[po+'%'])
        # verificaciones = cur.fetchall()
        # mysql.connection.commit()
        # primero es tener la consulta base
        data = json.loads(filtro)
        headers = []
        if data:
            consultaBase = 'select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.PO != "--" and v.IdEstado = 9 and'
            for clave in data.keys():
                headers.append(clave)
            consulta = ''
            contador = 0

            for value in data.values():
                if headers[contador] == "Fecha":
                    fechas = value
                    fechas = fechas.split('a')
                    print(fechas[0])
                    consulta += 'date(v.'+headers[contador]+') BETWEEN "' + \
                        fechas[0]+'" AND "'+''+fechas[1]+'" AND '
                elif headers[contador] == "IdProveedor":
                    a = 1
                    cur = mysql.connection.cursor()
                    cur.execute(
                        'SELECT Id_Proveedor from tb_proveedor where NombreProveedor = %s', [value])
                    idprov = cur.fetchone()
                    cur = mysql.connection.cursor()

                    print("aqui se muestra el proveedor:")
                    print(idprov)

                    if idprov:
                        cur = mysql.connection.cursor()
                        cur.execute(
                            'SELECT Id_Proveedor from tb_proveedor where NombreProveedor = %s', [value])
                        idprov = cur.fetchone()

                        consulta += 'v.'+headers[contador] + \
                            ' = '+str(idprov[0]) + ' AND '
                    else:
                        # llamar el id de oddo
                        idOddo = conexion.buscarIdProveedor(value)
                        # INSERTAMOS EL PROVEEDOR DE ODDO EN LA BASE
                        cur = mysql.connection.cursor()
                        cur.execute(
                            "INSERT INTO tb_proveedor (NombreProveedor,IdOddo,IdEstado) VALUES (%s,%s,1)", (value, idOddo))
                        proveedornuevo = cur.fetchone()
                        # LLAMAMOS AL PROVEEDOR DE NOMBRE TAL
                        cur = mysql.connection.cursor()
                        cur.execute(
                            'SELECT Id_Proveedor from tb_proveedor where NombreProveedor = %s', [value])
                        idprov = cur.fetchone()

                        consulta += 'v.'+headers[contador] + \
                            ' = '+str(idprov[0]) + ' AND '

                    # consulta += 'date(v.'+headers[contador]+') BETWEEN "'+fechas[0]+'" AND "'+''+fechas[1]+'" AND '
                else:
                    
                    
                    consulta += 'v.'+headers[contador]+' = '+value+' AND '
                contador += 1
                # consultaBase += ' AND '+data
            consulta_total = consultaBase+' '+consulta[:-4] + 'and v.PO != "--"'
            print(consulta_total)
        else:
            consulta_total = 'select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.PO != "--" and v.IdEstado = 9 group by v.PO'
            print('consulta: ',consulta_total)   
            cur = mysql.connection.cursor()
            cur.execute(" "+consulta_total)
            verificaciones = cur.fetchall()
            mysql.connection.commit()
            return render_template('tablas/tabla-filtracion-validador.html', opc="", verificaciones=verificaciones)
        consulta_total += 'group by v.PO'
        print('consulta: ',consulta_total)   
        cur = mysql.connection.cursor()
        cur.execute(" "+consulta_total)
        verificaciones = cur.fetchall()
        mysql.connection.commit()
        return render_template('tablas/tabla-filtracion.html', opc="", verificaciones=verificaciones)


#VERIFICADORES
@app.route('/AñadirFiltroVer', methods=["POST", "GET"])
def AñadirFiltroVer():
    if request.method == "POST":
        filtro = request.form['filtro']
        # cur = mysql.connection.cursor()
        # cur.execute("select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.PO like %s and v.IdEstado = 6",[po+'%'])
        # verificaciones = cur.fetchall()
        # mysql.connection.commit()
        # primero es tener la consulta base
        data = json.loads(filtro)
        headers = []
        if data:
            consultaBase = 'select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where'
            for clave in data.keys():
                headers.append(clave)
            consulta = ''
            contador = 0

            for value in data.values():
                if headers[contador] == "Fecha":
                    fechas = value
                    fechas = fechas.split('a')
                    print(fechas[0])
                    consulta += 'date(v.'+headers[contador]+') BETWEEN "' + \
                        fechas[0]+'" AND "'+''+fechas[1]+'" AND '
                elif headers[contador] == "IdProveedor":
                    a = 1
                    cur = mysql.connection.cursor()
                    cur.execute(
                        'SELECT Id_Proveedor from tb_proveedor where NombreProveedor = %s', [value])
                    idprov = cur.fetchone()
                    cur = mysql.connection.cursor()

                    print("aqui se muestra el proveedor:")
                    print(idprov)

                    if idprov:
                        cur = mysql.connection.cursor()
                        cur.execute(
                            'SELECT Id_Proveedor from tb_proveedor where NombreProveedor = %s', [value])
                        idprov = cur.fetchone()

                        consulta += 'v.'+headers[contador] + \
                            ' = '+str(idprov[0]) + ' AND '
                    else:
                        # llamar el id de oddo
                        idOddo = conexion.buscarIdProveedor(value)
                        # INSERTAMOS EL PROVEEDOR DE ODDO EN LA BASE
                        cur = mysql.connection.cursor()
                        cur.execute(
                            "INSERT INTO tb_proveedor (NombreProveedor,IdOddo,IdEstado) VALUES (%s,%s,1)", (value, idOddo))
                        proveedornuevo = cur.fetchone()
                        # LLAMAMOS AL PROVEEDOR DE NOMBRE TAL
                        cur = mysql.connection.cursor()
                        cur.execute(
                            'SELECT Id_Proveedor from tb_proveedor where NombreProveedor = %s', [value])
                        idprov = cur.fetchone()

                        consulta += 'v.'+headers[contador] + \
                            ' = '+str(idprov[0]) + ' AND '

                    # consulta += 'date(v.'+headers[contador]+') BETWEEN "'+fechas[0]+'" AND "'+''+fechas[1]+'" AND '
                else:
                    
                    
                    consulta += 'v.'+headers[contador]+' = '+value+' AND '
                contador += 1
                # consultaBase += ' AND '+data
            consulta_total = consultaBase+' '+consulta[:-4] + 'and v.PO != "--"'
            print(consulta_total)
        else:
            consulta_total = 'select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.IdVerificacador = %s group by v.PO'
            print('consulta: ',consulta_total)   
            cur = mysql.connection.cursor()
            cur.execute('select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.IdVerificador = %s group by v.PO',[session['userId']])
            verificaciones = cur.fetchall()
            mysql.connection.commit()
            return render_template('tablas/tabla-filtracion.html', opc="", verificaciones=verificaciones)

        print('consulta: ',consulta_total)   
        cur = mysql.connection.cursor()
        cur.execute(" "+consulta_total)
        verificaciones = cur.fetchall()
        mysql.connection.commit()
        return render_template('tablas/tabla-filtracion.html', opc="", verificaciones=verificaciones)


#DIGITADOR
@app.route('/AñadirFiltroDi', methods=["POST", "GET"])
def AñadirFiltroDi():
    if request.method == "POST":
        filtro = request.form['filtro']
        # cur = mysql.connection.cursor()
        # cur.execute("select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.PO like %s and v.IdEstado = 6",[po+'%'])
        # verificaciones = cur.fetchall()
        # mysql.connection.commit()
        # primero es tener la consulta base
        data = json.loads(filtro)
        headers = []
        if data:
            consultaBase = 'select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where'
            for clave in data.keys():
                headers.append(clave)
            consulta = ''
            contador = 0

            for value in data.values():
                if headers[contador] == "Fecha":
                    fechas = value
                    fechas = fechas.split('a')
                    print(fechas[0])
                    consulta += 'date(v.'+headers[contador]+') BETWEEN "' + \
                        fechas[0]+'" AND "'+''+fechas[1]+'" AND '
                elif headers[contador] == "IdProveedor":
                    a = 1
                    cur = mysql.connection.cursor()
                    cur.execute(
                        'SELECT Id_Proveedor from tb_proveedor where NombreProveedor = %s', [value])
                    idprov = cur.fetchone()
                    cur = mysql.connection.cursor()

                    print("aqui se muestra el proveedor:")
                    print(idprov)

                    if idprov:
                        cur = mysql.connection.cursor()
                        cur.execute(
                            'SELECT Id_Proveedor from tb_proveedor where NombreProveedor = %s', [value])
                        idprov = cur.fetchone()

                        consulta += 'v.'+headers[contador] + \
                            ' = '+str(idprov[0]) + ' AND '
                    else:
                        # llamar el id de oddo
                        idOddo = conexion.buscarIdProveedor(value)
                        # INSERTAMOS EL PROVEEDOR DE ODDO EN LA BASE
                        cur = mysql.connection.cursor()
                        cur.execute(
                            "INSERT INTO tb_proveedor (NombreProveedor,IdOddo,IdEstado) VALUES (%s,%s,1)", (value, idOddo))
                        proveedornuevo = cur.fetchone()
                        # LLAMAMOS AL PROVEEDOR DE NOMBRE TAL
                        cur = mysql.connection.cursor()
                        cur.execute(
                            'SELECT Id_Proveedor from tb_proveedor where NombreProveedor = %s', [value])
                        idprov = cur.fetchone()

                        consulta += 'v.'+headers[contador] + \
                            ' = '+str(idprov[0]) + ' AND '

                    # consulta += 'date(v.'+headers[contador]+') BETWEEN "'+fechas[0]+'" AND "'+''+fechas[1]+'" AND '
                else:
                    
                    
                    consulta += 'v.'+headers[contador]+' = '+value+' AND '
                contador += 1
                # consultaBase += ' AND '+data
            consulta_total = consultaBase+' '+consulta[:-4] + 'and v.PO != "--"'
            print(consulta_total)
        else:
            consulta_total = 'select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.IdDigitador = %s group by v.PO'
            print('consulta: ',consulta_total)   
            cur = mysql.connection.cursor()
            cur.execute('select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.IdDigitador = %s group by v.PO',[session['userId']])
            verificaciones = cur.fetchall()
            mysql.connection.commit()
            return render_template('tablas/tabla-filtracion.html', opc="", verificaciones=verificaciones)

        print('consulta: ',consulta_total)   
        cur = mysql.connection.cursor()
        cur.execute(" "+consulta_total)
        verificaciones = cur.fetchall()
        mysql.connection.commit()
        return render_template('tablas/tabla-filtracion.html', opc="", verificaciones=verificaciones)






@app.route('/AñadirFiltroDesbloqueo', methods=["POST", "GET"])
def AñadirFiltroDesbloqueo():
    if request.method == "POST":
        filtro = request.form['filtro']
        # cur = mysql.connection.cursor()
        # cur.execute("select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.PO like %s and v.IdEstado = 6",[po+'%'])
        # verificaciones = cur.fetchall()
        # mysql.connection.commit()
        # primero es tener la consulta base
        data = json.loads(filtro)
        headers = []
        verificaciones = []
        if data:
            consultaBase = 'select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado inner join tb_desbloqueos as des on v.Id_Verificacion = des.IdVerificacion where'
            for clave in data.keys():
                headers.append(clave)
            consulta = ''
            contador = 0

            for value in data.values():
                if headers[contador] == "Fecha":
                    fechas = value
                    fechas = fechas.split('a')
                    print(fechas[0])
                    consulta += 'date(v.'+headers[contador]+') BETWEEN "' + \
                        fechas[0]+'" AND "'+''+fechas[1]+'" AND '
                elif headers[contador] == "IdProveedor":
                    a = 1
                    cur = mysql.connection.cursor()
                    cur.execute(
                        'SELECT Id_Proveedor from tb_proveedor where NombreProveedor = %s', [value])
                    idprov = cur.fetchone()
                    cur = mysql.connection.cursor()

                    print("aqui se muestra el proveedor:")
                    print(idprov)

                    if idprov:
                        cur = mysql.connection.cursor()
                        cur.execute(
                            'SELECT Id_Proveedor from tb_proveedor where NombreProveedor = %s', [value])
                        idprov = cur.fetchone()

                        consulta += 'v.'+headers[contador] + \
                            ' = '+str(idprov[0]) + ' AND '
                    else:
                        # llamar el id de oddo
                        idOddo = conexion.buscarIdProveedor(value)
                        # INSERTAMOS EL PROVEEDOR DE ODDO EN LA BASE
                        cur = mysql.connection.cursor()
                        cur.execute(
                            "INSERT INTO tb_proveedor (NombreProveedor,IdOddo,IdEstado) VALUES (%s,%s,1)", (value, idOddo))
                        proveedornuevo = cur.fetchone()
                        # LLAMAMOS AL PROVEEDOR DE NOMBRE TAL
                        cur = mysql.connection.cursor()
                        cur.execute(
                            'SELECT Id_Proveedor from tb_proveedor where NombreProveedor = %s', [value])
                        idprov = cur.fetchone()

                        consulta += 'v.'+headers[contador] + \
                            ' = '+str(idprov[0]) + ' AND '

                    # consulta += 'date(v.'+headers[contador]+') BETWEEN "'+fechas[0]+'" AND "'+''+fechas[1]+'" AND '
                else:
                    
                    
                    consulta += 'v.'+headers[contador]+' = '+value+' AND '
                contador += 1
                # consultaBase += ' AND '+data
            consulta_total = consultaBase+' '+consulta[:-4] 
            print(consulta_total)
        else:
            
            
            cur = mysql.connection.cursor()
            cur.execute("select IdVerificacion from tb_desbloqueos")
            numerosBoletas = cur.fetchall()
            verificaciones = []
            for boleta in numerosBoletas:
                # LLAMAMOS LAS VERIFICACIONES DE CADA NUMERO DE BOLETA
                print('boleta: ',boleta[0])
                cur = mysql.connection.cursor()
                cur.execute(
                        "select v.Id_Verificacion,v.NoBoleta,v.Fecha,pc.NombrePuntoCompra, p.NombreProveedor,digi.NombreUsuario as digitador,veri.NombreUsuario as verificador from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_usuarios as digi on v.IdDigitador = digi.Id_Usuario inner join tb_usuarios as veri on v.IdVerificador = veri.Id_Usuario Where v.Id_Verificacion = %s", [boleta[0]])
                verifi1 = cur.fetchall()
                mysql.connection.commit()
                verificaciones.append(verifi1)
            print("Vacio")
            print(verificaciones)

            # consulta_total = 'select v.Id_Verificacion,v.NoBoleta,v.Fecha,pc.NombrePuntoCompra, p.NombreProveedor,digi.NombreUsuario as digitador,veri.NombreUsuario as verificador from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_usuarios as digi on v.IdDigitador = digi.Id_Usuario inner join tb_usuarios as veri on v.IdVerificador = veri.Id_Usuario Where v.Id_Verificacion = %s group by v.PO'
            # print('consulta: ',consulta_total)   
            # cur = mysql.connection.cursor()
            # cur.execute('select v.Id_Verificacion,v.Fecha,v.PO,v.NoBoleta,pc.NombrePuntoCompra,e.NombreEstado, p.NombreProveedor from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_estado as e on v.IdEstado = e.Id_Estado where v.IdDigitador = %s group by v.PO',[session['userId']])
            # verificaciones = cur.fetchall()
            # mysql.connection.commit()
            return render_template('tablas/tabla-filtro-desbloqueo.html', opc="", verificaciones=verificaciones)

        print('consulta: ',consulta_total)   
        cur = mysql.connection.cursor()
        cur.execute(" "+consulta_total)
        verificaciones = cur.fetchall()
        mysql.connection.commit()
        print(verificaciones)
        return render_template('tablas/tabla-filtro-desbloqueo.html', opc="", verificaciones=verificaciones)





# GENERAR EL REPORTE DE VERIFICACIONES


@app.route('/reporteVerifiGeneral',  methods=["POST", "GET"])
def reporteVerifiGeneral():
    if request.method == "POST":
        mi_array = request.form['valor']
        # AQUI SE NECEWSITA SABER SI LAS POS SON VAIDAS
        ids = json.loads(mi_array)
        retorno = conexion.GenerarExcel_1(session['pass'], ids, session['uid'])

        return jsonify({'url': ''+retorno})
        # usuarios = db1.execute('select u.Id_Usuario,u.Nombres,u.Apellidos,u.TelefonoFijo,u.Celular,u.Direccion,cred.Usuario,rol.NombreRol from Usuarios as u inner join Credenciales as cred ON u.IdCredenciales = cred.Id_Credenciales inner join Roles as rol ON cred.Rol = rol.Id_Rol inner join estado as est ON u.IdEstado = est.Id_Estado Where u.IdEstado = 1')

        # df_1 = pd.DataFrame((tuple(t) for t in usuarios),
        # columns=('Date ', 'name', 'username', 'description', '','','',''))

        # output = BytesIO()
        # writer = pd.ExcelWriter(output, engine='xlsxwriter')

        # #taken from the original question
        # df_1.to_excel(writer, startrow = -1,startcol=-1, merge_cells = False, sheet_name = "Backup")
        # workbook = writer.book
        # worksheet = writer.sheets["Backup"]
        # caption = 'Table with user defined column headers.'
        # formato = workbook.add_format()
        # formato.set_align('center')
        # # Set the columns widths.
        # worksheet.set_column('A:H', 20)

        # #haciendo la lista
        # row = 2
        # for i in usuarios:
        #     worksheet.write('A'+ str(row), i['Id_Usuario'],formato)
        #     worksheet.write('B'+ str(row), i['Nombres'],formato)
        #     worksheet.write('C'+ str(row), i['Apellidos'],formato)
        #     worksheet.write('D'+ str(row), i['TelefonoFijo'],formato)
        #     worksheet.write('E'+ str(row), i['Celular'],formato)
        #     worksheet.write('F'+ str(row), i['Direccion'],formato)
        #     worksheet.write('G'+ str(row), i['Usuario'],formato)
        #     worksheet.write('H'+ str(row), i['NombreRol'],formato)
        #     worksheet.write('I'+ str(row), None)
        #     row += 1

        #     tam = len(usuarios)
        #     tam = "A1:H" + str(tam)
        #     worksheet.add_table(tam, {
        #                             'columns': [{'header': 'Id Usuario'},
        #                                         {'header': 'Nombres'},
        #                                         {'header': 'Apellidos'},
        #                                         {'header': 'Telefono Fijo'},
        #                                         {'header': 'Celular'},
        #                                         {'header': 'Dirección'},
        #                                         {'header': 'Usuario'},
        #                                         {'header': 'Rol'}
        #     ]})

        #     #the writer has done its job
        #     writer.close()

        #     #go back to the beginning of the stream
        #     output.seek(0)

        # return send_file(output, download_name="Usuarios.xlsx", as_attachment=True)

@app.route('/reporteDesbloqueo',  methods=["POST", "GET"])
def reporteDesbloqueo():
    if request.method == "POST":
        mi_array = request.form['valor']
        # AQUI SE NECEWSITA SABER SI LAS POS SON VAIDAS
        ids = json.loads(mi_array)
        

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"

        # Añade el texto en la celda A1
        ws['A1'] = 'COMPAÑÍA RECICLADORA DE NICARAGUA - REPORTE DE VERIFICACIONES DESBLOQUEADAS'

        # Obtiene la celda A1
        celda = ws['A1']


        # Crea un objeto Font y ajusta la propiedad bold a True para establecer el texto en negritas
        fuente = Font(bold=True)
        celda.font = fuente

        # Crea un objeto Alignment y ajusta la propiedad vertical a 'top'
        alineacion = Alignment(horizontal='center', vertical='center', wrap_text=True)
        celda.alignment = alineacion

        # Asignamos bordes a la celda
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        celda.border = border

        # Une las tres celdas para crear una celda combinada
        ws.merge_cells('A1:F1')


        # Inserta la imagen en la celda combinada B2:D4
        img = Image('static/img/logo.png')
        img.width = 50
        img.height = 50
        ws.add_image(img, 'A1')



        # Añade los nombres y apellidos a la hoja de trabajo en dos columnas separadas
        ws['A2'] = 'ID'
        ws['B2'] = 'NO BOLETA'
        ws['C2'] = 'PUNTO COMPRA'
        ws['D2'] = 'FECHA/HORA'
        ws['E2'] = 'DIGITADOR'
        ws['F2'] = 'VERIFICADOR'

        ws.column_dimensions['A'].width = 11
        ws.column_dimensions['B'].width = 14
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 40

        ws.row_dimensions[2].height = 15
        ws.row_dimensions[1].height = 44

        columnas = ['A','B','C','D','E','F']
        #APLICAMOS EL FORMATO A LAS CELDAS DEL FOR ENCABEZADOS

        for col in columnas:
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            ws[col+'2'].border = border
            ws[col+'2'].alignment = alineacion
            ws[col+'2'].font = fuente
            


        filaCont = 3
        contador = 0
        # Añade los datos a la hoja de trabajo
        for id in ids:
            print('id del for: ',id)
            # Sumar las cantidades de cada línea de pedido de compra
            cur = mysql.connection.cursor()
            cur.execute(
                    'select v.Id_Verificacion,v.NoBoleta,v.Fecha,pc.NombrePuntoCompra, p.NombreProveedor,digi.NombreUsuario as digitador,veri.NombreUsuario as verificador from tb_verificacion as v inner join tb_proveedor as p ON v.IdProveedor = p.Id_Proveedor inner join tb_puntocompra as pc ON v.IdPuntoCompra = pc.Id_PuntoCompra inner join tb_usuarios as digi on v.IdDigitador = digi.Id_Usuario inner join tb_usuarios as veri on v.IdVerificador = veri.Id_Usuario Where v.Id_Verificacion = %s', [id])
            potemp = cur.fetchall()
            for row_num, fila in enumerate(potemp, filaCont):
                # Une las tres celdas para crear una celda combinada
                #ws.merge_cells('A'+str(row_num)+':C7')
                print(fila)
                ws.cell(row=row_num, column=1, value=fila[0])
                ws.cell(row=row_num, column=2, value=fila[1])
                ws.cell(row=row_num, column=3, value=fila[3])
                ws.cell(row=row_num, column=4, value=fila[2])
                ws.cell(row=row_num, column=5, value=fila[5])
                ws.cell(row=row_num, column=6, value=fila[6])
            filaCont += 1
            contador += 1




        # Guarda el archivo de Excel
        # Obtener la fecha actual
        fecha_hora_actual = datetime.now()
        fecha_hora_actual_formateada = fecha_hora_actual.strftime("%d-%m-%y_%H%M%S")
        

        
        # output = BytesIO()
        # wb.save(output)
        # output.seek(0)
        nombre = "static/Reportes/Reporte_"+str(fecha_hora_actual_formateada)+".xlsx"
        wb.save(nombre)
        return jsonify({'url': ''+nombre})






        # usuarios = db1.execute('select u.Id_Usuario,u.Nombres,u.Apellidos,u.TelefonoFijo,u.Celular,u.Direccion,cred.Usuario,rol.NombreRol from Usuarios as u inner join Credenciales as cred ON u.IdCredenciales = cred.Id_Credenciales inner join Roles as rol ON cred.Rol = rol.Id_Rol inner join estado as est ON u.IdEstado = est.Id_Estado Where u.IdEstado = 1')

        # df_1 = pd.DataFrame((tuple(t) for t in usuarios),
        # columns=('Date ', 'name', 'username', 'description', '','','',''))

        # output = BytesIO()
        # writer = pd.ExcelWriter(output, engine='xlsxwriter')

        # #taken from the original question2
        # df_1.to_excel(writer, startrow = -1,startcol=-1, merge_cells = False, sheet_name = "Backup")
        # workbook = writer.book
        # worksheet = writer.sheets["Backup"]
        # caption = 'Table with user defined column headers.'
        # formato = workbook.add_format()
        # formato.set_align('center')
        # # Set the columns widths.
        # worksheet.set_column('A:H', 20)

        # #haciendo la lista
        # row = 2
        # for i in usuarios:
        #     worksheet.write('A'+ str(row), i['Id_Usuario'],formato)
        #     worksheet.write('B'+ str(row), i['Nombres'],formato)
        #     worksheet.write('C'+ str(row), i['Apellidos'],formato)
        #     worksheet.write('D'+ str(row), i['TelefonoFijo'],formato)
        #     worksheet.write('E'+ str(row), i['Celular'],formato)
        #     worksheet.write('F'+ str(row), i['Direccion'],formato)
        #     worksheet.write('G'+ str(row), i['Usuario'],formato)
        #     worksheet.write('H'+ str(row), i['NombreRol'],formato)
        #     worksheet.write('I'+ str(row), None)
        #     row += 1

        #     tam = len(usuarios)
        #     tam = "A1:H" + str(tam)
        #     worksheet.add_table(tam, {
        #                             'columns': [{'header': 'Id Usuario'},
        #                                         {'header': 'Nombres'},
        #                                         {'header': 'Apellidos'},
        #                                         {'header': 'Telefono Fijo'},
        #                                         {'header': 'Celular'},
        #                                         {'header': 'Dirección'},
        #                                         {'header': 'Usuario'},
        #                                         {'header': 'Rol'}
        #     ]})

        #     #the writer has done its job
        #     writer.close()

        #     #go back to the beginning of the stream
        #     output.seek(0)

        # return send_file(output, download_name="Usuarios.xlsx", as_attachment=True)


@app.route('/reporteValidador',  methods=["POST", "GET"])
def reporteValidador():
    if request.method == "POST":
        mi_array = request.form['valor']
        # AQUI SE NECEWSITA SABER SI LAS POS SON VAIDAS
        ids = json.loads(mi_array)
        print("MI ARRAY: ",mi_array)
        print("json con ids: ",ids)
        #MANDAR A LLAMAR LOS PESOS BASCULAS Y TODOS LAS VARIACIONES
        array_datos = []
        valores_extras = []
        contador = 0
        for po in ids:
            cur = mysql.connection.cursor()
            cur.execute("select Id_Verificacion from tb_verificacion Where PO = %s group by PO",[po])
            idtemp = cur.fetchone()


            cur = mysql.connection.cursor()
            cur.execute("select Id_Verificacion from tb_verificacion Where PO = %s ORDER by Id_Verificacion DESC limit 1",[po])
            idtemp1 = cur.fetchone()
            print("ID TEMP: ",idtemp1)
            #MANDAMOS A TRAER LOS VALORES BASCULA DE ESA PO
           # print(idtemp[0])
           # TRAEMOS LAS FECHAS DE LAS CREACION Y CIERRE
            cur = mysql.connection.cursor()
            cur.execute("select v.Fecha,con.FechaConciliacion,p.NombrePuntoCompra from tb_verificacion as v inner join tb_conciliacion as con on v.Id_Verificacion = con.IdVerificacion inner join tb_puntocompra as p on v.IdPuntoCompra = p.Id_PuntoCompra Where v.Id_Verificacion = %s",(idtemp1))
            valores_extras1 = cur.fetchone()
            valores_extras.append(valores_extras1)
           
           #============================================
            cur = mysql.connection.cursor()
            cur.execute("select TipoMaterial,PesoBascula,Variacion1,Variacion2 from tb_validacion Where IdVerificacion = %s",(idtemp))
            valores = cur.fetchall()
            array_datos.append(valores)
            contador +=1
        print('arreglo con los datos: ',array_datos)
        print("VALORES EXTRAS",valores_extras)
        retorno = conexion.GenerarExcel_3(session['pass'], ids, session['uid'],array_datos,valores_extras)


        return jsonify({'url': ''+retorno})
       
#TRAEMOS EL VALOR NUEVO PARA EL CODIGO
@app.route('/traerNumero',  methods=["POST", "GET"])
def traerNumero():
    if request.method == "POST":
        id = request.form['data']
        # NECESITAMOS LA LISTA DE VERIFICADORES, DIGITADORES
        # CONSULTA PARA LOS VERIFICADORES
        if session['punto'] == 'CASETA: Recepciones':
            # cur = mysql.connection.cursor()
            # cur.execute("select ROUND(COUNT(NoBoleta)/2,0) from tb_verificacion Where NoBoleta like 'CS%'")
            # contador = cur.fetchone()

            #TRAER EL ULTIMO VALOR DEL NUM BOLETA
            cur = mysql.connection.cursor()
            cur.execute("SELECT NoBoleta FROM `tb_verificacion` WHERE NoBoleta like 'CS%' group by NoBoleta order by Id_Verificacion DESC limit 1")
            contador = cur.fetchone()
            # MANDAMOS A SACAR EL NUMERO DEL ULTIMO NUMERO DE BOLETA
            caracteres_a_eliminar = "CS"
            if contador:
                cont = contador[0].replace(caracteres_a_eliminar, "")
            else:
                cont = 0

            resultado = int(cont)+1
            #resultado = int(contador[0])+1
            print('entrastes aqui')
            return 'CS'+str(resultado)
        elif session['punto'] == 'GRANEL/PLANTA: Receipts':
            # cur = mysql.connection.cursor()
            # cur.execute("select ROUND(COUNT(NoBoleta)/2,0) from tb_verificacion Where NoBoleta like 'GR%'")
            # contador = cur.fetchone()
            # resultado = contador[0]+1

            #TRAER EL ULTIMO VALOR DEL NUM BOLETA
            cur = mysql.connection.cursor()
            cur.execute("SELECT NoBoleta FROM tb_verificacion WHERE NoBoleta like 'GR%' group by NoBoleta order by Id_Verificacion DESC limit 1")
            contador = cur.fetchone()
            # MANDAMOS A SACAR EL NUMERO DEL ULTIMO NUMERO DE BOLETA
            caracteres_a_eliminar = "GR"
            if contador:
                cont = contador[0].replace(caracteres_a_eliminar, "")
            else:
                cont = 0

            resultado = int(cont)+1
            print(resultado)
            return 'GR'+str(resultado)
        elif session['punto'] == 'MOVIL: Recepciones':
            # cur = mysql.connection.cursor()
            # cur.execute("select ROUND(COUNT(NoBoleta)/2,0) from tb_verificacion Where NoBoleta like 'MV%'")
            # contador = cur.fetchone()
            # resultado = contador[0]+1

            #TRAER EL ULTIMO VALOR DEL NUM BOLETA
            cur = mysql.connection.cursor()
            cur.execute("SELECT NoBoleta FROM tb_verificacion WHERE NoBoleta like 'MV%' group by NoBoleta order by Id_Verificacion DESC limit 1")
            contador = cur.fetchone()
            # MANDAMOS A SACAR EL NUMERO DEL ULTIMO NUMERO DE BOLETA
            caracteres_a_eliminar = "MV"
            if contador:
                cont = contador[0].replace(caracteres_a_eliminar, "")
            else:
                cont = 0

            resultado = int(cont)+1

            return 'MV'+str(resultado)
       
        
        # CONSULTA PARA LOS DIGITADOR
    


# MIS VERIFICACIONES
@app.route('/misVerificaciones')
def misVerificaciones():
    print("mis ver")
    # NECESITAMOS LA LISTA DE VERIFICADORES, DIGITADORES
    # CONSULTA PARA LOS VERIFICADORES
    cur = mysql.connection.cursor()
    cur.execute("select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 2")
    verificador = cur.fetchall()
    # CONSULTA PARA LOS DIGITADOR
    cur = mysql.connection.cursor()
    cur.execute("select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 1")
    digitador = cur.fetchall()
    return render_template('misverificaciones.html', verificadores=verificador, digitadores=digitador)

# REPORTERIA
@app.route('/reporteria')
def reporteria():
    
    return render_template('reporteria.html')



# MIS VERIFICACIONES


@app.route('/validar')
def validar():
    print("mis ver")
    # NECESITAMOS LA LISTA DE VERIFICADORES, DIGITADORES
    # CONSULTA PARA LOS VERIFICADORES
    cur = mysql.connection.cursor()
    cur.execute("select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 2")
    verificador = cur.fetchall()
    # CONSULTA PARA LOS DIGITADOR
    cur = mysql.connection.cursor()
    cur.execute("select * from tb_usuarios Where IdEstado = 1 AND IdCargo = 1")
    digitador = cur.fetchall()
    return render_template('misverificaciones.html', verificadores=verificador, digitadores=digitador)

# DESLOGUEO


@app.route('/deslog')
def deslog():
    session.clear()
    return redirect(url_for('Index'))


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=3134, debug=True)
