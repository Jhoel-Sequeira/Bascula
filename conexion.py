from io import BytesIO
import xmlrpc.client
from flask import send_file

#IMPORTS PARA GENERAR EL EXCEL
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment,Font,Border,Side
import datetime


#=======================================================
# PRIMERA PRUEBA
# CONEXION A LA BASE DE DATOS
# TENEMOS LOS DATOS DE LA API Y EL USUARIO CON SU CONTRASEÑA
url = 'https://recicladora.odoo.com/'
db = 'fdelanuez-itc-recicladora-master-668849'
username = 'soporte@crn.com.ni'
password = 'CRN!2023@bdserver'
#HACEMOS EL LINK DE LA CONEXION CON LA API DE ODO FORMATEANDOLO
info = xmlrpc.client.ServerProxy('https://recicladora.odoo.com/xmlrpc/common')
info.version()
uid = info.authenticate(db, username, password,{})

# PRUEBAS PARA INSERCION EN UNA TABLA
#PRUEBAS DE PERMISOS DE CADA USUARIO
#models = xmlrpc.client.ServerProxy('{}/xmlrpc/object'.format(url))
models = xmlrpc.client.ServerProxy('https://recicladora.odoo.com/xmlrpc/object')








# #BASE DE DATOS DE PRUEBA
# url = 'https://recicladora-250523-8393110.dev.odoo.com/'
# db = 'recicladora-250523-8393110'

# username = 'soporte@crn.com.ni'
# password = '123crn123'
# info = xmlrpc.client.ServerProxy(
#         'https://recicladora-250523-8393110.dev.odoo.com/xmlrpc/common')

# uid =  info.authenticate(db, username, password, {})
# models= xmlrpc.client.ServerProxy(
#         'https://recicladora-250523-8393110.dev.odoo.com/xmlrpc/object')

def conectar(user,contra):
    global username 
    username = ''+user
    global password 
    password = ''+contra
    # HACEMOS EL LINK DE LA CONEXION CON LA API DE ODO FORMATEANDOLO
    info1 = xmlrpc.client.ServerProxy(
        'https://recicladora.odoo.com/xmlrpc/common')
    # BASE DE DATOS DE PRUEBA
    # info1 = xmlrpc.client.ServerProxy(
    #     'https://recicladora-250523-8393110.dev.odoo.com/xmlrpc/common')
    info1.version()
    global info 
    info = info1
    uid1 = info1.authenticate(db, username, password, {})
    global uid 
    uid = uid1
    # PRUEBAS PARA INSERCION EN UNA TABLA
    # PRUEBAS DE PERMISOS DE CADA USUARIO
    # models = xmlrpc.client.ServerProxy('{}/xmlrpc/object'.format(url))
    models1 = xmlrpc.client.ServerProxy(
        'https://recicladora.odoo.com/xmlrpc/object')
    # BASE DE DATOS DE PRUEBA
    # models1 = xmlrpc.client.ServerProxy(
    #     'https://recicladora-250523-8393110.dev.odoo.com/xmlrpc/object')
    global models 
    models = models1
    # permisos = models.execute_kw(db, uid, password, 'res.partner', 'check_access_rights', [
    #                              'write'], {})  # esta seria como la consulta que mostraria
    # CREACION DE REGISTRO DENTRO DE ODDO MEDIANTE LA API
    # LAS CONSULTAS SE EJECclsUTAN SEGUN EL COMANDO
    # -WRITE, READ, CREATE, SEARCH
    # id = models.execute_kw(db, uid, password, 'res.partner', 'create', [{'name': "Prueba2"},{'vat': "123"}]) #ESTA ES LA CONSULTA OARA CREAR LOS REGISTROS
    # datos = models1.execute_kw(db, uid, password, 'res.partner', 'search', [
    #                         [['name', '=', 'Prueba']]])  # ESTA ES PARA BUSCAR POR NOMBRE

    # args = [[['1', '=', 1]]]
    # ids = models.execute_kw(db, uid, password, 'res.partner', 'search', args)
    # models.execute_kw(db, uid, password, 'res.partner', 'check_access_rights', ['read'], {'raise_exception': False})
    #print(datos)

# print(id)
# traer el cargo del empleado que se logueo


def obtenerUid(user,contra):
    
    return info.authenticate(db, user, contra, {})


def Autenticar(user, contra,uid):
    cargo = models.execute_kw(db, uid, contra, 'res.users', 'search_read', 
                              [[['login', '=', ''+user]]], {'fields': ['x_studio_field_xql4c','almacen']})
    # mandamos a llamar el cargo del usuario logueado
    print(cargo)
    return cargo

def conectarTemp(user, contra,prov):
    # primero autenticamos como un usuario
    uidtemp = info.authenticate(db, user, contra, {})
    #BUSCAMOS EL PROVEEDOR
    proveedores = models.execute_kw(db, uidtemp, contra, 'res.partner', 'search_read', [
                                        [['supplier', '=', True], ['name', 'ilike', ''+prov+'%']]], {'fields': ['id', 'name'], 'limit': 5})
    models.execute_kw(db, uid, password, "res.users", "write", [[uidtemp], {'session_ids': [(1, uidtemp, {'active': False})]}])
    return proveedores

def conectarTempCuadrilla(user, contra,prov):
    # primero autenticamos como un usuario
    uidtemp = info.authenticate(db, user, contra, {})
    #BUSCAMOS EL PROVEEDOR
    proveedores = models.execute_kw(db, uidtemp, contra, 'res.partner', 'search_read', [
                                        [['supplier', '=', True],['category_id','=',25], ['name', 'ilike', ''+prov+'%']]], {'fields': ['id', 'name'], 'limit': 5})
    models.execute_kw(db, uid, password, "res.users", "write", [[uidtemp], {'session_ids': [(1, uidtemp, {'active': False})]}])
    return proveedores




def buscarIdProveedor(prov):
    id = models.execute_kw(db, uid, password, 'res.partner', 'search_read', [
                           [['supplier', '=', True], ['name', '=', ''+prov]]], {'fields': ['id']})
    return id[0]['id']

def buscarIdCuadrilla(prov):
    id = models.execute_kw(db, uid, password, 'res.partner', 'search_read', [
                           [['supplier', '=', True],['category_id','=',25], ['name', '=', ''+prov]]], {'fields': ['id']})
    return id[0]['id']


def TraerUsuario(id,uid1,password1):
    print("adentro")
    print(uid1)
    print(password1)
    print(id)
    info = models.execute_kw(db, uid1, password1, 'res.users', 'search_read', [
                             [['id', '=', ''+id]]], {'fields': ['id', 'name',]})
    return info


def buscarProveedor(prov,cargo):
    if cargo != 2:
        proveedores = models.execute_kw(db, uid, password, 'res.partner', 'search_read', [
                                        [['supplier', '=', True], ['name', 'ilike', ''+prov+'%']]], {'fields': ['id', 'name'], 'limit': 5})

        return proveedores
    else:
        proveedores = conectarTemp('soporte@crn.com.ni','CRN!2023@bdserver',prov) #ESTA LINEA ES PARA EL DE PRODUCCION
        #proveedores = conectarTemp('doris.fonseca@crn.com.ni','123',prov) #ESTA LINEA ES PARA EL DE PRUEBA
        return proveedores
    
def buscarCuadrilla(cuad,cargo):
    if cargo != 2:
        cuadrilla = models.execute_kw(db, uid, password, 'res.partner', 'search_read', [
                                        [['supplier', '=', True],['category_id','=',25], ['name', 'ilike', ''+cuad+'%']]], {'fields': ['id', 'name'], 'limit': 5})

        return cuadrilla
    else:
        cuadrilla = conectarTempCuadrilla('soporte@crn.com.ni','CRN!2023@bdserver',cuad) #ESTA LINEA ES PARA EL DE PRODUCCION
        #proveedores = conectarTempCuadrilla('doris.fonseca@crn.com.ni','123',prov) #ESTA LINEA ES PARA EL DE PRUEBA
        return cuadrilla


def CrearOrdenCompra(proveedorId,puntoCompra,NoBoleta,rechazo,jumbo,devolucion,liquido,rechazoPet,primera,segunda,uid1,contra1,jefe,destare):
    
    if jefe:
        print(jumbo)
        pOrder = models.execute_kw(db, uid1, contra1, 'purchase.order',
                                'create', [{'picking_type_id': puntoCompra,
                                            'supplier': True,
                                            'partner_id': proveedorId,
                                            'x_studio_field_WLD1C':NoBoleta,
                                            'x_studio_rechazo_1': rechazo,
                                            'x_studio_field_8Fq79': jefe,
                                            'x_studio_jumbo': jumbo,
                                            'x_studio_destare_lb':destare,
                                            'x_studio_rechazo_pet': devolucion,
                                            'x_studio_lquido': liquido,
                                            'x_studio_material_de_primera': primera,
                                            'x_studio_material_de_segunda': segunda}])
        
        
        return pOrder
    else:
        print(jumbo)
        pOrder = models.execute_kw(db, uid1, contra1, 'purchase.order',
                                'create', [{'picking_type_id': puntoCompra,
                                            'supplier': True,
                                            'partner_id': proveedorId,
                                            'x_studio_field_WLD1C':NoBoleta,
                                            'x_studio_rechazo_1': rechazo,
                                            'x_studio_jumbo': jumbo,
                                            'x_studio_destare_lb':destare,
                                            'x_studio_rechazo_pet': devolucion,
                                            'x_studio_lquido': liquido,
                                            'x_studio_material_de_primera': primera,
                                            'x_studio_material_de_segunda': segunda}])
        
        
        return pOrder
def CrearAlbaran(pOrder,uid1,contra1):
    models.execute_kw(db, uid1, contra1, 'purchase.order', 'button_confirm', [[pOrder]])


def IngresarMaterialOrdenCompra(material,monto,pOrder,uid1,contra1):
    print("material:",material)
    id = models.execute_kw(db, uid1, contra1, 'product.product', 'search_read', [[['name', '=',material],['purchase_ok','=',True]]],{'fields':['pricelist_id']}) 
    print("id vacio")
    print(id)
    if id:
        idMaterial = id[0]['id']
    else:
        
        id1 = models.execute_kw(db, uid1, contra1, 'product.product', 'search_read', [[['name', '=',material+' '],['purchase_ok','=',True]]],{'fields':['pricelist_id']}) 
        print(id)
        if id1:

            idMaterial = id1[0]['id']
        else:
            id2 = models.execute_kw(db, uid1, contra1, 'product.product', 'search_read', [[['product_tmpl_id','=',([material+'  '])],['pricelist_id','=',9]]],{'fields':['price'],'limit':1}) #AQUI AGARRAMOS EL PRICE LIST PUBLICO
            idMaterial = id2[0]['id']
        


    precio = models.execute_kw(db, uid1, contra1, 'product.pricelist.item', 'search_read', [[['product_tmpl_id','=',([material])],['pricelist_id','=',9]]],{'fields':['price'],'limit':1}) #AQUI AGARRAMOS EL PRICE LIST PUBLICO

    if precio:
        precioUnidad = precio[0]['price']
        Id = precio[0]['id']
        print(precio)
    else:
        precio1 = models.execute_kw(db, uid1, contra1, 'product.pricelist.item', 'search_read', [[['product_tmpl_id','=',([material+' '])],['pricelist_id','=',9]]],{'fields':['price'],'limit':1}) #AQUI AGARRAMOS EL PRICE LIST PUBLICO
        if precio1:

            precioUnidad = precio1[0]['price']
            Id = precio1[0]['id']
            print("id dentro del if: ",precio1)
        else:
            precio2 = models.execute_kw(db, uid1, contra1, 'product.pricelist.item', 'search_read', [[['product_tmpl_id','=',([material+'  '])],['pricelist_id','=',9]]],{'fields':['price'],'limit':1}) #AQUI AGARRAMOS EL PRICE LIST PUBLICO
            precioUnidad = precio2[0]['price']
            Id = precio2[0]['id']
        
        


    #FORMATEAR EL STRING
    print(float(precioUnidad[0:10]))
    fecha_plan = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    line_data = {
        'order_id':pOrder,
        'name':material,
        'date_planned':fecha_plan,
        'product_uom':12, #valor de libras es constante
        'price_unit':float(precioUnidad[0:10]),
        'product_qty':monto, #aqui van las libras totales
        'company_id':1,
        'state':'done',
        'product_id':id[0]['id']
    }
    
    
    datos = models.execute_kw(db,uid1,contra1,'purchase.order.line','create',[line_data])
    print("bloqueamos aqui")
    

def traerPo(idOrden):
   
    po = models.execute_kw(db, uid, password, 'purchase.order', 'search_read', [[['id', '=',''+str(idOrden)]]],{'fields':['name']})
    return po

def GenerarExcel_1(contra,ids,uid1):
    # #GENERANDO UN EXCEL CON LA INFORMACION DE ODDO
    # Crea un libro de trabajo y una hoja de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    # Añade el texto en la celda A1
    ws['A1'] = 'COMPAÑÍA RECICLADORA DE NICARAGUA - REPORTE DE GRANEL PLANTA'

    # Obtiene la celda A1
    celda = ws['A1']


    # Crea un objeto Font y ajusta la propiedad bold a True para establecer el texto en negritas
    fuente = Font(bold=True)
    celda.font = fuente

    # Crea un objeto Alignment y ajusta la propiedad vertical a 'top'
    alineacion = Alignment(horizontal='center', vertical='center', wrap_text=True)
    celda.alignment = alineacion

    # Asignamos bordes a la celda
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    celda.border = border

    # Une las tres celdas para crear una celda combinada
    ws.merge_cells('A1:O1')


    # Inserta la imagen en la celda combinada B2:D4
    img = Image('static/img/logo.png')
    img.width = 50
    img.height = 50
    ws.add_image(img, 'A1')



    # Añade los nombres y apellidos a la hoja de trabajo en dos columnas separadas
    ws['A2'] = '#PO'
    ws['B2'] = 'PROVEEDOR'
    ws['C2'] = '#BOLETA'
    ws['D2'] = 'PESO PLANTA'
    ws['E2'] = 'RECHAZO'
    ws['F2'] = 'JUMBO'
    ws['G2'] = 'LIQUIDO'
    ws['H2'] = 'DEVOLUCION PET'
    ws['I2'] = 'DESTARE LBS'
    ws['J2'] = 'PESO TOTAL'
    ws['K2'] = 'PESO BÁSCULA'
    ws['L2'] = 'DIFERENCIA'
    ws['M2'] = '%'
    ws['N2'] = 'MATERIAL'
    ws['O2'] = 'TURNO'

    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 11
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 11
    ws.column_dimensions['F'].width = 11
    ws.column_dimensions['G'].width = 11
    ws.column_dimensions['H'].width = 19
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 22
    ws.column_dimensions['K'].width = 14
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 11
    ws.column_dimensions['N'].width = 11
    ws.column_dimensions['O'].width = 11

    ws.row_dimensions[2].height = 15
    ws.row_dimensions[1].height = 44

    columnas = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O']
    #APLICAMOS EL FORMATO A LAS CELDAS DEL FOR ENCABEZADOS

    for col in columnas:
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws[col+'2'].border = border
        ws[col+'2'].alignment = alineacion
        ws[col+'2'].font = fuente
        


    filaCont = 3
    print(ids)
    # Añade los datos a la hoja de trabajo
    for id in ids:
        potemp = models.execute_kw(db, uid1, contra, 'purchase.order', 'search_read', [[['name', '=',''+str(id)]]])
        # Recuperar las líneas de pedido de compra asociadas a la orden de compra
        print(potemp[0]['id'])
        order_lines = models.execute_kw(db, uid1, contra, 'purchase.order.line', 'search_read', [[('order_id', '=', potemp[0]['id'])]], {'fields': ['product_qty']})
        
        # Sumar las cantidades de cada línea de pedido de compra
        total_cantidad = sum(line['product_qty'] for line in order_lines)
        for row_num, fila in enumerate(potemp, filaCont):
            # Une las tres celdas para crear una celda combinada
            #ws.merge_cells('A'+str(row_num)+':C7')
            ws.cell(row=row_num, column=1, value=fila['name'])
            ws.cell(row=row_num, column=2, value=fila['partner_id'][1])
            ws.cell(row=row_num, column=3, value=fila['x_studio_field_WLD1C'])
            ws.cell(row=row_num, column=4, value=total_cantidad)
            ws.cell(row=row_num, column=5, value=fila['x_studio_rechazo_1'])
            ws.cell(row=row_num, column=6, value=fila['x_studio_jumbo'])
            ws.cell(row=row_num, column=7, value=fila['x_studio_lquido'])

            ws.cell(row=row_num, column=8, value='0')
            ws.cell(row=row_num, column=9, value='0')
            ws.cell(row=row_num, column=10, value='=SUMA(D'+str(row_num)+',E'+str(row_num)+',F'+str(row_num)+',G'+str(row_num)+')')
            ws.cell(row=row_num, column=11, value='0')
            ws.cell(row=row_num, column=12, value='0')
            ws.cell(row=row_num, column=13, value='0')
            ws.cell(row=row_num, column=14, value='0')
            ws.cell(row=row_num, column=15, value='0')
        filaCont += 1




    # Guarda el archivo de Excel
    # Obtener la fecha actual
    fecha_hora_actual = datetime.datetime.now()
    fecha_hora_actual_formateada = fecha_hora_actual.strftime("%d-%m-%y_%H%M%S")
    

    
    # output = BytesIO()
    # wb.save(output)
    # output.seek(0)
    nombre = "static/Reportes/Reporte_"+str(fecha_hora_actual_formateada)+".xlsx"
    wb.save(nombre)
    return nombre
    #return output

def GenerarExcel_3(contra,ids,uid1,datos,valores_extras):
    # #GENERANDO UN EXCEL CON LA INFORMACION DE ODDO
    # Crea un libro de trabajo y una hoja de trabajo
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    # Añade el texto en la celda A1
    ws['A1'] = 'COMPAÑÍA RECICLADORA DE NICARAGUA - REPORTE DEL VALIDADOR'

    # Obtiene la celda A1
    celda = ws['A1']


    # Crea un objeto Font y ajusta la propiedad bold a True para establecer el texto en negritas
    fuente = Font(bold=True)
    celda.font = fuente

    # Crea un objeto Alignment y ajusta la propiedad vertical a 'top'
    alineacion = Alignment(horizontal='center', vertical='center', wrap_text=True)
    celda.alignment = alineacion

    # Asignamos bordes a la celda
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    celda.border = border

    # Une las tres celdas para crear una celda combinada
    ws.merge_cells('A1:R1')


    # Inserta la imagen en la celda combinada B2:D4
    img = Image('static/img/logo.png')
    img.width = 50
    img.height = 50
    ws.add_image(img, 'A1')



    # Añade los nombres y apellidos a la hoja de trabajo en dos columnas separadas
    ws['A2'] = '#PO'
    ws['B2'] = 'PROVEEDOR'
    ws['C2'] = '#BOLETA'
    ws['D2'] = 'PESO PLANTA'
    ws['E2'] = 'RECHAZO'
    ws['F2'] = 'JUMBO'
    ws['G2'] = 'LIQUIDO'
    ws['H2'] = 'DEVOLUCIÓN PET'
    ws['I2'] = 'DESTARE LBS'
    ws['J2'] = 'PESO TOTAL'
    ws['K2'] = 'PESO BÁSCULA'
    ws['L2'] = 'DIFERENCIA'
    ws['M2'] = '%'
    ws['N2'] = 'MATERIAL'
    ws['O2'] = 'TURNO'
    ws['P2'] = 'FECHA CREACION'
    ws['Q2'] = 'FECHA CIERRE'
    ws['R2'] = 'PUNTO COMPRA'

    ws.column_dimensions['A'].width = 11
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 11
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 11
    ws.column_dimensions['F'].width = 11
    ws.column_dimensions['G'].width = 11
    ws.column_dimensions['H'].width = 19
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 22
    ws.column_dimensions['K'].width = 14
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 11
    ws.column_dimensions['N'].width = 11
    ws.column_dimensions['O'].width = 11
    ws.column_dimensions['P'].width = 11
    ws.column_dimensions['Q'].width = 11
    ws.column_dimensions['R'].width = 11

    ws.row_dimensions[2].height = 15
    ws.row_dimensions[1].height = 44

    columnas = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R']
    #APLICAMOS EL FORMATO A LAS CELDAS DEL FOR ENCABEZADOS
    materiales = 0
    for col in columnas:
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        ws[col+'2'].border = border
        ws[col+'2'].alignment = alineacion
        ws[col+'2'].font = fuente
        


    filaCont = 3
    contador = 0
    # Añade los datos a la hoja de trabajo
    print(valores_extras)
    for id in ids:
        potemp = models.execute_kw(db, uid1, contra, 'purchase.order', 'search_read', [[['name', '=',''+str(id)]]])
        # Recuperar las líneas de pedido de compra asociadas a la orden de compra
        print(potemp[0]['id'])
        #print("valores: ",datos[1][0])
        
        order_lines = models.execute_kw(db, uid1, contra, 'purchase.order.line', 'search_read', [[('order_id', '=', potemp[0]['id'])]], {'fields': ['product_qty']})
        print('order_Line: ',order_lines)
        # Sumar las cantidades de cada línea de pedido de compra
        total_cantidad = sum(line['product_qty'] for line in order_lines)
        
        for row_num, fila in enumerate(potemp, filaCont):
            # Une las tres celdas para crear una celda combinada
            #ws.merge_cells('A'+str(row_num)+':C7')
            ws.cell(row=row_num, column=1, value=fila['name'])
            ws.cell(row=row_num, column=2, value=fila['partner_id'][1])
            ws.cell(row=row_num, column=3, value=fila['x_studio_field_WLD1C'])
            ws.cell(row=row_num, column=4, value=total_cantidad)
            ws.cell(row=row_num, column=5, value=fila['x_studio_rechazo_1'])
            ws.cell(row=row_num, column=6, value=fila['x_studio_jumbo'])
            ws.cell(row=row_num, column=7, value=fila['x_studio_lquido'])

            ws.cell(row=row_num, column=8, value='0')
            ws.cell(row=row_num, column=9, value='0')
            ws.cell(row=row_num, column=10, value='=SUMA(D'+str(row_num)+',E'+str(row_num)+',F'+str(row_num)+',G'+str(row_num)+')')
            ws.cell(row=row_num, column=11, value=datos[contador][0][1])
            ws.cell(row=row_num, column=12, value=datos[contador][0][3])
            ws.cell(row=row_num, column=13, value=datos[contador][0][2])
            ws.cell(row=row_num, column=14, value=datos[contador][0][0])
            ws.cell(row=row_num, column=15, value='0')
            try:
                if valores_extras[materiales][0] is None:
                    ws.cell(row=row_num, column=16, value="-")
                else:
                    ws.cell(row=row_num, column=16, value=valores_extras[materiales][0])
            except:
                 ws.cell(row=row_num, column=16, value="-")

            try:
                if valores_extras[materiales][1] is None:
                    ws.cell(row=row_num, column=17, value="-")
                else:
                    ws.cell(row=row_num, column=17, value=valores_extras[materiales][1])
            except:
                ws.cell(row=row_num, column=17, value="-")
            
            try:
                if valores_extras[materiales][2] is None:
                    ws.cell(row=row_num, column=18, value="-")
                else:
                    ws.cell(row=row_num, column=18, value=valores_extras[materiales][2])
            except:
                ws.cell(row=row_num, column=18, value="-")
        filaCont += 1
        materiales+=1
        contador += 1
       




    # Guarda el archivo de Excel
    # Obtener la fecha actual
    fecha_hora_actual = datetime.datetime.now()
    fecha_hora_actual_formateada = fecha_hora_actual.strftime("%d-%m-%y_%H%M%S")
    

    
    # output = BytesIO()
    # wb.save(output)
    # output.seek(0)
    nombre = "static/Reportes/Reporte_"+str(fecha_hora_actual_formateada)+".xlsx"
    wb.save(nombre)
    return nombre